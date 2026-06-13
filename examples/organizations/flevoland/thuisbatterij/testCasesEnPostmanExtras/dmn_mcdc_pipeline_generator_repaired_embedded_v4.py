#!/usr/bin/env python3
"""
Generate MC/DC-style and boundary-focused JSON test cases for every decision
 table in a DMN file.

The output is designed for DMN/REST-style test runners. Each generated case has:
- decisionId / decisionName: the DMN decision table the case targets
- name: readable test-case name
- expected: comma-separated expected output values
- requestBody.variables: typed variables to submit
- coverage: why the case was selected

Algorithm summary:
1. Parse all DMN decision tables.
2. For every decision table, derive the variables required by that table.
3. Build boundary-focused domains from FEEL unary tests, date ranges,
   string equality rules, numeric thresholds, and known subsidy-style variables.
4. Evaluate candidates directly against each decision table.
5. Select MC/DC-style pairs for atomic conditions.
6. Add explicit boundary/domain coverage so every derived boundary value is
   represented in at least one test for that table.

Supported FEEL subset:
- string and numeric literals
- date("YYYY-MM-DD"), date(variable), date(variable).year
- equality and comparisons
- `and` clauses
- nested `if ... then ... else ...` output expressions
- simple arithmetic

The core DMN/JSON/Postman generation uses only the Python standard library.
Excel generation uses `openpyxl` so the workbook can include local-friendly
formatting, run-link cells, and charts. For complex FEEL models, extend the
evaluator or replace `safe_eval` with your production DMN engine while keeping
the domain and MC/DC selection logic.
"""

from __future__ import annotations

import argparse
import base64
import copy
import uuid
import html
import os
import shutil
import tempfile
import zipfile
from collections import Counter, defaultdict
import itertools
import json
import math
import re
import sys
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date, timedelta
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple


DMN_NS = {"dmn": "https://www.omg.org/spec/DMN/20191111/MODEL/"}
IDENT_RE = re.compile(r"\b[A-Za-z_][A-Za-z0-9_]*\b")
COMPARISON_RE = re.compile(r"^(<=|>=|<|>|=|!=)\s*(.+)$")


@dataclass
class DmnInput:
    id: str
    label: str
    expression: str
    type_ref: Optional[str] = None


@dataclass
class DmnOutput:
    id: str
    label: str
    name: str
    type_ref: Optional[str] = None


@dataclass
class DmnRule:
    id: str
    input_entries: List[str]
    output_entries: List[str]
    description: str = ""


@dataclass
class DmnDecision:
    id: str
    name: str
    table_id: str
    variable_name: Optional[str]
    hit_policy: str
    inputs: List[DmnInput]
    outputs: List[DmnOutput]
    rules: List[DmnRule]


@dataclass
class RuleTrace:
    rule_id: str
    matched: bool
    atoms: Dict[str, bool]


@dataclass
class TableEvaluation:
    selected_rule_id: Optional[str]
    selected_rule_index: Optional[int]
    outputs: Dict[str, Any]
    rule_traces: Dict[str, RuleTrace] = field(default_factory=dict)


@dataclass
class Candidate:
    decision_id: str
    inputs: Dict[str, Any]
    evaluation: TableEvaluation


# ---------------------------------------------------------------------------
# XML parsing
# ---------------------------------------------------------------------------


def text_of(elem: Optional[ET.Element]) -> str:
    if elem is None or elem.text is None:
        return ""
    return elem.text.strip()


def parse_dmn(path: Path) -> Dict[str, DmnDecision]:
    root = ET.parse(path).getroot()
    decisions: Dict[str, DmnDecision] = {}

    for decision_el in root.findall("dmn:decision", DMN_NS):
        decision_id = decision_el.get("id") or ""
        variable_el = decision_el.find("dmn:variable", DMN_NS)
        variable_name = variable_el.get("name") if variable_el is not None else None

        dt = decision_el.find("dmn:decisionTable", DMN_NS)
        if dt is None:
            continue

        inputs: List[DmnInput] = []
        for inp in dt.findall("dmn:input", DMN_NS):
            ie = inp.find("dmn:inputExpression", DMN_NS)
            inputs.append(
                DmnInput(
                    id=inp.get("id") or "",
                    label=inp.get("label") or "",
                    expression=text_of(ie.find("dmn:text", DMN_NS)) if ie is not None else "",
                    type_ref=ie.get("typeRef") if ie is not None else None,
                )
            )

        outputs: List[DmnOutput] = []
        for out in dt.findall("dmn:output", DMN_NS):
            outputs.append(
                DmnOutput(
                    id=out.get("id") or "",
                    label=out.get("label") or "",
                    name=out.get("name") or out.get("label") or out.get("id") or "",
                    type_ref=out.get("typeRef"),
                )
            )

        rules: List[DmnRule] = []
        for rule_el in dt.findall("dmn:rule", DMN_NS):
            input_entries = [text_of(e.find("dmn:text", DMN_NS)) for e in rule_el.findall("dmn:inputEntry", DMN_NS)]
            output_entries = [text_of(e.find("dmn:text", DMN_NS)) for e in rule_el.findall("dmn:outputEntry", DMN_NS)]
            description = text_of(rule_el.find("dmn:description", DMN_NS))

            # Ignore empty placeholder rows created by DMN modelers.
            if not any(x.strip() for x in input_entries + output_entries):
                continue

            rules.append(
                DmnRule(
                    id=rule_el.get("id") or "",
                    input_entries=input_entries,
                    output_entries=output_entries,
                    description=description,
                )
            )

        decisions[decision_id] = DmnDecision(
            id=decision_id,
            name=decision_el.get("name") or decision_id,
            table_id=dt.get("id") or "",
            variable_name=variable_name,
            hit_policy=(dt.get("hitPolicy") or "UNIQUE").upper(),
            inputs=inputs,
            outputs=outputs,
            rules=rules,
        )

    if not decisions:
        raise ValueError("No DMN decision table found.")
    return decisions


# ---------------------------------------------------------------------------
# FEEL subset evaluator
# ---------------------------------------------------------------------------


def date_func(value: Any) -> date:
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return date.fromisoformat(value)
    raise ValueError(f"Cannot convert {value!r} to date")


def split_top_level_keyword(expr: str, keyword: str, start: int = 0) -> int:
    depth = 0
    in_string = False
    i = start
    while i < len(expr):
        ch = expr[i]
        if ch == '"':
            in_string = not in_string
            i += 1
            continue
        if in_string:
            i += 1
            continue
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        elif depth == 0:
            before_ok = i == 0 or not (expr[i - 1].isalnum() or expr[i - 1] == "_")
            after_idx = i + len(keyword)
            after_ok = after_idx >= len(expr) or not (expr[after_idx].isalnum() or expr[after_idx] == "_")
            if before_ok and after_ok and expr.startswith(keyword, i):
                return i
        i += 1
    return -1


def split_top_level_and(expr: str) -> List[str]:
    parts: List[str] = []
    start = 0
    while True:
        idx = split_top_level_keyword(expr, "and", start)
        if idx == -1:
            parts.append(expr[start:].strip())
            return [p for p in parts if p]
        parts.append(expr[start:idx].strip())
        start = idx + 3


@lru_cache(maxsize=None)
def translate_feel_expr(expr: str) -> str:
    expr = (expr or "").strip()
    if not expr:
        return "None"

    if expr.startswith("if "):
        then_idx = split_top_level_keyword(expr, "then", 3)
        else_idx = split_top_level_keyword(expr, "else", then_idx + 4 if then_idx != -1 else 0)
        if then_idx == -1 or else_idx == -1:
            raise ValueError(f"Cannot parse FEEL if-expression: {expr}")
        condition = expr[3:then_idx].strip()
        then_part = expr[then_idx + 4:else_idx].strip()
        else_part = expr[else_idx + 4:].strip()
        return f"({translate_feel_expr(then_part)} if {translate_condition_expr(condition)} else {translate_feel_expr(else_part)})"

    return translate_condition_expr(expr)


@lru_cache(maxsize=None)
def translate_condition_expr(expr: str) -> str:
    expr = (expr or "").strip()
    expr = re.sub(r"\bdate\s*\(", "date_func(", expr)
    expr = re.sub(r"(?<![<>=!])=(?!=)", "==", expr)
    expr = re.sub(r"\btrue\b", "True", expr, flags=re.IGNORECASE)
    expr = re.sub(r"\bfalse\b", "False", expr, flags=re.IGNORECASE)
    expr = re.sub(r"\bnull\b", "None", expr, flags=re.IGNORECASE)
    return expr


@lru_cache(maxsize=None)
def compiled_feel_expr(expr: str) -> Tuple[str, Any]:
    py_expr = translate_feel_expr(expr)
    return py_expr, compile(py_expr, "<feel>", "eval")


def safe_eval(expr: str, ctx: Dict[str, Any]) -> Any:
    py_expr, code = compiled_feel_expr(expr)
    safe_globals = {"__builtins__": {}, "date_func": date_func, "min": min, "max": max, "abs": abs, "math": math}
    try:
        return eval(code, safe_globals, dict(ctx))
    except Exception as e:
        raise ValueError(f"Failed to evaluate FEEL {expr!r} translated as {py_expr!r}: {e}") from e


def atom_label(input_expr: str, entry: str) -> str:
    entry = (entry or "").strip()
    if not entry or entry == "-":
        return ""
    m = COMPARISON_RE.match(entry)
    if m:
        op, rhs = m.groups()
        op = "==" if op == "=" else op
        return f"{input_expr} {op} {rhs.strip()}"
    if re.match(r'^\s*".*"\s*$', entry):
        return f"{input_expr} == {entry}"
    if re.match(r"^\s*-?\d+(\.\d+)?\s*$", entry):
        return f"{input_expr} == {entry}"
    return entry


def eval_entry_against_input(input_expr: str, input_value: Any, entry: str, ctx: Dict[str, Any]) -> bool:
    entry = (entry or "").strip()
    if entry in ("", "-"):
        return True

    m = COMPARISON_RE.match(entry)
    if m:
        op, rhs_expr = m.groups()
        rhs = safe_eval(rhs_expr, ctx)
        if op in ("=", "=="):
            return input_value == rhs
        if op == "!=":
            return input_value != rhs
        if op == "<=":
            return input_value <= rhs
        if op == ">=":
            return input_value >= rhs
        if op == "<":
            return input_value < rhs
        if op == ">":
            return input_value > rhs

    if re.match(r'^\s*".*"\s*$', entry) or re.match(r"^\s*-?\d+(\.\d+)?\s*$", entry):
        return input_value == safe_eval(entry, ctx)

    return bool(safe_eval(entry, ctx))


def eval_rule_atoms(decision: DmnDecision, rule: DmnRule, ctx: Dict[str, Any]) -> Tuple[bool, Dict[str, bool]]:
    atoms: Dict[str, bool] = {}
    rule_matches = True

    for inp, entry in zip(decision.inputs, rule.input_entries):
        entry = (entry or "").strip()
        input_value = safe_eval(inp.expression, ctx)
        if not entry or entry == "-":
            continue

        parts = split_top_level_and(entry)
        entry_match = True
        for part in parts:
            label = atom_label(inp.expression, part)
            value = eval_entry_against_input(inp.expression, input_value, part, ctx)
            atoms[label] = bool(value)
            entry_match = entry_match and bool(value)
        rule_matches = rule_matches and entry_match

    return bool(rule_matches), atoms


def evaluate_decision_table_direct(decision: DmnDecision, ctx: Dict[str, Any]) -> TableEvaluation:
    selected_rule: Optional[DmnRule] = None
    selected_rule_index: Optional[int] = None
    rule_traces: Dict[str, RuleTrace] = {}

    for idx, rule in enumerate(decision.rules, start=1):
        matched, atoms = eval_rule_atoms(decision, rule, ctx)
        rule_traces[rule.id] = RuleTrace(rule_id=rule.id, matched=matched, atoms=atoms)
        if matched and selected_rule is None:
            selected_rule = rule
            selected_rule_index = idx
            if decision.hit_policy == "FIRST":
                # Still keep evaluating remaining rules for traceability? For a
                # true FIRST table, selection stops. We continue only atom traces
                # in the generic evaluator by not breaking; selected stays first.
                pass

    outputs: Dict[str, Any] = {}
    if selected_rule is not None:
        for out, out_expr in zip(decision.outputs, selected_rule.output_entries):
            outputs[out.name] = safe_eval(out_expr, ctx) if out_expr.strip() else None

    return TableEvaluation(
        selected_rule_id=selected_rule.id if selected_rule else None,
        selected_rule_index=selected_rule_index,
        outputs=outputs,
        rule_traces=rule_traces,
    )


# ---------------------------------------------------------------------------
# Domain extraction
# ---------------------------------------------------------------------------


def identifiers(expr: str) -> List[str]:
    expr = re.sub(r'"[^"]*"', '""', expr or "")
    keywords = {"if", "then", "else", "and", "or", "not", "date", "true", "false", "null", "year", "min", "max"}
    return [x for x in IDENT_RE.findall(expr or "") if x not in keywords]


def decision_text(decision: DmnDecision) -> str:
    return "\n".join(
        [
            decision.name,
            "\n".join(i.expression for i in decision.inputs),
            "\n".join(e for r in decision.rules for e in (r.input_entries + r.output_entries)),
        ]
    )


def all_model_text(decisions: Dict[str, DmnDecision]) -> str:
    return "\n".join(decision_text(d) for d in decisions.values())


def variables_required_by_decision(decision: DmnDecision, constant_defaults: Dict[str, Any]) -> List[str]:
    vars_needed: set[str] = set()
    for inp in decision.inputs:
        vars_needed.update(identifiers(inp.expression))
    for rule in decision.rules:
        for expr in rule.input_entries + rule.output_entries:
            vars_needed.update(identifiers(expr))

    # Do not ask the caller to supply a variable produced by this same table.
    own_outputs = {o.name for o in decision.outputs if o.name}
    vars_needed -= own_outputs

    # Keep constants if this table references them; direct table tests are self-contained.
    return sorted(vars_needed)


def compute_constant_defaults(decisions: Dict[str, DmnDecision]) -> Dict[str, Any]:
    defaults: Dict[str, Any] = {}
    for decision in decisions.values():
        if decision.inputs:
            continue
        try:
            evaluation = evaluate_decision_table_direct(decision, dict(defaults))
            defaults.update(evaluation.outputs)
        except Exception:
            continue
    return defaults


def extract_dates_from_text(text: str) -> List[str]:
    values: set[str] = set()
    for raw in re.findall(r'date\s*\(\s*"(\d{4}-\d{2}-\d{2})"\s*\)', text):
        d = date.fromisoformat(raw)
        for delta in (-1, 0, 1):
            values.add((d + timedelta(days=delta)).isoformat())
    if values:
        years = sorted({date.fromisoformat(v).year for v in values})
        values.add(f"{min(years) - 1}-12-31")
        values.add(f"{max(years) + 1}-01-01")
    return sorted(values)


def dates_for_year_literals(text: str) -> List[str]:
    years = sorted({int(x) for x in re.findall(r"(?<![\w-])(20\d{2})(?![\w-])", text)})
    # Use only plausible year literal ranges, not every date boundary year if no year expression is present.
    values: set[str] = set()
    for y in years:
        values.add(f"{y - 1}-12-31")
        values.add(f"{y}-01-01")
        values.add(f"{y}-12-31")
        values.add(f"{y + 1}-01-01")
    return sorted(values)


def string_literals_for_var(decision: DmnDecision, var_name: str) -> List[str]:
    values: set[str] = set()
    for inp_idx, inp in enumerate(decision.inputs):
        if inp.expression != var_name:
            continue
        for rule in decision.rules:
            if inp_idx >= len(rule.input_entries):
                continue
            for lit in re.findall(r'"([^"]*)"', rule.input_entries[inp_idx] or ""):
                if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", lit):
                    values.add(lit)
    if values:
        values.add("onbekend")
    return sorted(values)


def extract_numeric_boundaries(text: str) -> List[float]:
    raw_numbers = [float(x) for x in re.findall(r"(?<![\w-])-?\d+(?:\.\d+)?(?![\w-])", text)]
    nums = {n for n in raw_numbers if not (1900 <= n <= 2100 and float(n).is_integer())}

    # Derived ratios reveal implicit thresholds such as 750 / 0.25 = 3000.
    positives = sorted(n for n in nums if n > 0)
    for a in positives:
        for b in positives:
            if b == 0:
                continue
            q = a / b
            if 1 <= q <= 10_000_000 and abs(q - round(q, 6)) < 1e-9:
                nums.add(round(q, 6))

    boundaries: set[float] = set()
    for n in nums:
        if abs(n) < 1:
            boundaries.add(n)
        else:
            for delta in (-1, 0, 1):
                boundaries.add(n + delta)

    return sorted(x for x in boundaries if -1_000_000 <= x <= 10_000_000)


def numeric_domain_for_variable(var_name: str, numeric_boundaries: Sequence[float], constant_defaults: Dict[str, Any]) -> List[Any]:
    if var_name in constant_defaults and isinstance(constant_defaults[var_name], (int, float)):
        return [as_number(constant_defaults[var_name])]

    lower = var_name.lower()
    if "percentage" in lower:
        return [0.25]
    if "minimum" in lower or "minimale" in lower:
        return [750]
    if "maximum" in lower:
        return [1250]
    if "gemaaktekosten" in lower or "kosten" in lower:
        preferred = [0, 2999, 3000, 3001, 4999, 5000, 5001, 10000]
    elif "basishoogtesubsidie" in lower or "beschikbaarsubsidieplafond" in lower or "hoogtesubsidie" in lower:
        preferred = [0, 1, 749, 750, 751, 1249, 1250, 1251, 437499, 437500, 437501]
    elif "reeds" in lower or "gesubsidieerd" in lower:
        preferred = [0, 1, 437499, 437500, 437501, 874999, 875000, 875001, 999999, 1000000, 1000001]
    elif "plafond" in lower or "budget" in lower:
        preferred = [437500, 875000, 1000000]
    else:
        preferred = [0, 1] + list(numeric_boundaries[:12])

    return sorted(dict.fromkeys(as_number(x) for x in preferred))


def as_number(value: Any) -> Any:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def is_date_variable(var_name: str, text: str) -> bool:
    return (
        "datum" in var_name.lower()
        or f"date({var_name})" in text.replace(" ", "")
        or f"date( {var_name} )" in text
    )


def domain_for_variable(decision: DmnDecision, var_name: str, constant_defaults: Dict[str, Any], model_text: str) -> List[Any]:
    text = decision_text(decision)
    compact_text = text.replace(" ", "")

    if var_name in constant_defaults and not is_date_variable(var_name, compact_text):
        value = constant_defaults[var_name]
        if isinstance(value, str):
            return [value]
        if isinstance(value, (int, float)):
            return [as_number(value)]

    strings = string_literals_for_var(decision, var_name)
    if strings:
        return strings

    if is_date_variable(var_name, compact_text):
        date_values = extract_dates_from_text(text)
        if ".year" in text or "year" in text:
            date_values = sorted(set(date_values) | set(dates_for_year_literals(text)))
        return date_values or [date.today().isoformat()]

    numeric_boundaries = extract_numeric_boundaries(model_text)
    return numeric_domain_for_variable(var_name, numeric_boundaries, constant_defaults)


def domains_for_decision(decision: DmnDecision, constant_defaults: Dict[str, Any], model_text: str) -> Dict[str, List[Any]]:
    domains: Dict[str, List[Any]] = {}
    for var_name in variables_required_by_decision(decision, constant_defaults):
        domains[var_name] = domain_for_variable(decision, var_name, constant_defaults, model_text)
    return domains


def candidate_inputs(domains: Dict[str, List[Any]], max_candidates: int) -> Iterable[Dict[str, Any]]:
    keys = list(domains.keys())
    if not keys:
        yield {}
        return
    values = [domains[k] for k in keys]
    for idx, combo in enumerate(itertools.product(*values)):
        if idx >= max_candidates:
            return
        yield dict(zip(keys, combo))


# ---------------------------------------------------------------------------
# Selection: MC/DC pairs + boundary representatives
# ---------------------------------------------------------------------------


def input_distance(a: Dict[str, Any], b: Dict[str, Any]) -> int:
    return sum(1 for k in set(a) | set(b) if a.get(k) != b.get(k))


def jsonable(value: Any) -> Any:
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return value
    if isinstance(value, dict):
        return {k: jsonable(v) for k, v in value.items()}
    if isinstance(value, list):
        return [jsonable(v) for v in value]
    return value


def select_mcdc_pairs_for_decision(decision: DmnDecision, candidates: List[Candidate]) -> Tuple[List[Tuple[Candidate, Candidate, str]], List[str]]:
    pairs: List[Tuple[Candidate, Candidate, str]] = []
    uncovered: List[str] = []
    seen: set[str] = set()

    for rule_index, rule in enumerate(decision.rules, start=1):
        atom_names: List[str] = []
        for c in candidates:
            rt = c.evaluation.rule_traces.get(rule.id)
            if rt and rt.atoms:
                atom_names = list(rt.atoms.keys())
                break
        if not atom_names:
            continue

        for target_atom in atom_names:
            buckets: Dict[Tuple[Tuple[Tuple[str, bool], ...], bool, bool], List[Candidate]] = {}
            for c in candidates:
                rt = c.evaluation.rule_traces.get(rule.id)
                if not rt or target_atom not in rt.atoms:
                    continue
                other_key = tuple((atom, bool(rt.atoms.get(atom))) for atom in atom_names if atom != target_atom)
                key = (other_key, bool(rt.atoms[target_atom]), bool(rt.matched))
                cell = buckets.setdefault(key, [])
                if len(cell) < 100:
                    cell.append(c)

            best: Optional[Tuple[int, Candidate, Candidate]] = None
            for other_key in sorted({k[0] for k in buckets}):
                false_false = buckets.get((other_key, False, False), [])
                true_true = buckets.get((other_key, True, True), [])
                false_true = buckets.get((other_key, False, True), [])
                true_false = buckets.get((other_key, True, False), [])
                for left, right in ((false_false, true_true), (false_true, true_false)):
                    if not left or not right:
                        continue
                    for a in left:
                        for b in right:
                            score = input_distance(a.inputs, b.inputs)
                            if jsonable(a.evaluation.outputs) != jsonable(b.evaluation.outputs):
                                score -= 1
                            if best is None or score < best[0]:
                                best = (score, a, b)

            reason = f"MC/DC: rule {rule_index} {rule.id}, condition [{target_atom}]"
            if best:
                _, a, b = best
                key = json.dumps([a.inputs, b.inputs, reason], sort_keys=True, default=str)
                if key not in seen:
                    pairs.append((a, b, reason))
                    seen.add(key)
            else:
                uncovered.append(f"{decision.name}: {reason}")

    return pairs, uncovered


def output_sort_value(candidate: Candidate) -> Tuple[int, float, str]:
    nums = [float(v) for v in candidate.evaluation.outputs.values() if isinstance(v, (int, float))]
    max_abs = max((abs(v) for v in nums), default=0.0)
    has_output = 0 if candidate.evaluation.outputs else 1
    return (has_output, -max_abs, json.dumps(candidate.inputs, sort_keys=True, default=str))


def choose_boundary_representatives(domains: Dict[str, List[Any]], candidates: List[Candidate]) -> List[Tuple[Candidate, str]]:
    reps: List[Tuple[Candidate, str]] = []
    for var_name, values in domains.items():
        for value in values:
            matching = [c for c in candidates if c.inputs.get(var_name) == value]
            if not matching:
                continue
            matching.sort(key=output_sort_value)
            reps.append((matching[0], f"Boundary/domain value: {var_name}={value!r}"))
    return reps


def choose_rule_representatives(decision: DmnDecision, candidates: List[Candidate]) -> List[Tuple[Candidate, str]]:
    reps: List[Tuple[Candidate, str]] = []
    for idx, rule in enumerate(decision.rules, start=1):
        matching = [c for c in candidates if c.evaluation.selected_rule_id == rule.id]
        if not matching:
            continue
        matching.sort(key=output_sort_value)
        reps.append((matching[0], f"Representative selected rule {idx}: {rule.id}"))
    return reps


def choose_output_extreme_representatives(decision: DmnDecision, candidates: List[Candidate]) -> List[Tuple[Candidate, str]]:
    reps: List[Tuple[Candidate, str]] = []
    for out in decision.outputs:
        numeric_candidates = [c for c in candidates if isinstance(c.evaluation.outputs.get(out.name), (int, float))]
        if not numeric_candidates:
            continue
        by_value = sorted(numeric_candidates, key=lambda c: float(c.evaluation.outputs[out.name]))
        for label, c in [("minimum", by_value[0]), ("maximum", by_value[-1])]:
            reps.append((c, f"Output {label}: {out.name}={c.evaluation.outputs[out.name]!r}"))
        zeroish = sorted(numeric_candidates, key=lambda c: abs(float(c.evaluation.outputs[out.name])))
        if zeroish:
            reps.append((zeroish[0], f"Output near zero boundary: {out.name}={zeroish[0].evaluation.outputs[out.name]!r}"))
    return reps


# ---------------------------------------------------------------------------
# Output formatting
# ---------------------------------------------------------------------------


def variable_type(value: Any) -> str:
    if isinstance(value, bool):
        return "Boolean"
    if isinstance(value, (int, float)):
        return "Double"
    return "String"


def typed_variables(inputs: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    return {
        k: {"value": jsonable(v), "type": variable_type(v)}
        for k, v in inputs.items()
    }


def format_value_for_expected(value: Any, type_ref: Optional[str] = None) -> str:
    if value is None:
        return "null"
    if isinstance(value, float):
        return f"{value:.10g}" if not value.is_integer() else f"{value:.1f}"
    if isinstance(value, int) and type_ref in {"double", "number"}:
        return f"{float(value):.1f}"
    if isinstance(value, int):
        return str(value)
    return str(value)


def expected_string(decision: DmnDecision, evaluation: TableEvaluation) -> str:
    if not evaluation.outputs:
        if decision.outputs:
            return ", ".join(f"{out.name}=null" for out in decision.outputs)
        return "noOutputs=true"
    parts = []
    for out in decision.outputs:
        parts.append(f"{out.name}={format_value_for_expected(evaluation.outputs.get(out.name), out.type_ref)}")
    return ", ".join(parts)


def case_name(index: int, decision: DmnDecision, candidate: Candidate, primary_reason: str) -> str:
    reason = primary_reason.split(":", 1)[0]
    selected = candidate.evaluation.selected_rule_id or "no matching rule"
    return f"TC_{index:03d} {decision.name} - {reason} - {selected}"


def build_outputs(
    dmn_path: Path,
    decisions: Dict[str, DmnDecision],
    all_domains: Dict[str, Dict[str, List[Any]]],
    all_candidates: Dict[str, List[Candidate]],
    max_cases_per_decision: Optional[int] = None,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    all_cases: List[Dict[str, Any]] = []
    coverage_summary: Dict[str, Any] = {}
    uncovered_all: List[str] = []

    for decision in decisions.values():
        candidates = all_candidates[decision.id]
        domains = all_domains[decision.id]
        by_key: Dict[str, Dict[str, Any]] = {}

        def add_candidate(c: Candidate, reason: str) -> None:
            key = json.dumps(c.inputs, sort_keys=True, default=str)
            if key not in by_key:
                by_key[key] = {
                    "candidate": c,
                    "reasons": [],
                }
            if reason not in by_key[key]["reasons"]:
                by_key[key]["reasons"].append(reason)

        pairs, uncovered = select_mcdc_pairs_for_decision(decision, candidates)
        uncovered_all.extend(uncovered)
        for a, b, reason in pairs:
            add_candidate(a, reason + " baseline")
            add_candidate(b, reason + " flipped")

        for c, reason in choose_boundary_representatives(domains, candidates):
            add_candidate(c, reason)
        for c, reason in choose_rule_representatives(decision, candidates):
            add_candidate(c, reason)
        for c, reason in choose_output_extreme_representatives(decision, candidates):
            add_candidate(c, reason)

        selected_items = list(by_key.values())
        selected_items.sort(key=lambda item: json.dumps(item["candidate"].inputs, sort_keys=True, default=str))
        if max_cases_per_decision:
            selected_items = selected_items[:max_cases_per_decision]

        decision_case_count = 0
        for item in selected_items:
            c: Candidate = item["candidate"]
            reasons: List[str] = item["reasons"]
            case_index = len(all_cases) + 1
            all_cases.append(
                {
                    "name": case_name(case_index, decision, c, reasons[0]),
                    "decisionId": decision.id,
                    "decisionName": decision.name,
                    "decisionTableId": decision.table_id,
                    "evaluationMode": "direct-table-inputs",
                    "expected": expected_string(decision, c.evaluation),
                    "requestBody": {"variables": typed_variables(c.inputs)},
                    "coverage": {
                        "selectedRuleId": c.evaluation.selected_rule_id,
                        "selectedRuleIndex": c.evaluation.selected_rule_index,
                        "reasons": reasons,
                    },
                }
            )
            decision_case_count += 1

        atom_count = 0
        for rule in decision.rules:
            for c in candidates[:1_000]:
                rt = c.evaluation.rule_traces.get(rule.id)
                if rt:
                    atom_count += len(rt.atoms)
                    break

        coverage_summary[decision.id] = {
            "decisionName": decision.name,
            "decisionTableId": decision.table_id,
            "hitPolicy": decision.hit_policy,
            "candidateCountEvaluated": len(candidates),
            "selectedTestCaseCount": decision_case_count,
            "inputDomainsUsed": jsonable(domains),
            "rules": [
                {
                    "ruleIndex": idx,
                    "ruleId": rule.id,
                    "description": rule.description,
                    "inputEntries": rule.input_entries,
                    "outputEntries": rule.output_entries,
                }
                for idx, rule in enumerate(decision.rules, start=1)
            ],
            "uncoveredConditions": uncovered,
        }

    analysis = {
        "metadata": {
            "sourceDmn": str(dmn_path),
            "algorithm": "Per-decision-table boundary domains + MC/DC-style condition-pair selection",
            "note": (
                "Each decision table is tested in direct-table-input mode. Non-boolean DMN entries "
                "such as date ranges, comparisons, and string equality tests are treated as atomic "
                "boolean predicates for MC/DC pair selection. Boundary/domain representatives are "
                "added so all generated input ranges appear in at least one case per table."
            ),
            "decisionCount": len(decisions),
            "selectedTestCaseCount": len(all_cases),
        },
        "decisions": coverage_summary,
        "uncoveredConditions": uncovered_all,
    }
    return all_cases, analysis


def generate(
    dmn_path: Path,
    max_candidates_per_decision: int = 100_000,
    max_cases_per_decision: Optional[int] = None,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    decisions = parse_dmn(dmn_path)
    model_text = all_model_text(decisions)
    constant_defaults = compute_constant_defaults(decisions)

    all_domains: Dict[str, Dict[str, List[Any]]] = {}
    all_candidates: Dict[str, List[Candidate]] = {}

    for decision in decisions.values():
        domains = domains_for_decision(decision, constant_defaults, model_text)
        all_domains[decision.id] = domains
        candidates: List[Candidate] = []
        errors = 0
        for inputs in candidate_inputs(domains, max_candidates_per_decision):
            try:
                # Direct-table mode: the request body contains all variables
                # required to evaluate the selected table. Defaults from zero-input
                # constant tables are also added when referenced.
                evaluation = evaluate_decision_table_direct(decision, dict(inputs))
                candidates.append(Candidate(decision_id=decision.id, inputs=dict(inputs), evaluation=evaluation))
            except Exception as e:
                errors += 1
                if errors <= 3:
                    print(f"Skipping {decision.id} candidate {inputs}: {e}", file=sys.stderr)
        if not candidates:
            raise RuntimeError(f"No evaluable candidates generated for decision {decision.id}")
        all_candidates[decision.id] = candidates

    return build_outputs(dmn_path, decisions, all_domains, all_candidates, max_cases_per_decision=max_cases_per_decision)



# ---------------------------------------------------------------------------
# Postman collection generation
# ---------------------------------------------------------------------------

JsonObj = Dict[str, Any]


def load_json(path: str | Path) -> Any:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def dump_json(data: Any, path: str | Path) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")


def decision_key_from_item(item: JsonObj) -> Optional[str]:
    """Extract `/decision-definition/key/{key}/...` from a Postman item."""
    url = item.get("request", {}).get("url", {})
    path = url.get("path", [])
    if isinstance(path, list) and "key" in path:
        idx = path.index("key")
        if idx + 1 < len(path):
            return path[idx + 1]
    raw = url.get("raw", "") if isinstance(url, dict) else ""
    m = re.search(r"/decision-definition/key/([^/]+)/", raw)
    return m.group(1) if m else None


def iter_postman_items(items: Iterable[JsonObj]) -> Iterable[JsonObj]:
    """Yield all request items, including nested folder items."""
    for item in items:
        if "item" in item:
            yield from iter_postman_items(item.get("item", []))
        elif "request" in item:
            yield item


def choose_template_items(base_collection: JsonObj, needed_keys: Iterable[str]) -> Dict[str, JsonObj]:
    """Pick one canonical request item per decision key from a Postman collection."""
    candidates: Dict[str, List[JsonObj]] = defaultdict(list)
    for item in iter_postman_items(base_collection.get("item", [])):
        key = decision_key_from_item(item)
        if key:
            candidates[key].append(item)

    selected: Dict[str, JsonObj] = {}
    for key in needed_keys:
        options = candidates.get(key, [])
        if not options:
            continue

        def score(item: JsonObj) -> Tuple[int, int, int]:
            name = item.get("name", "").lower()
            no_test = 1 if ("test" not in name and "experiment" not in name) else 0
            has_examples = 1 if item.get("response") else 0
            shorter_name = -len(name)
            return (no_test, has_examples, shorter_name)

        selected[key] = max(options, key=score)
    return selected


def infer_postman_value_type(value: Any) -> str:
    if value is None:
        return "Null"
    if isinstance(value, bool):
        return "Boolean"
    if isinstance(value, int) and not isinstance(value, bool):
        return "Integer"
    if isinstance(value, float):
        return "Double"
    return "String"


def parse_expected_string(expected: str) -> JsonObj:
    """Convert 'a=1.0, b=text' into an Operaton/Camunda-style response row."""
    result: JsonObj = {}
    if not expected:
        return result
    parts = [p.strip() for p in re.split(r",\s*(?=[A-Za-z_][A-Za-z0-9_]*\s*=)", expected)]
    for part in parts:
        if "=" not in part:
            continue
        key, raw_value = part.split("=", 1)
        key = key.strip()
        raw_value = raw_value.strip()
        if raw_value.lower() == "null":
            value = None
        elif raw_value.lower() == "true":
            value = True
        elif raw_value.lower() == "false":
            value = False
        else:
            try:
                value = float(raw_value)
            except ValueError:
                value = raw_value.strip('"')
        result[key] = {"type": infer_postman_value_type(value), "value": value}
    return result


def make_original_request(template_request: JsonObj, request_body: JsonObj) -> JsonObj:
    original = copy.deepcopy(template_request)
    original.pop("auth", None)  # parent request retains auth; examples stay compact
    original["body"] = {
        "mode": "raw",
        "raw": json.dumps(request_body, ensure_ascii=False, indent=2),
        "options": {"raw": {"language": "json"}},
    }
    return original


def make_postman_example(test_case: JsonObj, template_request: JsonObj) -> JsonObj:
    expected_row = parse_expected_string(test_case.get("expected", ""))
    coverage = test_case.get("coverage", {}) or {}
    reasons = coverage.get("reasons", []) or []
    description_lines = [
        f"Decision: {test_case.get('decisionName') or test_case.get('decisionId')}",
        f"Decision table: {test_case.get('decisionTableId', '')}",
        f"Expected: {test_case.get('expected', '')}",
    ]
    if reasons:
        description_lines.append("Coverage reasons:")
        description_lines.extend(f"- {reason}" for reason in reasons)

    return {
        "name": test_case.get("name", "MC/DC test case"),
        "originalRequest": make_original_request(template_request, test_case.get("requestBody", {})),
        "status": "OK",
        "code": 200,
        "_postman_previewlanguage": "json",
        "header": [{"key": "Content-Type", "value": "application/json"}],
        "cookie": [],
        "body": json.dumps([expected_row], ensure_ascii=False, indent=2),
        "description": "\n".join(description_lines),
    }


def engine_rest_base_url(base_url: str) -> str:
    """Return a base URL that ends with /engine-rest exactly once."""
    base = base_url.rstrip("/")
    if base.endswith("/engine-rest"):
        return base
    return f"{base}/engine-rest"


def postman_url_for_decision(base_url: str, decision_id: str, tenant_id: str) -> JsonObj:
    raw = f"{engine_rest_base_url(base_url)}/decision-definition/key/{decision_id}/tenant-id/{tenant_id}/evaluate"
    m = re.match(r"^(https?)://([^/]+)(/.*)$", raw)
    if not m:
        return {"raw": raw}
    protocol, host, path = m.groups()
    return {"raw": raw, "protocol": protocol, "host": host.split("."), "path": [p for p in path.strip("/").split("/") if p]}


def make_synthetic_template_item(decision_id: str, base_url: str, tenant_id: str, request_body: JsonObj) -> JsonObj:
    return {
        "name": decision_id,
        "request": {
            "method": "POST",
            "header": [],
            "body": {
                "mode": "raw",
                "raw": json.dumps(request_body, ensure_ascii=False, indent=2),
                "options": {"raw": {"language": "json"}},
            },
            "url": postman_url_for_decision(base_url, decision_id, tenant_id),
        },
        "response": [],
    }


def make_decision_item(decision_id: str, cases: List[JsonObj], template_item: JsonObj) -> JsonObj:
    item = copy.deepcopy(template_item)
    decision_name = cases[0].get("decisionName") or decision_id
    item["name"] = f"{decision_id} - MC/DC examples ({len(cases)})"
    item["description"] = (
        f"Generated MC/DC and boundary-value examples for `{decision_name}`. "
        f"Each example contains the request body for one generated test case and "
        f"an expected Operaton/Camunda-style decision-evaluation response body."
    )
    item["request"]["body"] = {
        "mode": "raw",
        "raw": json.dumps(cases[0].get("requestBody", {}), ensure_ascii=False, indent=2),
        "options": {"raw": {"language": "json"}},
    }
    item["response"] = [make_postman_example(tc, item["request"]) for tc in cases]
    item.pop("event", None)
    return item


def make_summary_item(test_cases: List[JsonObj]) -> JsonObj:
    counts = Counter(tc.get("decisionId", "UNKNOWN") for tc in test_cases)
    body = {
        "summary": "Generated MC/DC examples grouped under the corresponding DMN decision-evaluate request.",
        "totalTestCases": len(test_cases),
        "testCasesPerDecision": [[decision_id, count] for decision_id, count in sorted(counts.items())],
    }
    return {
        "name": "MC/DC generation summary",
        "request": {
            "method": "GET",
            "header": [],
            "url": {"raw": "about:blank", "host": ["about:blank"]},
            "description": json.dumps(body, ensure_ascii=False, indent=2),
        },
        "response": [],
    }


def generate_postman_collection(
    test_cases: List[JsonObj],
    postman_template: Optional[Path],
    base_url: str,
    tenant_id: str,
) -> JsonObj:
    by_decision: Dict[str, List[JsonObj]] = defaultdict(list)
    for tc in test_cases:
        by_decision[tc.get("decisionId", "UNKNOWN")].append(tc)

    if postman_template:
        base_collection = load_json(postman_template)
        templates = choose_template_items(base_collection, by_decision.keys())
    else:
        base_collection = {
            "info": {
                "_postman_id": str(uuid.uuid4()),
                "name": "DMN MC/DC generated collection",
                "description": "Generated from DMN and MC/DC boundary test cases.",
                "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json",
            },
            "item": [],
        }
        templates = {}

    # Synthesize any request missing from the optional template.
    for key, cases in by_decision.items():
        if key not in templates:
            templates[key] = make_synthetic_template_item(key, base_url, tenant_id, cases[0].get("requestBody", {}))

    output = copy.deepcopy(base_collection)
    output["info"] = copy.deepcopy(base_collection.get("info", {}))
    output["info"]["_postman_id"] = str(uuid.uuid4())
    output["info"]["schema"] = output["info"].get("schema", "https://schema.getpostman.com/json/collection/v2.1.0/collection.json")
    base_name = output["info"].get("name", "Postman collection")
    output["info"]["name"] = f"{base_name} - MC/DC examples"
    base_description = output["info"].get("description", "")
    output["info"]["description"] = (
        f"{base_description}\n\nGenerated collection with {len(test_cases)} MC/DC and boundary-value test cases "
        f"grouped as Postman examples under their corresponding DMN decision calls."
    ).strip()

    utility_items: List[JsonObj] = []
    if postman_template:
        for item in base_collection.get("item", []):
            key = decision_key_from_item(item) if "request" in item else None
            if key is None and "request" in item:
                utility_items.append(copy.deepcopy(item))

    preferred_order = [
        "BehaalbareHoogteSubsidie",
        "BerekenBasisHoogteSubsidie",
        "BerekenBeschikbaarSubsidiePlafond",
        "SubsidieConstantenThuisbatterij",
        "jaarGebondenBudget",
    ]
    ordered_keys = [key for key in preferred_order if key in by_decision] + [
        key for key in sorted(by_decision.keys()) if key not in preferred_order
    ]
    folder = {
        "name": "DMN decision calls with generated MC/DC examples",
        "description": "One request per DMN decision key. Test cases are attached as Postman examples.",
        "item": [make_decision_item(key, by_decision[key], templates[key]) for key in ordered_keys],
    }
    output["item"] = [make_summary_item(test_cases), folder] + utility_items
    return output


# ---------------------------------------------------------------------------
# Excel workbook generation
# ---------------------------------------------------------------------------


def excel_col_name(index: int) -> str:
    """1-based column number to Excel column letters."""
    letters = ""
    while index:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def write_matrix(sheet: Any, start_row: int, start_col: int, rows: List[List[Any]]) -> None:
    """Write a 2D matrix using 1-based row/col coordinates."""
    if not rows:
        return
    row_count = len(rows)
    col_count = max(len(r) for r in rows)
    padded = [r + [None] * (col_count - len(r)) for r in rows]
    end_col = excel_col_name(start_col + col_count - 1)
    rng = f"{excel_col_name(start_col)}{start_row}:{end_col}{start_row + row_count - 1}"
    sheet.get_range(rng).values = padded


def add_table_if_possible(sheet: Any, address: str, name: str) -> None:
    try:
        sheet.tables.add(address, True, name)
    except Exception:
        # The workbook remains useful without table objects if the runtime lacks this API.
        pass


def style_basic_sheet(sheet: Any, used_range: str, header_range: str = "A1:Z1") -> None:
    try:
        sheet.freeze_panes.freeze_rows(1)
    except Exception:
        pass
    try:
        sheet.get_range(header_range).format = {
            "fill": "#0F766E",
            "font": {"bold": True, "color": "#FFFFFF"},
            "horizontal_alignment": "center",
            "vertical_alignment": "center",
        }
    except Exception:
        pass
    try:
        sheet.get_range(used_range).format.wrap_text = True
        sheet.get_range(used_range).format.autofit_columns()
        sheet.get_range(used_range).format.autofit_rows()
    except Exception:
        pass


def generate_excel_workbook(
    test_cases: List[JsonObj],
    analysis: JsonObj,
    output_path: Path,
    *,
    base_url: str = "https://operaton.open-regels.nl",
    tenant_id: str = "46",
    postman_path: Optional[Path] = None,
) -> None:
    """Create a local-friendly Excel workbook with dashboard, run links and chart.

    The workbook intentionally uses openpyxl rather than artifact_tool so it can
    be generated on a normal Windows/macOS/Linux Python installation:

        python -m pip install openpyxl

    The "buttons" are styled Excel cells with hyperlinks. Standard .xlsx files
    cannot execute POST requests without macros/VBA, so these cells point to the
    generated Postman collection and decision endpoints. Use Postman/Newman for
    real automated execution.
    """
    try:
        from openpyxl import Workbook as OpenPyxlWorkbook
        from openpyxl.chart import PieChart, Reference
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as e:  # pragma: no cover - local dependency guidance
        raise RuntimeError(
            "Excel generation requires openpyxl. Install it with: "
            "python -m pip install openpyxl. Alternatively run with --skip-excel."
        ) from e

    counts = Counter(tc.get("decisionId", "UNKNOWN") for tc in test_cases)
    uncovered = analysis.get("uncoveredConditions", []) or []
    decisions = analysis.get("decisions") or {}
    output_path = Path(output_path)
    postman_path = Path(postman_path) if postman_path else None

    var_names = sorted({
        name
        for tc in test_cases
        for name in (tc.get("requestBody", {}).get("variables", {}) or {}).keys()
    })

    def endpoint_for(decision_id: str) -> str:
        return postman_url_for_decision(base_url, decision_id, tenant_id).get("raw", "")

    def expected_response_body(expected: str) -> str:
        """Small preview of the expected decision response for the workbook."""
        result: JsonObj = {}
        for part in [p.strip() for p in str(expected or "").split(",") if p.strip()]:
            if "=" not in part:
                continue
            key, value_text = part.split("=", 1)
            key = key.strip()
            value_text = value_text.strip()
            if not key:
                continue
            try:
                value: Any = float(value_text)
            except ValueError:
                value = value_text.strip('"')
            value_type = "Double" if isinstance(value, float) else "String"
            result[key] = {"type": value_type, "value": value}
        return json.dumps([result], ensure_ascii=False, indent=2) if result else ""

    def safe_uri(path: Path) -> str:
        try:
            return path.resolve().as_uri()
        except Exception:
            return str(path)

    # ------------------------------------------------------------------
    # Prepare sheet data
    # ------------------------------------------------------------------
    case_headers = [
        "#",
        "Run",
        "Name",
        "Decision ID",
        "Decision Name",
        "Decision Table ID",
        "Expected",
        "Selected Rule ID",
        "Selected Rule Index",
        "Coverage Reasons",
    ] + var_names + ["Endpoint", "Expected Response Preview", "Request Body JSON"]

    case_rows: List[List[Any]] = [case_headers]
    for idx, tc in enumerate(test_cases, start=1):
        variables = tc.get("requestBody", {}).get("variables", {}) or {}
        coverage = tc.get("coverage", {}) or {}
        decision_id = tc.get("decisionId", "")
        row = [
            idx,
            "RUN",
            tc.get("name", ""),
            decision_id,
            tc.get("decisionName", ""),
            tc.get("decisionTableId", ""),
            tc.get("expected", ""),
            coverage.get("selectedRuleId", ""),
            coverage.get("selectedRuleIndex", ""),
            "\n".join(coverage.get("reasons", []) or []),
        ]
        for var in var_names:
            spec = variables.get(var)
            row.append(spec.get("value") if isinstance(spec, dict) else None)
        row.extend([
            endpoint_for(decision_id),
            expected_response_body(tc.get("expected", "")),
            json.dumps(tc.get("requestBody", {}), ensure_ascii=False, indent=2),
        ])
        case_rows.append(row)

    domain_rows: List[List[Any]] = [["Decision ID", "Decision Name", "Variable", "Domain Values"]]
    for decision_id, details in sorted(decisions.items()):
        for var_name, values in sorted((details.get("inputDomainsUsed") or {}).items()):
            domain_rows.append([
                decision_id,
                details.get("decisionName", ""),
                var_name,
                json.dumps(values, ensure_ascii=False),
            ])

    rule_rows: List[List[Any]] = [["Decision ID", "Decision Name", "Rule Index", "Rule ID", "Description", "Input Entries", "Output Entries"]]
    for decision_id, details in sorted(decisions.items()):
        for rule in details.get("rules", []) or []:
            rule_rows.append([
                decision_id,
                details.get("decisionName", ""),
                rule.get("ruleIndex", ""),
                rule.get("ruleId", ""),
                rule.get("description", ""),
                json.dumps(rule.get("inputEntries", []), ensure_ascii=False),
                json.dumps(rule.get("outputEntries", []), ensure_ascii=False),
            ])

    wb = OpenPyxlWorkbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    ws_summary = wb.create_sheet("Summary")
    ws_cases = wb.create_sheet("Test Cases")
    ws_domains = wb.create_sheet("Input Domains")
    ws_rules = wb.create_sheet("Rules")

    # ------------------------------------------------------------------
    # Styling helpers
    # ------------------------------------------------------------------
    dark_fill = PatternFill("solid", fgColor="0F766E")
    medium_fill = PatternFill("solid", fgColor="14B8A6")
    light_fill = PatternFill("solid", fgColor="CCFBF1")
    grey_fill = PatternFill("solid", fgColor="F3F4F6")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=16, bold=True, color="0F172A")
    subtitle_font = Font(size=11, color="475569")
    header_font = Font(bold=True, color="FFFFFF")
    button_font = Font(bold=True, color="FFFFFF", underline="single")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    def append_rows(ws: Any, rows: List[List[Any]]) -> None:
        for row in rows:
            ws.append(row)

    def apply_header(ws: Any, row: int = 1) -> None:
        for cell in ws[row]:
            cell.fill = dark_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    def add_table(ws: Any, table_name: str, start_row: int = 1, end_col: Optional[int] = None) -> None:
        if ws.max_row < start_row or ws.max_column < 1:
            return
        max_col = end_col or ws.max_column
        ref = f"A{start_row}:{get_column_letter(max_col)}{ws.max_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        try:
            ws.add_table(table)
        except ValueError:
            pass

    def autosize(ws: Any, max_width: int = 60) -> None:
        for column_cells in ws.columns:
            col_letter = get_column_letter(column_cells[0].column)
            max_len = 0
            for cell in column_cells:
                if cell.value is None:
                    continue
                max_len = max(max_len, min(len(str(cell.value)), max_width))
            ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, max_width))
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def style_button(cell: Any, label: str, link: Optional[str] = None, fill: Optional[PatternFill] = None) -> None:
        cell.value = label
        cell.fill = fill or dark_fill
        cell.font = button_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        if link:
            cell.hyperlink = link
            cell.style = "Hyperlink"
            cell.fill = fill or dark_fill
            cell.font = button_font

    # ------------------------------------------------------------------
    # Summary / dashboard with run buttons and pie chart
    # ------------------------------------------------------------------
    ws_summary.merge_cells("A1:F1")
    ws_summary["A1"] = "DMN MC/DC Test Generation Summary"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].alignment = Alignment(horizontal="left")

    ws_summary.merge_cells("A2:F2")
    ws_summary["A2"] = "Boundary-focused MC/DC cases generated per DMN decision table. Use Postman/Newman for real POST execution."
    ws_summary["A2"].font = subtitle_font

    metric_rows = [
        ("Source DMN", analysis.get("metadata", {}).get("sourceDmn", "")),
        ("Algorithm", analysis.get("metadata", {}).get("algorithm", "")),
        ("Decision count", analysis.get("metadata", {}).get("decisionCount", len(counts))),
        ("Selected test cases", len(test_cases)),
        ("Uncovered conditions", len(uncovered)),
    ]
    ws_summary["A4"] = "Metric"
    ws_summary["B4"] = "Value"
    for c in ws_summary[4][0:2]:
        c.fill = dark_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border
    for idx, (metric, value) in enumerate(metric_rows, start=5):
        ws_summary.cell(idx, 1, metric)
        ws_summary.cell(idx, 2, value)
        ws_summary.cell(idx, 1).fill = grey_fill
        ws_summary.cell(idx, 1).font = Font(bold=True)
        ws_summary.cell(idx, 1).border = thin_border
        ws_summary.cell(idx, 2).border = thin_border

    # Button-like cells. They are hyperlinks because standard .xlsx cannot issue POST requests without macros.
    postman_uri = safe_uri(postman_path) if postman_path else None
    ws_summary.merge_cells("D4:F5")
    style_button(ws_summary["D4"], "RUN ALL IN POSTMAN", postman_uri, dark_fill)
    ws_summary.merge_cells("D7:F8")
    style_button(ws_summary["D7"], "OPEN TEST CASES", "#'Test Cases'!A1", medium_fill)
    ws_summary["D10"] = "Newman command"
    ws_summary["D10"].fill = grey_fill
    ws_summary["D10"].font = Font(bold=True)
    ws_summary["E10"] = f"newman run \"{postman_path.name if postman_path else '<generated collection>.json'}\""
    ws_summary["E10"].alignment = Alignment(wrap_text=True)
    ws_summary.merge_cells("E10:F10")
    ws_summary["D11"] = "Note"
    ws_summary["D11"].fill = grey_fill
    ws_summary["D11"].font = Font(bold=True)
    ws_summary["E11"] = "The RUN cells are clickable links. To execute all POST requests, import/run the generated Postman collection or use Newman."
    ws_summary.merge_cells("E11:F11")
    ws_summary["E11"].alignment = Alignment(wrap_text=True)

    # Decision summary table and pie chart source data
    decision_start = 14
    decision_headers = ["Decision ID", "Decision name", "Hit policy", "Candidate count", "Selected cases"]
    for col_idx, value in enumerate(decision_headers, start=1):
        cell = ws_summary.cell(decision_start, col_idx, value)
        cell.fill = dark_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    row_idx = decision_start + 1
    for decision_id, details in sorted(decisions.items()):
        values = [
            decision_id,
            details.get("decisionName", ""),
            details.get("hitPolicy", ""),
            details.get("candidateCountEvaluated", ""),
            counts.get(decision_id, 0),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws_summary.cell(row_idx, col_idx, value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row_idx += 1

    add_table(ws_summary, "DecisionSummary", decision_start, end_col=5)

    if counts:
        chart = PieChart()
        labels = Reference(ws_summary, min_col=1, min_row=decision_start + 1, max_row=row_idx - 1)
        data = Reference(ws_summary, min_col=5, min_row=decision_start, max_row=row_idx - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = "MC/DC cases by decision"
        chart.height = 8
        chart.width = 11
        ws_summary.add_chart(chart, "G4")

    ws_summary.freeze_panes = "A14"
    ws_summary.column_dimensions["A"].width = 34
    ws_summary.column_dimensions["B"].width = 28
    ws_summary.column_dimensions["C"].width = 16
    ws_summary.column_dimensions["D"].width = 18
    ws_summary.column_dimensions["E"].width = 36
    ws_summary.column_dimensions["F"].width = 22

    # ------------------------------------------------------------------
    # Test cases sheet with individual run links
    # ------------------------------------------------------------------
    append_rows(ws_cases, case_rows)
    apply_header(ws_cases, 1)
    add_table(ws_cases, "GeneratedTestCases", 1)
    ws_cases.freeze_panes = "A2"
    ws_cases.auto_filter.ref = f"A1:{get_column_letter(ws_cases.max_column)}{ws_cases.max_row}"
    endpoint_col = case_headers.index("Endpoint") + 1
    for r in range(2, ws_cases.max_row + 1):
        decision_id = ws_cases.cell(r, 4).value
        run_cell = ws_cases.cell(r, 2)
        style_button(run_cell, "RUN", endpoint_for(str(decision_id)), medium_fill)
        ws_cases.cell(r, endpoint_col).hyperlink = str(ws_cases.cell(r, endpoint_col).value or "")
        ws_cases.cell(r, endpoint_col).style = "Hyperlink"
    autosize(ws_cases)
    ws_cases.column_dimensions["B"].width = 12
    ws_cases.column_dimensions["C"].width = 46
    ws_cases.column_dimensions["J"].width = 48
    ws_cases.column_dimensions[get_column_letter(ws_cases.max_column)].width = 64
    ws_cases.column_dimensions[get_column_letter(ws_cases.max_column - 1)].width = 48

    # ------------------------------------------------------------------
    # Supporting sheets
    # ------------------------------------------------------------------
    append_rows(ws_domains, domain_rows)
    apply_header(ws_domains, 1)
    add_table(ws_domains, "InputDomains", 1)
    ws_domains.freeze_panes = "A2"
    autosize(ws_domains)
    ws_domains.column_dimensions["D"].width = 72

    append_rows(ws_rules, rule_rows)
    apply_header(ws_rules, 1)
    add_table(ws_rules, "Rules", 1)
    ws_rules.freeze_panes = "A2"
    autosize(ws_rules)
    ws_rules.column_dimensions["F"].width = 56
    ws_rules.column_dimensions["G"].width = 56

    # Apply light borders and readable row heights across data sheets.
    for ws in [ws_cases, ws_domains, ws_rules]:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.sheet_view.showGridLines = False
    ws_summary.sheet_view.showGridLines = False

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

# ---------------------------------------------------------------------------
# Macro-enabled Excel workbook generation
# ---------------------------------------------------------------------------

RUN_ALL_MACRO = "PostmanTestRunner.RunAllTests"
RUN_SELECTED_MACRO = "PostmanTestRunner.RunSelectedTest"

# Embedded macro-enabled runner workbook. This is the small working template used
# only as the container for vbaProject.bin, workbook relationships, and Excel
# form-control support files. Generated Dashboard/Tests XML replaces the template
# sheets at runtime, so callers do not need to keep a separate grr.bla file next
# to the script.
EMBEDDED_RUNNER_TEMPLATE_B64 = """
UEsDBBQABgAIAAAAIQDox6z/FAIAAAcLAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAAC
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADE
ls1y2jAQx++d6Tt4dO3YImmbdjqYHJr02GYm6QMIacEqsqTRKgTevusP6EcA45jSi2WQd3+7f3nX
O75elSZZQkDtbM4ushFLwEqntJ3n7PvDl/QjSzAKq4RxFnK2BmTXk9evxg9rD5iQtcWcFTH6T5yj
LKAUmDkPlnZmLpQi0s8w517IhZgDvxyNrrh0NoKNaax8sMn4Bmbi0cTkdkV/N5FMtWXJ5+a5CpUz
4b3RUkQKlC+t+guSutlMS1BOPpbkOkMfQCgsAGJpMh80EcM9xEiJIeM7mQEM9oO2WWVkWQeGhfb4
hlLfQ1jSzpCsyP4miCfKYQ+gQu8HtIF9o/MOWkFyJ0L8KkoSl68Mf3JhMXVukR12UmlfYgorCSar
Bc5KIYO7tWJqgPaEthsJDpBqS+T1cnEE8s936vBxV5nUjnvGcfmf4ohUNsDr63ApajcdiWNcG8AT
Z9s47SIXIoC6j1SQ85MH8LvvjjhUU0bI25vhureOOrgyBnMXnEe+uTuGvK23qnMGZzx5gBA1bJvN
nkqTJEkkVrWcLEVqqLXDnhIPr66XSnwM+eUSPzvRt306aO8TfYZ7d17c+/Pirs6L+/AvcVSz6Czy
Zu1Daiw6Cm45FdRafoCMWff4RO97MzJlv8yqoWJXI6Gpqu1ZLkD/r8ZmSpJknR7XurZEmvoGf6ag
GisVqB1sXo+xk58AAAD//wMAUEsDBBQABgAIAAAAIQC1VTAj9AAAAEwCAAALAAgCX3JlbHMvLnJl
bHMgogQCKKAAAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAArJJNT8MwDIbvSPyHyPfV3ZAQQkt3QUi7IVR+gEncD7WNoyQb3b8nHBBUGoMDR3+9
fvzK2908jerIIfbiNKyLEhQ7I7Z3rYaX+nF1ByomcpZGcazhxBF21fXV9plHSnkodr2PKqu4qKFL
yd8jRtPxRLEQzy5XGgkTpRyGFj2ZgVrGTVneYviuAdVCU+2thrC3N6Dqk8+bf9eWpukNP4g5TOzS
mRXIc2Jn2a58yGwh9fkaVVNoOWmwYp5yOiJ5X2RswPNEm78T/XwtTpzIUiI0Evgyz0fHJaD1f1q0
NPHLnXnENwnDq8jwyYKLH6jeAQAA//8DAFBLAwQUAAYACAAAACEAZVC1/4kEAADmCgAADwAAAHhs
L3dvcmtib29rLnhtbKxWa0/jOBT9vtL+h2zEV5M4cZ6ijPLUIsEIQRd2P1Vu4hKLvNZxoQjNf5/r
pC1lWK26zEqQxK/j43vPue7Zl01Ta09MDLxrZzo+NXWNtUVX8vZhpv8xz5Gva4OkbUnrrmUz/YUN
+pfzX385e+7E47LrHjUAaIeZXknZh4YxFBVr6HDa9ayFkVUnGiqhKR6MoReMlkPFmGxqwzJN12go
b/UJIRTHYHSrFS9Y2hXrhrVyAhGsphLoDxXvhx1aUxwD11DxuO5R0TU9QCx5zeXLCKprTRFePLSd
oMsajr3BjrYR8OfCPzbhYe12gqEPWzW8EN3QreQpQBsT6Q/nx6aB8bsQbD7G4DgkYgj2xFUO96yE
+0lW7h7LfQPD5k+jYZDWqJUQgvdJNGfPzdLPz1a8ZneTdDXa919pozJV61pNB5mVXLJypnvQ7J7Z
uw6x7uM1r2HUNk0LjgmCZ9PyV8u3/SDzPOS5xERBhn0UpU6CssDPrJxEfuzb33TjfG+Aa3GwfF7x
4X7rDF0r2YquazkHS+yIznTLtGBXhQASi2rJREslS7pWgqK3EfpZ9Y7YSdWBV7Qb9veaCwYWBaVC
1OBJi5Auh2sqK20t6ikXA5iXF3K9ZGBQgZqX06GigvUdbycV9xDprqW10XMGpBey6li7UEsWbW3s
LDnsv1hr9C8wqTWKrq5ZoSyabWjT12yYd3PISEIHNhgHPqIfTfsfnEQLFVQDojqdfPr+McIQABHu
3HIthQbfF+klKOaWPoF+sA0Vr9wWmAuQiL94jZKEpJmTIkwyDxE/8EATiYcyN3PN1LbiIHa+wTmE
GxYdXctqq0qFOtMJSPDD0BXd7EawGa55+cbgNcqDKI8CjHDiRogQN0Ox7RPkxWngp3Zs2ok1SlDV
3zvOnoc3Naqmtrnnbdk9z3TXIuC6l10TYdd2HV17HofveSkr0GOAbZg09f3O+EMFnLHrqE4KWXti
c7qEHnUISzGd6a+R68SxZ8UIp1mAiOP6KLC9DFkZIUlu2wTYjwyNA4pj5Qeq41trR7emdKiWHRUl
3DLqYlABh8oqQrWNuCixSunhAqUbKPL7yUBrP9ka87/bBszHW1YqW8OmB63t1otN3Tani5wrC6ZU
0iXIURWLgta3Oy4AX/GyZOpu1M/HzX87iU5IeHJ9Qs6MA1RQ2/sdAaZQtQFe6lQ4wKYVKIZsIy8H
Ob7BfhzCiYkZeWZAkJnZjtKXhXxiWwhUZ2WOl6XZpC9104b/x30zVoFwd4UrluB1ORe0eISL/4at
YoiFOrIyFPA9JBs7PigQKJIc54jgwERx7BLkpLnteDhNMidXZpjIquOvPlntfWNczahcQ/1SpWts
h+qZb3v3naupY5vcd2UjvEnVQbar/23iLZy+ZkdOzu+OnJh8vZpfHTn3Mpsv7vNjJ0dXcRodPz+6
uYn+mmd/7rYw/jGgPyY8xSQw7SxCtp0QRLzcQ35uOsgmHkkcEmfY9N4SXj8XT5/Lt0WMnSKTwx9k
22Kq8q/Aw+2vVW1gcjukasZeqYr+6K892vl3AAAA//8DAFBLAwQUAAYACAAAACEAc8li2DUBAAA+
BAAAGgAIAXhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzIKIEASigAAEAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAvJNNTsMwEIX3SNzB8p44CdAi1KQbhNQdgnIA15n80MQTeUwht8dKoGmkNAgp
YjnP8nufZzyr9WdVsgMYKlBHPPB8zkArTAqdRfx1+3h1xxlZqRNZooaIN0B8HV9erJ6hlNZdoryo
iTkXTRHPra3vhSCVQyXJwxq0O0nRVNK60mSilmovMxCh7y+EOfXg8cCTbZKIm01yzdm2qV3y796Y
poWCB1TvFWg7EiGs4wJnKE0GNuJt2YmB50C5GGdYnmGoCmWQMLWewkp08WOxh518MvgGyvbZvebt
Cn0uOZzz9R9o9pQDnFAcJRLtSTjVhuCfYSZnsvjbTIKlCPzhhxO1+/ao+5F0NX3rU524nbMTlEsD
yYs1bueopxnIUzA3s8LYpnQrflwRauufeDHY+vgLAAD//wMAUEsDBBQABgAIAAAAIQDVYfwQAQQA
AGgLAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sxFbbbuM2EH0v0H8Q+G7rbjuG5YVjxWiA
XoLstvvMSJRFRBJVkr6kRf+9M5QlWXYWcLYPDZJAGs6cOUMOz2jx6VgW1p5JxUUVEXfsEItViUh5
tY3I7182oxmxlKZVSgtRsYi8MUU+LX/8YXEQ8lXljGkLECoVkVzrem7bKslZSdVY1KyClUzIkmp4
lVtb1ZLR1ASVhe05zsQuKa9IgzCXt2CILOMJi0WyK1mlGxDJCqqBv8p5rVq0Y3oTXirpAWpt+ZxR
jJuVDs8NrviVPJFCiUyPE1HaDbXrKu/su0GdZXJLoSWVr7t6BMA1FPfCC67fTLnEKpP547YSkr4U
cCJHN6CJdZTw68Gff0aYXme6nTJNOqTrnbwJxg1syfYcW6uH8r5vF92ww/J6MP87wSYdGG6XnO94
GpG/wzhYTVerzch379ejwN/MRvdgGfleHHoPgTONnfU/ZLkwHfwkLbgm7FdawhncFzR1ib1cpBza
Egu2JMsisnLnD1O0m5A/ODuos2dLi/pnluk1KwrwBSp4qV6EeEXPR6DkYKx9FbwxlwoYpCyju0I/
i8NPjG9zDTc4GIdQEvbEPH2LmUrgmgDQ2A8RKxEFMID/VsnxvkMz0WNEIOTAU52DZTae+N7M9cCU
7JQW5dfTgqHSxBtCMdV0uZDiYEF7AJCqKcqAO4dIZOL532AShs7gJ4AtTRBkhSgQSCwIV2DdL52F
vYcdSOAPMnXpwOU83fvltqjoHBHY3Q7VfR81+AgqOpu6O1SvQzXV3LceWMdkuBZfR194PFx7TN9n
Ddt9+16g85C1f8G69UDWF2sxrHXVzoZxD6E5rz5icF6Tj3BE5yHH4IJj64Ec+7M0ux7DWsfx7oLj
xHD8xulPP8IRnYccwwuOrQdyvOiMGNb6Xuxb3BQAejEMaTayEYHmztV0y36hcssrZRWgH3i9AVI2
CmCeQVmMFU7sRWi4xe1bDhOYwV0DQSBWJoRuX0AcEPcz07vaEpKDbJihGpFaSC0p15BhjjopH09a
1wzI3uqhRBRsS5O30/Ds13xcg8G1KiBjRTVbi0pDjtPs/68z0WCvcwHfBtYz+3PHJYPzAV1AbYFM
0uje/5rfUjmtGaq666C6dpsJN71qpsgOTqqy3J5zL/Gf+V8wZ+DrrJbcCDqx6E6LDcfRAWZ8eeIJ
9gJqOnyXROR5V62K4gtTGsWUVkkupFWKPfvKdY5Dp2lhBdADy3KRSVEuF/AJNQfBRxFuH1vbb1l2
ZsY3s4IaDcKA3kauTw9n3mA23naTQos+DdzL99IE04kX9hQGudyT/TKZP3OdfqlJiKnsZhNwojRN
8ST7Z3iCBmmaqHm+bNbeF4boDc7dN/LyXwAAAP//AwBQSwMEFAAGAAgAAAAhALkLmJwZBwAAYyUA
ABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0Mi54bWzkWtly2kgUfZ+q+QeV3m2hDWMKSBnjBY8TU0lm
8iyLBlTRQiThJVPz73O6GyT1lTAk43EqcVUcRN/Tdz296Nq9Nw9RqN2xNAuSuK+bhy1dY7GfTIN4
3tf//Hh+0NG1LPfiqRcmMevrjyzT3wx+/613n6SfswVjuQYNcdbXF3m+7BpG5i9Y5GWHyZLFkMyS
NPJyfE3nRrZMmTcVk6LQsFqtthF5QaxLDd10Hx3JbBb4bJT4q4jFuVSSstDL4X+2CJbZRtvDdC99
09S7R6wbfyoujqSk0Gc6Nf+iwE+TLJnlh34SGdK1epTHxrESZ+TvE2jkpZ9XywMoXiK42yAM8kcR
rq5Ffnc8j5PUuw1RkQfT8XztIcU/Cz92xWGvbml/lz2/0FTP5F5qTMdI2V3AqVWqsr4vi6Zb6LJK
ZfZ3KmsXyni60u4qmPb1v9v2ydFo6NgHo6E9PHCGFthvu2cHxx3bdk/b9pl7NPpHH/QEgyephmXC
3nkRajAMvamlG4PeNAAtecBaymZ9/cTsTswWF4g5fwXsPqs8a7l3+4GFzM8Z7Ju6lifLazbLT1kY
9vURBvgqu02Sz3zmGJgWzC+9mGmPH5ZgRF93yKQTV9c8Pw/u2ASwvn6b5HkScaViHecYmqXJVxYL
p4Rt7i3XqYKlEunJJZRmX0REeEQ0RhFO9XkT2rlY80jQlM28VZi/T+4vWTBfwF3TOYQuQdnu9HHE
Mh+rGGEd2kKvn4RQgv+1KODbEbjuPYjP+2CaL/q63Tls21bHtKDFX2UI7ZMUmNytYiIoIibicz3R
bO01EXQQE/G5nnhUmZjlj3zFHT1pGxURKoBaq7CsvWxjpxUTj4uJjltO3Ms2uLbOGOfOOjP7WTc3
OTPbpeNOaX8RTKdM1qSeeEMWTrBi5OXeoJcm9xq2DbiRgVs4HswuV8w5YGGwgQEuqO3zOSd8kmR2
X88wejcwzZ5xB6r5+IHmQj18VtQ3E2ujl6NBIayYQq/VrBc1/Aa9HN3XwchSr13oFSENGyCOCjmV
EBS/1OKqkFGDlrYKOWuAHKmQ8wZIR4VcNECOVchl3V2rpULGDZCyjiIvV3VDVlkSAfmjAUKye90A
Idl92wAh2X3XACHZvWmAkOxOGiBldhXy8k2Vro0jLGC5vIodc9t2uWF1lXcWqdKwKrRJfU65A2I3
Kxhnk/qMJAQLt4SQ+pw1QEh9zusQlzKuQQupjziEVHfbxJdxPaI2CfpKQkx+2G52AZK0ayVppL5v
FSGJ4p0iJGpvqkKHuDWpe+6W8xXaoBjPQZtqTR1S9qFScEob7sAO2kiIYoLSpg6xKW0aIGRZXzRA
KG3q7jrE0Hg35EpC+Pla0IYU/1pJGqWNIqS0UYSUNkoWKW3qnrulZYU2uI48B22gpkiAQ6oxrApr
uw13YAdtJEQJmFTzrA6p0aYBQmnTAKG0aXCXnAbj3ZArCXmSNkrSKG0UIaWNIqS0UQpFaVP3vL3l
hsVf/5/hkIKakjYkj8OqsEYb7sAO2kiIQhuSx7M6pEabBgilTQOE0qburkPKNt4NuZKQJ2mjJI3S
RhFS2ihCShulUJQ2dc/b5Vaq7Da4xj4Hbaq3YYfebarCGm24AyptXBLNSEKe3G0kpCPeLfmryTkd
uKADlw2G6aVXueMTr66VqGhdFSGta1XokuP2RhGS82/S4HNJfaWunJLPUViF2i5ZRENFWiut8GFH
bdcY5WClN5A1plLd2shFbeSyyTotsOK/SSusRkdLrEppjdW80SKrUlrlJs/LrViWWTZ25Cu8t8qT
8yDMWbpuZjld/nbzje0ytGWW3py99dJ5EGdaiE4U7/fgBEplS0g8o/ElRnENlh2rzbcFOsYMKw8d
Il2bJQm8kV94s012aLW0y9t36XgqmkAhm3v+47p7W8pEew6d0xMeUOzl7DSJczSf1s3n/9qUHfSg
+3SRoDmtvWdfVkHKsP+gH8G7GrCUis7WD7WvZQtvyXgX0Wo5WPJF2pDZWLYxV2gWxppZ+lw28T4E
X9H6wvJfpoFo2aHPKAmCViWG+ZdJ4PPi8h4UGuN9/f0q3rQ3P7Ish1ov9hdJqkXJHfsU5AveXuS9
IWwqUK+MDHroU0aDHvr4XXSXBnwZbZ43gzez2aAc5t+EhHeIsH9xtGgWrR+4vG27rfUMiMQMQ9rJ
k4otLOgmW3x2IdnHnN3u2DV73JIhE8FbWpIck7R8xlNBJvlMSSvY9uOYXGWSi3IXTML2oDDJ+gWY
hMPpJZm0xdxrYBK2gYJJOAgUJtm/AJNw2L4kk7aYew1MQqe/YBLugAqTKifyT3u64dL4kkzaYu41
MAl3o4JJuLMqTBK/sdoc35tfdv5k9yS8YLwkk7aYew1Mwt2oYBIaKQqT2r/A6Yb2zEsyaYu5/49J
m5s6/kBh9/XcKP48avAvAAAA//8DAFBLAwQUAAYACAAAACEAeeDWrcUHAAARIgAAEwAAAHhsL3Ro
ZW1lL3RoZW1lMS54bWzsWs2PG7cVvwfI/0DMXdbM6HthOdCnN/bueuGVXeRISZSGXs5wQFK7KxQB
CufUS4ECadFLgd56KIoGaIAGueSPMWAjTf+IPHJGmuGKir3+QJJidy8z1O89/ua9x/fekHP3k6uY
oQsiJOVJ1wvu+B4iyYzPabLsek8m40rbQ1LhZI4ZT0jXWxPpfXLv44/u4gMVkZggkE/kAe56kVLp
QbUqZzCM5R2ekgR+W3ARYwW3YlmdC3wJemNWDX2/WY0xTTyU4BjUPsSJ4lxoldi7t1E+YjBDoqQe
mDFxplWTXOLRYkFnxGDn54FGyLUcMIEuMOt6MM+cX07IlfIQw1LBD13PN39e9d7dKj7IhZjaI1uS
G5u/XC4XmJ+HZk6xnG4n9Udhux5s9RsAU7u4UVv/b/UZAJ7N4EkzLmWdQaPpt8McWwJllw7dnVZQ
s/El/bUdzkGn2Q/rln4DyvTXd59x3BkNGxbegDJ8Ywff88N+p2bhDSjDN3fw9VGvFY4svAFFjCbn
u+hmq91u5ugtZMHZoRPeaTb91jCHFyiIhm106SkWPFH7Yi3Gz7gYA0ADGVY0QWqdkgWeQRT3UsUl
GlKZMrz2UIoTLmHYD4MAQq/uh9t/Y3F8QHBJWvMCJnJnSPNBciZoqrreA9DqlSAvv/nmxfOvXzz/
z4svvnjx/F/oiC4jlamy5A5xsizL/fD3P/7vr79D//3333748k9uvCzjX/3z96++/e6n1MNSK0zx
8s9fvfr6q5d/+cP3//jSob0n8LQMn9CYSHRCLtFjHsMDGlPY/MlU3ExiEmFqSeAIdDtUj1RkAU/W
mLlwfWKb8KmALOMC3l89s7ieRWKlqGPmh1FsAY85Z30unAZ4qOcqWXiySpbuycWqjHuM8YVr7gFO
LAePVimkV+pSOYiIRfOUQbrGS5IQhfRv/JwQx9N9Rqll12M6E1zyhUKfUdTH1GmSCZ1agVQIHdIY
/LJ2EQRXW7Y5for6nLmeekgubCQsC8wc5CeEWWa8j1cKxy6VExyzssGPsIpcJM/WYlbGjaQCTy8J
42g0J1K6ZB4JeN6S0x9iSGxOtx+zdWwjhaLnLp1HmPMycsjPBxGOUydnmkRl7KfyHEIUo1OuXPBj
bq8QfQ9+wMledz+lxHL36xPBE0hwZUpFgOhfVsLhy/uE2+txzRaYuLJMT8RWdu0J6oyO/mpphfYR
IQxf4jkh6MmnDgZ9nlo2L0g/iCCrHBJXYD3Adqzq+4RIgkxfs5sij6i0QvaMLPkePsfra4lnjZMY
i32aT8DrVuhOBSxGx3M+YrPzMvCEQvsH8eI0yiMJOkrBPdqn9TTCVu3S99Idr2th+e9N1hisy2c3
XZcgQ24sA4n9jW0zwcyaoAiYCaboyJVuQcRyfyGi66oRWznlFvaiLdwAjZHV78Q0eV3zc4KF4Jc/
T+/zwboet+J36Xf25ZXDa13OPtyvsLcZ4lVySqCc7Cau29bmtrWBLYD/89Zm31q+bWhuG5rbhsb1
CvZBGpqih4H2ptjqMRs/8d59nwVl7EytGTmSZutHwmvNfAyDZk/KbExu9wHTCC7188AEFm4psJFB
gqvfUBWdRTiF/aHA7GIuZa56KVHKJWwbmWGzn0qu6TabT6v4mM+z7U6zv+RnJpRYFeN+AzaesnHY
qlIZutnKBzW/DXXDdmm2WjcEtOxNSJQms0nUHCRam8HXkNA7Z++HRcfBoq3Vb1y1YwqgtvUKFCcE
b+tdr1HPGMGOHPToc+2nzNUb72rnvFdP7zMmK0cAbC3uerqjue59PP10Wai9gactEsYpWVjZJIyv
TIMnI3gbzqOzvO/+UwF3U193Cpda9LQpNquhoNFqfwhf6yRyLTewpJwpWIIuYY2HsOg8NMNp11vA
vjFcxikEj9TvXpgt4fBlpkS24t8mtaRCqiGWUWZxk3Uy/8RUEYEYjbuefv5tOLDEJJGMXAeW7i+V
XKgX3C+NHHjd9jJZLMhMlf1eGtGWzm4hxWfJwvmrEX97sJbkK3D3WTS/RFO2Eo8xhFijFWjvzqmE
44Mgc/WcwnnYNpMV8XetMuXZ3zrkKvIxZmmE85JSzuYZ3BSULR1zt7VB6S5/ZjDorgmnS11h37ns
vr5Wa8sV9bFTFE0rreiy6c6mH67Kl1gVVdRileXu6zm3s0l2EKjOMvHutb9ErZjMoqYZ7+ZhnbTz
UZvae+wIStWnucdu2yLhtMTbln6Qux61ukJsGksT+ObgvHy2zafPIHkM4RRxxbLTbpbAnWkt01Nh
fDvl83V+yWSWaDKf66Y0S+WPyQLR+VXXC12dY354nHcDLAG06XlhhW0Fnd2eLaiLXS6aLditcNbG
XutXbeGtxOaYdStsthZdtNXV5kRd9+pmZu2w7KlNGjaWgqtdK8Lxv8DQOmeHuVnuhTxzpfJOG67Q
StCu91u/0asPwsag4rcbo0q9Vvcr7UavVuk1GrVg1Aj8YT/8HOipKA4a2ZcPYzgNYuv8+wczvvMN
RLw58Loz43GVmy8Wqsb75huIILS+gci+aEAT/ZGDB44EWuEoqIe9cFAZDINmpR4Om5V2q9arDMLm
MOxB0W6Oe5976MKAg/5wOB43wkpzALi632tUev3aoNJsj/rhOBjVhz6A8/JzBW8xOufmtoBLw+ve
jwAAAP//AwBQSwMEFAAGAAgAAAAhAGbonXKrAwAAYQ4AAA0AAAB4bC9zdHlsZXMueG1sxFdZb+M2
EH4v0P9A6F3REcmxDUmL+BCwwHZRICmwr7REOcTyECg6kVv0v3dIybbStbPOAa/9IHJIfvPNwSGZ
fGo5Q49ENVSK1AmufAcRUciSinXq/HWfu2MHNRqLEjMpSOpsSeN8yn7/LWn0lpG7B0I0AgjRpM6D
1vXU85rigXDcXMmaCBippOJYQ1etvaZWBJeNWcSZF/r+yOOYCqdDmPLiHBCO1fdN7RaS11jTFWVU
by2Wg3gx/bwWUuEVA6ptEOECtcFIhahVOyVW+oMeTgslG1npK8D1ZFXRgvxId+JNPFwckAD5bUhB
7PnhM9tb9UakyFPkkZrwOVlSSaEbVMiN0KkzAqLGBdPvQj6J3AxBhPtZWdL8jR4xA0ngeFlSSCYV
0hA68JyVCMxJN+O21rJBX7FS8snMrTCnbNuNhUZgQ95P5hQCYISeIdNRypKVmfXLFI4vYCH9pRaO
LmDhsfip9Sp18ty3P8PhFVnTJ4j9NJAolLF97t6YNAVBlsAm10SJHDqob99va0hSAfWoyzM77yez
1wpvgzA+f0EjGS0Ni/Xcbo3e0rlv/gZm1Q9QUZKWlLDfIos+IGw2wTnkTujK4TefX0aX78/8/EJ2
zUfLfL68lA/nN6/XZcMGObmSqoSzcVdRTfHsRFnCSKUhCxRdP5ivlrXJCak1nB9ZUlK8lgIzUwd3
K/oGwBaEsTtzfn6rnmG3FRIbnnP9GdIJTmJTQXdNyKO+2eF1HYM/ROuwh7AhcH49LmqrvYJTqwMg
eJzVfjXCdc225ujpD5VTWOEHYl1/IFb0gVgDf4G5wyie8FfvPfDxMe9Fk6MJAta/AP11w1dE5fZO
ZkBfVnFWCv6P/C2ja8FJF/EswbsuelK4viftLhO8tjorMSECL7nqhHMGzN/p9s5DH21V/F6rjIH9
5oN73nkespUCasOgAD0rP/tCgswpnjp39t6PVQkXyb4coNWGMk3FkeoDsGV7qGf2LNHmJm4r3V4R
EC9JhTdM3+8HU+fQ/oOUdMNhg/Sz/qSPUluI1Dm0v5iyG9hbD6TUlwYumvBFG0VT55/l7GayWOah
O/ZnYze6JrE7iWcLN47ms8Uin/ihP/938B54x2vAPl8gj4No2jB4M6je2J783UEG7jx0Ovr2ugC0
h9wn4ci/jQPfza/9wI1GeOyOR9exm8dBuBhFs2WcxwPu8RtfDb4XBN37w5CPp5pywqjYxWoXoaEU
ggTdF4zwdpHwDm/D7D8AAAD//wMAUEsDBBQABgAIAAAAIQCM3O17wQMAACAPAAAUAAAAeGwvc2hh
cmVkU3RyaW5ncy54bWzcV91vGkcQf4+U/2G1jxVwQAEnBIjs2HFbtdgyOFKfooUdYOO73ct+OKaW
//fO3l1svAvEaeU8hKdjvuc3M7+DwdubLCXXoI1QckhbjSYlIOeKC7kc0svp+/orSoxlkrNUSRjS
NRj6dvTyxcAYS9BXmiFdWZv3k8TMV5Ax01A5SNQslM6Yxa96mZhcA+NmBWCzNGk3m70kY0JSMldO
WszbalHipPjs4F0p6XXoaGDEaGBH58rYjElycsOyPAUyBUx94aQEPUjsaJB4s9L00gCxKyAzZ62S
ZAap+kKsItpJwtKUWHQ1DfK3cmSOEYELS4TMnTUEzb2nD25IUWkjjD5VlqWh8JwZAzyUvmcijaVj
ZX0pofEFGJfaUFoAEWU7nEyiXIe//xnKxmdTcnE5jjrw3YXCE4+C730NqcerAIQsBKTc1LxCEoe4
IuIEZ4pAAdFoho9edIiwlqBVCB4zs5oppnmE35hlECb/C+xKRfBdXkQdMSavNWPLY2ZdFkYplUvQ
03UepdAA3JwixjMjuADQ/EQsQTIN0SRi09+c0xzPIwLtJoe5jUd8OLcuXpKJxaqjGAjflgXWcgtM
fsu+KB0BdebsXGVAPLRR/A9MCzbzF4OoRNpTwANi2AL5Y3I2JkeKr8NqjkDDFcgjwNMWVzPG9KQC
8TxlCyU5aTfbB/Vmt97qRpt6NpmGMk8UBpkCGQIzK1lQRV3DElLTkGkCcikkoMDYhMNceFKqc1gI
KSw+JlewTr5ZU2JxttLWBU86vQSuWeqwy7CU3YVDsRws4pZus9kMo3S8MJLOduI17PzaRpdG5MIh
U2Hw+WfWWeS8/Wn1zy82X3XPP7wLTXZn2n4h1fn0jdVI8G8enU1/Ugr3nEv/WDncpze7z6Sy+I+L
1GpuWaRiUls1e3DuFpOJcX7SRve2bzQWskPzpKivdkbdodnT3pYN+oTXuYQZHiXy9ZHjS7D+OnvP
cYM+12mVq0z1tKPLS9a4p188h4PiHGqkUn2l23tNWH4YohY47t/7csVDm9uPN3jFxx9fviCEXlek
aWifbCpQ9egVFKkL39QBKujDqtDaRnCCH2qRjL1NWQt9pL7b+PbwfBcWvGPWBz+ksYrvn6Wx2034
b0PAvbYA2ON7UC84gdYqQGnJafSuVrlVvwZo/8HtK7U/OFVDQKc9tLcRwvP9fcaSDX3G3YS44Vu+
LCLvu2i8mKQVjfJ7kCleyD8emdazI7P/zdp5/Tpm/O9YqYrdf8aV+jZw/2fjqhfYzwVcgv9wR/8C
AAD//wMAUEsDBBQABgAIAAAAIQCvbhdYhQQAADkLAAAYAAAAeGwvZHJhd2luZ3MvZHJhd2luZzEu
eG1sxFZLb9s4EL4vsP+B4F3Rw7L8QOQifqgo0E2DpIs9FrREWUQoUiWp2N6i/32HlBQ7bg7JHnYN
RBmSw2+Gw5lveP3hUHP0RJVmUqQ4vAowoiKXBRO7FP/5NfOmGGlDREG4FDTFR6rxh8Xvv10fCjXf
67VCACD0HIYproxp5r6v84rWRF/JhgpYLaWqiYGh2vmFInuArrkfBUHi60ZRUuiKUrPuVnCPR/4F
Wk2YwIvrOp/fcEOVIIaupDBUmB60zt+CWhP12DZeLuuGGLZlnJmjc7fDXlWS5XRwM4x/gaxZrqSW
pbkCCF+WJagPBwecMPCdo+iefm+ZojrFJIwB20bU7OWKcn4j8kqqbqpUsu6kXPJFcO1bPSu6DSB8
KcuzaTtyK0ruF9NO24rD3Jk2TDtth3gyY+TJXPi6uXiSROOTKy9shv38pdHRNAxOSyfDgzndoJpA
5FKMkaEHw5l4BLnzRTw9NHd9QPLbpzuFWAHpGkRjjASpITGXrTFSoBCjihUFtckMe8kckD5r00uo
VSzFP5LRcrRcJjMvGq0yL443I28WzGbeKg5WSZJl02Sy+ml3hzFE2qbB5mCQbqzNb4cAft+0s+0v
rn1n4iV+lkXL8SaLvQwkLw6WsbfcxDMvi0bTTTTJVtEo6fCTeQ4VYKD4PhXPKZW8O6XiPqWsgz+s
f/bnwd/UfpzUfSD7YjvzE5+53h3BRalLLhvgPgduXeCtsk2782vQcCFou/9DFhB90hrpwn0oFWQr
mUPao0OKgU6O9gsALk4o7ybzYRZsD1sapc1HKmtkhRQrmhsHSZ7Atc7dQcWiCZkxzh0wF2if4tkY
ssGu1AyqH3FWp3jqItGZr4BqNqJwOwxhvJPBAS7649kj9XV4WMriaMG28B8OChRpvsCn5BJM5Zw1
GO0VaVKsv7dEUYz4JwGlPErGkwQS+Hygzgfb8wFxdQ54RmHUNortKjh52PnLtXkwR06dw431xUac
8B3kttugDLextSsFLe9hUf9tiyJw8YaDueOQOZwJPrDMiWV1Krzbz8DqVheUMdq6e4LCALlNsQCi
t6Sv2CPcrJAPTgI1oikUJcx1NrXkrLB3YOG12m1XHOJErE/PUQcnXqhxyHWBzLGhJckB6aYxUqNb
ooAt+pQERy2gWdy3At1wjr5SbbStM6hi+MKqPVp/Zaa/KJeeGmYdO3IGpL8mhqDyTjFh9F/MVA+2
zfTJ6NQvyNaHxtGRu+shGeF8S/JHm3mvtJQ38vXov+XrnvAv28NoNpoM5Nt3jOceEY1fbxLhJIwH
jn/RKQbC3kH6VyzPFPDvM3UPdP3xbPFX4o4G1rZ3e091y80do6uKqK7kXyPt/5dUJ6txMI2nibfe
TDJvvFmvvXEGfSOC5jGezrJJZlvGm0n1ZXhO9Ppa2N5NqbYEhk39JdmK6kVXFq4Pvvu5lvcXlM+d
1DesNz2qLh5+HVIP8KaXY/eOWsu8raG2u+ejopZPpNAVazRGam77n/pUOAIFkjg7sSON52C48j/P
30veGG7kkiKGV+/iHwAAAP//AwBQSwMEFAAGAAgAAAAhAJI7gS+lAgAAXAUAABsAAAB4bC9kcmF3
aW5ncy92bWxEcmF3aW5nMS52bWyMVF1r2zAUfR/sPwj1oS8JsbM1HaptyDrCXvbBKHsZoyi2nKiV
dY10nbj99buynDSBUhYcW7o69+OcKynrG8Pob73Y5bxzVvhyqxrpp40uHXiocVpCI3aN4e/fjUh4
Cwl1rUsl4ufFp/8PH9WXyvCC8mQg/Fa2ysgn6JDthOox56rSOCyHdV01sj1bYZVEmfOUz4YQs7MY
RbaLIfGpVUxXOb/vE/rd4zxJOSsBXOX1s8r5PF0kyWR4c0YxWsocMFQWayVuc95MTFx3EWvip1dj
cZQJHTwq9gDaenwyFLXRqFysjFEpIRAjihXs4THndchEHF3nNdjREoOMk1obc0SWYK0qMVDJuaPR
ITAIA+XjuSqDkBE64rLZiRZHYU5F8Wkyv+Isel2cKRX5XLbgNVKtQq49mA7VTdCnkW6j7dSoGkVy
M84QWpHOr1q82esKtyL9mNB4q/Rmi+LDnMbPU20r1Ys0BvEw3TvZTodUAgPukvRZd4hgc46cBTVK
MOByHq0rWSr2Z3H9NwhJoitsoCJxZIdw7Erwok6T2/xVv0ohadlA51VpdEl9eVNYByiDBCewnUDq
4hp6NsrUEJdKhxYRcCoNilBRYOO13ZhDnjrWyLJK7w6uIRJ56I0VpbK0ey6LrAaLrCauOV+2CJ59
l87BnrNx686TYZtGkjm/CDs8SXjxq7NsSeTvlEefzUKYIptRsnBSWNgPY91x3otboynnFzpQ7Mf6
gcq/G/ba56EFh2J7sbTlFtzgxFgyCc+n4Z1O2PXVhKVkWGSzc2DWi59OW4yBi5U0XgXMqTGUxQi4
JLVW1LYX1NFygKwaI79JuqwCSyI5cuzFy8IBekeSfl0GSYvbQdKQ9sR4ivv9Gm40RtFOVYo3znis
SFq6Uot/AAAA//8DAFBLAwQUAAYACAAAACEAz2+WoI4EAABcCwAAFAAAAHhsL2NoYXJ0cy9jaGFy
dDEueG1sxFZLb9s4EL4vsP9BK/Sq6C3ZQuzCluxFsW5ixGnvtETZQihRJak03qL/fYfUw0riRZO9
rA8yOZwZzjec+cjrj08l0R4x4wWtZrp9ZekarlKaFdVhpn+5XxsTXeMCVRkitMIz/YS5/nH++2/X
aZQeERO7GqVYAycVj9KZfhSijkyTp0dcIn5Fa1zBWk5ZiQRM2cHMGPoOzktiOpYVmMqJ3jlA/8FB
iYqqt2dvsad5XqQ4oWlT4kq0UTBMkIAM8GNR895bagfMeeWxLFJGOc3FVUpLs3XWgwJntm8OqOaQ
pAwJbE8tT3tEZKZbuimFBFWHVlAR42bTChltqgxnMWUVHMdIv0yjBRGYVeAqppWAqLt8lW/KeInY
Q1MbEG4NIPcFKcRJwdbn1+A7PlLIh3aHvzUFw3ymp7Z3ToH33gRYoTkxne5cAaztRVycCG4B2ZYj
0ZrDviqENSJkj9IHmZuR8qB6XpeGL5MhrVQZyYEoBMFq8CS/rEiP82sU7Wl22jLYGUWEi50MSE1q
Kam3TP5lOL9TOuYgUnK2ZZo8sZk+nBaKxPwec6FBwhoi+DWYCICFIrCQ5vDtN4eBULFQaDOCTqOT
lUt9vKgR9F5OEkywwNmzgqkJFQuGUVs8J9oIVUd1gWPZhFL8iNgppoT2pWO3VcUxRJRGRfb0zCFl
GWbPJNlmT7g6gC4dFV0XhLQpq2R+BgEAJEqC8xynYsNlNICFK0v4P9LvG3zAVfYXHsNtV74iIBzZ
DH2EoB0jcYPKrka6JpFedphdlG8xS2UbjJtK6i+b/Z7gXfH3a1cbjADypqhw11vd9vhJApDAYaQ1
rJjpP+JV4LvhwjaSYB0bXh74xjSZ2kboOF7sTT1/slz+PPeI/26S6PujzVsfAqRwOIUUqZi4YHc4
l9Hl8wTx454ilv3xIfngR/AJZdbVKujFCEhXatYiBirpkuO2ZVALDUpAEpCqlfl2sdtJ40d1cLXa
q9exO5314tPm33ScTufm9l67+3LzXE2WwhBPO1EgYNjBgnOTcVRNeQndSqJbjdCB3oCuvUtimuH5
nxiYEhGVBHXDKOkbM+D+Cr79K+zOK9TnOAFqD05muYV7sdRcd2LFsWf4SbAyPGsaG8uV6xir0Emc
aej6bhyPSi14d6l5IyoOoqYqvjX4U8cuP6z2Zxu2DXt6K6jx6SRYGIk3idfr2F3HE/+nImzVHqpY
RtXakcuZOt7W+KMG/98af9xrecHgTiBwCy76e1mFCDpnhpWTMQsrflN8rEZb2tEKaxvuItmTwUi6
+lrw24p0BNmxUVbwegmXzQNfdP4OqO6YHJKbSJK+hXvkM3rO55dKyw+W01VsJ4Y78aG0bG9iTC17
bYSTZGJZThiGsTUuLea+u7hC03LH9cVc+eBBCaCAm+62Vm8qO5DPgECugfwGLbhCOLoEzG75oikk
HtC9LL3hwq9ZUYkdFgIeler+OiqiX1MKT6b2lkQH/BmxQ1FxbQ8UeBX6uiYfY1ehrsGTUf2LfqE1
lzNX13LlpZ0MvmCzpm6Z+8XmkuCGR/H8HwAAAP//AwBQSwMEFAAGAAgAAAAhANZK18B1BAAAKicA
ABgAAAB4bC9kcmF3aW5ncy9kcmF3aW5nMi54bWzsWm1v4jgQ/n7S/QfL39OEJKSAGlaFktNKu92q
vdN+XJnEKVYdO2ubArfa/35jJ2xB20p71QKqlEg147dnxuPxPLjh4t264uiRKs2kSHHvLMCIilwW
TNyn+J+/M2+AkTZEFIRLQVO8oRq/G//5x8W6UKOVvlIIAIQeQTXFC2Pqke/rfEEros9kTQX0llJV
xEBV3fuFIiuArrgfBkHi61pRUugFpeaq6cEtHnkFWkWYwOOLKh9dckOVIIZOpTBUmBa0yn8FtSLq
YVl7uaxqYticcWY2ztwGe7qQLKdbM3vxT5AVy5XUsjRnAOHLsoTh24UDTi/wnaHoln5dMkV1ikkv
BmzrUbOSU8r5pcgXUjVNpZJVI+WSj3vBhW8HWtnNAOFTWY6fmm3N9Si5GsfNaCtu22x/EvWDdgZ0
uRkO9UmXkTs6e8/rtCg/en5FbZQMomf1brXpGlUEvJdijAxdG87EA8iNKeLxrr5pnZJfP94oxIoU
h0E8xEiQCoJzsjRGCtTDaMGKgtqAhrlkBEgftGkltFQsxd+SaBJNJsnQC6Np5sXxLPKGwXDoTeNg
miRZNkjOp9/t7F4MzrahMFsbpGur88s6gOeLdrr98YXvVOzjZ1k46c+y2MtA8uJgEnuTWTz0sjAa
zMLzbBpGSYOfjHI4BQYO4PviR1gl/zus4jasrIHfrH328eBvYAuQek01AGFgW77jHdObJTgvNfFl
HdyGwLVzvB1sI293GzRsCJqvPsoCvE+WRjp3r0sFEUtGEPponWJIKRtbAoDzE8qbxnzbCrq3U2ql
zV9UVsgKKVY0Nw6SPIJpjbnbIRZNyIxx7oC5QKsUD/th302oGGQAxFmVYrfYVv0C0s1MFG6GIYw3
MhjARbs8u6T2LK4nsthYNXP4hIVCmjSfoCi5BFU5ZzVGK0XqFOuvS6IoRvy9gOMcJf3zBAJ4t6J2
K/PdCnFnHfCMwmhZK3a/gJX3Gndxbe7MhlNncG1tsR4n/B5i201Qhlvf2p6ClrfQqf+F2bC97e66
5ZARrAkK6ObEZnYqvOsPkNntWBiM0dztExwMkJcpFpDsbeJX7AF2Vsg7J8EwoikcSmhrdGrJWWH3
wMJrdT+fcvATsTa5pzVibxiHWBfIbGpakhyQLmsjNbomClJROx4MtYBmfLsU9nTB2YUS2qAEL7g4
NO32uIqGVpcWOYN0f0UMQeWNYsLoz8ws7izBtCHYzN1Psz5QRpPWHXtkhPM5yR9svD1DJh3D7DJM
/zQM84LawzBMHw7FHsOEx2MY0P32GCbsGMblr45hOoZxX+Q7hnn9HSY5DcO8oPZADAN3lj2GiY7I
MO675hu7w0Qdw3QM091hnv5V1DHM6xnm/DQM84LaAzEM3Fn2GCY+IsOEb/AOE3cM0zFMxzAdw/yO
9zCD0zDMC2oPxDBwZ9ljmP4RGSZ6gwzT7ximY5iOYTqG+R0MMzwNw7yg9kAMA3eWPYaBt6zHetPf
j98gwyQdw3QM0zHM6xnG/VrA/sxt/B8AAAD//wMAUEsDBBQABgAIAAAAIQDhOKvZLAMAABIWAAAb
AAAAeGwvZHJhd2luZ3Mvdm1sRHJhd2luZzIudm1s7JjPT9swFMfvk/Y/WObAha5JaMswSSXGhHbZ
D21ol2lCJnGowcmL4tcS+Ov3HKfQSBvqoZdqqdqmtr/v+flrfxKpcVMYRp/SilXCl3UpbLpQhbSj
Qqc1WMhxlEIhVoXhb990SnhNCXmuUyX85SWm2SJGNakyfE7zxCDsQlbKyEdYIlsJ1WDCVaaxHXbj
Oitk1RthmUSZ8IiP2xTjXo55vPIp8bFSTGcJv24Cel1jFIScpQB1ZvWTovhwFgRH7TdnlKOimZ2G
ymKVxEXCiyPjx2uvNf7SqK44mglruFfsDnRp8dFQ1kKjqn1ljEpxiRgtMYMHuE947maiNdZLq6Hs
enySrpFrY56VKZSlStEtJeE1/VonBmEgve+70hrppZ0uHm948WzMpik2CiannPmog55Tfj2HFViN
VKuQNxbMEtWZ86eQ9a0uR0blKMKT6LTCs64LoRKz03dT6nnQGS7WjYXStwsU0YQGnka6zFQjQp/L
wuihltWonVGg0x2STTdLRCgTjpw5U1IwUCfc917KVLFfs5Pfzk/yXmEBGXkklwjPm+OiaMMpLPpr
XKaQLC1gaVVqdErb86q/NaB0TmzIVgJpM2+gYZ1bBa0l026nSDiSBoWryK3G6vLWrOfJfY0szvRq
HeoyUYS+LUWqSjpEh/M4hxJZTmtN+HmFYNkXWdfwwFl3gqOgPa1+kQk/cAc9CPj8+7KMxy54Ho9p
CocJc4ehq9a3G3FhNM30kWhiX2/uqOir9qB9aI1fl9iI8zJdQN0GMRYGR4zek6P2VxjStW1N38fj
vjRuxLdal+hTzy+lscppNjtdYYyE5+TSJW3Xi+q5Zy25LIz8LOle5Vb3Q5GXqLIrZdHlfBlcy6/I
zk/nzs75RWunk210bup+/k3XdXrrNr3yN52OrH9RNQ12TlV4vCVW0YDVvmI17WFFrQGrlyc4Paym
9Ajf8cMqCrbE6njAal+xmvWwotaAVR+raPdYnWyJ1WTAal+xOulhRa0Bqz5WxzvH6niyJVbTAat9
xep9DytqDVj1sZrsHKtJuCVWswGrfcXqtIcVtf5brMb0R/D8DwAAAP//AwBQSwMEFAAGAAgAAAAh
APUKwQ/HZgAAAA4BABEAAAB4bC92YmFQcm9qZWN0LmJpbux9D3hc1XXnnTeyPLZlM5aFEcbYz0JG
Y1soY2ODcDGMJNsIMLYiCaMEAyNLY0sgS5OZkRH/khFWQaEkFYS4akqJCAnxZl1W6ZeyMpumghCi
JSRVKNv112VrhVKi7f5BScuuvmyL9/e7f2bevHmyBd39drNfnn3nnnfeuefee+6555775z1N/vny
qWe/vepnwnVdJ/ziw7OLRKEDbwH2mfugELxn+PDs2bMGfT8A3P7m+jWRwD+hnAvRZvcgFCAsQGCb
BxB2ISxCWIywBKEdYSnC5QgXIAQRliMUI6xAKEG4EGElwkUIpQgXI6xCuARhNcKlCGsQ1iLYCOsQ
yhAuQyhHWI9QhFCBEELYgLARYRNCJcIVCFUIn0AII2xG2IJwJcJWhG0IVyFcjVCNcA3CdoTfQrgW
YQfCdQjXI0QQahBqEeoQdiJQl+9GfANCPcKNCDch3IzQjHALwl6EfQgNCJ9EaERoQtiNcCvCfoTb
EFoQPoXwaYTbEQ4g3IFwJ8JdCFGEVoSDCG0IzDuG+BBCJ0KHxh3TMZ8fAXwUoQchjvAZhARCEiGF
0ItwGOFehD6E+3RatvODgB9C+CzC5xDSCI3g1IOUNlq7G3FCpsCDeV4roTFG1akT57rS6x+//j/8
j7d9fuZNZcC1H9JnG3zcKwCLZPKnrnpd0X//OzJfk7+TphlS7oT0boMUEugFB6U82Bvmd62CBXR2
9fmmO8nOhusu/GtCKzQgpv5+1CsoLN9HzZ999QQNNy7ac4JsE2f/p02gDXD2f/ZL9v9lCO7+T92c
b/8nLfs/Y6/+T/syn/7P9P+c/s/07v5Pm8d+7NX/9wDv7v/kYfo/7cNc/Z905+v/tK/s/6RlHzb9
n/aIGtmFwP7fjeDs/6T/uJdvpR/2HlpwEBrwLhgHoQ1DgJ9B+CW1AheV5PV4EPU/e5aGehDhRcA+
PpDqr+I4YCGeXTbz6vqnonX/4qLbkkUf/sm/uTigdG1cMaN0cy4yubhuc8nDm66p+Ub8gZ88+p9f
uCEIHDPncERVNNXz+aiUvKNxC7iTubNWfsktaFTFQBVT/arCn5FF/n/BZ+FoZ4mmPSwpRd4UzEBP
KQgjW5mCjOh9GIvC4gEE/tuCUI0R8RqMkgoTzoHqNDb71FAx3opx8yEpbCUNn2+cJUor7bK12MPE
weJQhkqOWcmmMHJxfOcYTmxExxyHrWOKj0nD5sttlSwfPrkU3gP5PDxSbvGJ0fFR8gJ343MaxZDs
HOmIjxpk1mvNs9UmvTPOlkxh00DQF+BFZZWqJ+9mA7MBLw5nzwYlLWv+/Yequ9DDyuHGHIChP6T/
bYTIN0PkbRB8NdyVbRh6wuKM7FGylNnSZ6rxkYAPXycvli630ygN8o2OippUKtF5sDclYvb+2rv2
th4BsMMua+7AgHhbT+Kegz09wXvKli5e/Hhta9JnPR5+IBwObwk3VG++5goRDPsDdQtFiVXsKw+H
t171UP/SB2/o6jnY2uWbDDbFW9t8T+5u7SpIxora6xKx1lSg9WDXhWsbErH2oVhbV6sovLFdfDvQ
nOhdWrurL94DwpLVzbEj8S7RmortjCU6j/qt4idre5OpniOdhfcX7PTTzcu7ErdGXDhz78azLXgt
d9BnYRgaJfocbj6J9IkaDn36+owDftYBGz1Y7sAF0LBRpPupA/eyNqrM0MfhVF+5GuitVb/B/jpI
ADMfH+cSuOhmn/Oav/5yFsUrq7MFedYsa5nUrOL8JXD6n5w7fdSL/icVmnNEZVnOz4HzP/o0vGrh
1bRiZkk/7uNcRcifc1P3OHAuXjYevqd7ocnf07jMo0DMn04KO/J86x925N8ATy6JedcRSKEbM8yY
vGvEPK4b/2JybneuYpR/jPkH/feff0Zxpf/nmwH8V3DS3iIOMSvyM8Tbcvy/FZiHntv/m68A8isU
BGoe/t75GmT+/l5pTt/5v2FV5uPvWdcqkZ7L36PX9/H9PbW+p7wE+jgfxd97CyPmXP6e1q5Mg+WO
bt7+Nom9WsIwYeuq53NRKmn55MqP8rvOnE9lzvX8PC7UN+d0oWq7Wu32zXSdpmpbraQVV65TNFy9
JXyFsAuC0nNanec5vZLrOU3M6Tl96zyeU/n5PSfdQvPv/9WY2/6m//9vMxW/6f9uUf5/1v+3/Dr0
f3cbhDgS64tgIOPVzKX4HANGsRyo57nFep67GfFlmNtyScJz8rtNTn6vxGrFIUx+20Hqy027BS7R
lfOdOLvSbga7VjfDzWC3zQvZ7kay2NfkU26Fm5qX0ZWg9UDm8bwSRUJ1cutI5JY5Kr5MSzIsJXlI
+II5CFbRl0sTy6e5Kh91dT4KojLMTbvlFZXNsUWcKZLbPh9vVOV4esIn0mqhXg7W6YBcfpb8nPiB
gNxW4mqfAMwtJZuwi4bzajmxBw1huUyj4RYPek49Olz42v4iLIGukauaIh0oYJ5DkkbBww54xAGf
cPPvDxSUiD1i1ANfjE2uCeL7A4GQeEZM5dMElmGDLMBFiP5A0XXidUwiFb2F0rUQdtYdfIqwsRbV
9JRWXMMWOm2fI21awQWLIcVBNx+5/jsD5jeFlf8/of3/zZgR/Sl2q5iA4fU49+zOnjWrF17rvzQW
rNhHv4JIcn7/X+9ZzM1+/v7/qMvTnP8KsLKI/3w3IAKpLhRhrO1eh3r5RCh49uw7sobXLTYTqa/o
9V7soep14RyowdCtyqwQe9Jl0yIPJaNVXpBcy6R0V3lBWS5eTzPjRJazde2YVgjr2vEMNJGBJjPQ
6Qw0lYGmM9BMBprNQLnzIrN4p+YxM5VCvDO8xWrAqMQ9LWLrZEq1V8camvYzuvTO8FVWOo9erbB4
0S8QzbFkKon9sZ2tyY6DPa2Jdqn6xRLfGEv2dqUaOmN1Ha0JbsuKTcXbrSH2I7Xtm5e/KcfIChgf
LG6bHchp2TFUbUuxJm4WEh+31HrlhmM75Po4jRappAEBHEReA4jN/KzzWETSBQ2dliTTkC46UCuf
T2nB+PXz6MBOiT+t8aaLRwd2S/xkHr5e4idc+CXFn7YapJ7LBdE562/0k+tJvEw7YQaMswQXyzMC
xKPM0hLZA9fK/DLtqW2EqXd04EDO82y97pwDH/XELynus2bBG/+jznLpYmaiDgi4BYUx7Zctv4V9
02Xy/EFu+VX7zWrC8bzy3y/LM5MnZ1X+aVd7LSmOW3EsSEE45yzn1HKcH3DoWQMSKFZ+nI+4Tp53
kOX0KTl3HlP5GT0zcjT6Yw/k1mNQjxPZdnh4jnoMWMzb1COrX4/k6GMWPziHPj42hz4+7qmP0YEv
Svx4nlyHrBMoz1ge/knLhjEZzesfT1ncLj2RR39c8h/R+CWZ/jRszYB+OI/PVyT9UB7+aYkfzMM/
I/HpPPyIxPfl4b9mxZFvPA//dasa9erIwz9vjVOJ8vAnrCD0izrOy+jBpuJvWdNYTAKrc+sdPLlx
h95NgJcaUQvk7uTndU+ay75Nc7EV1wziXPt2p5XGoSN0P6nHbv009s3dX6ID35lDz16cQ8/GrBDq
afSmKNOuL0n6fL35rsQbvcnq8fesCPgYvcnqx9PWCPBGb7L2atwaQ/3cerOp+BWrFJ0S4ji3XcJB
LBvtZexSBLvqSu4L4EauEqdccu88puqTGVdMPZGOcu88puRm5G3qRZ3gc3tgIkeuY9zFl1dabuGb
O3d8SWBYliw68HpOvzVytgfeyOm3QTlDIJfz8Z2RfDcVv2ZNIQ1EfE55xWG4Ig49jSONklch/OC1
4i/y5NUgy2XkZco7BISSl9InG+koI7e8ogOTc9ivNz3t16biV6045lGo1DnrMYGDd7QFpt2rkUbV
YyH2LJaKX5p6qK1OtGu7PE/kbnfTrtEBVR63vV5SnLAmMWgj3bnLg8N8fZArz6/w0mZESmQzptdm
/pAd3/+jrL8ZH6f0IS8zrtgDZ+TzGc1IbkSBr3v8d/d7e2AqRz+rtcNxKWrAcwfRgXdz+n+2H76X
005Z/LQVAY+JPHv5d9YEymzsRZY+asVBb+xFFq/st7EXWft60qLzB1bnlG8LFCziaO8QNsdUey/C
FHeNCCsfVDwuD/2wH6tx2d3e9Uin9HZA1jfo0tuOzHM1Tpe6ng9mnqvxF8XK0fsTmedqHA65nk9k
nj9pTa5TZz3leKPPnhB+GHWJDvwqp38s0nZqU/ELVjU25SDic8prrAy7tQ55tUDGSl6Lcc71crHX
yEv7253HAn7Vf3Pr04d0Sl5DnvLK2sUimX5K68lk5mTA+ezXqCxZdGCZTH86T88KfadBManx2XFG
2SWjl1l8UPIxemnwm4qL/WFsEEL855RbGJ3ktMM+DkHWSm5L4KOXi9Q85TaKdEpuJbI8bj3L2p2V
OXIz7RwdKPVPsiy63ln8KkmfL4/VEu+Wx6biMomHCM9Z7yEcTHSOC4OQlao3F8aCYtjUe972tDKn
PbP+hGpnU/6sfajyLP87Q1WWjfkrC09R1Ot+QBiojJ0lzOudobCcaXTk0KuztJ70w1utMB7Q/SJP
NZ9mqebiXy37QV8Of3VW2Iv/DOTI+TrHZuof8+J8nPzp3zBm/RhznGHM+TRjUwb6n4Q5T8vg0aC0
A5wTEcc56BB4c+zkZR2rlOU0ZWoJZkvnLifTX4o25vhAeFiXx+TF2Cyf8XytwU85YCetE7YxH+tA
IK5ax3PRzgcfcPCossqtLituJfTckm0YKfjwbANopjQz9htb49PAm/5BfTT4UeDHPeiHgTd+rJOe
pEOa3omfAD19El5u+rgHf6JaXPQNmGgMIxhZTAGeRWhB5UYRTyCEAEcQ0oADmrYD8WnMa0y64WLQ
rVD3o4AN/uPGaQcPyn279S3rpPWCtceKSL+Xcp8ATYNHPaeAr55DLmEP+lnQc+xyyzGI+piKOOXb
Avy0PkjpxEeAn/TADwI/5oGPAz/igWc5BufAc63BWc4TJVisd/RfI+8+4AlzATsCmP6Ksy244G3u
Q5hLpi9EXRGIm9ZpnfTngsM6naFhfpMILQjRgu05fictHX2kDjwjnbmGdZ7m3qyDzGimQ3iexkPj
j47jnnU0V3RArcs5/WnS16NslJG5yG6PdZO1xdpvtViNsifz7Ab1aRi0ozo/Z7uWAj/sgZ8CPu2B
bwHe9EvW1wb/OHCUk5lmMT9esp+uVfN/0g2iLZjeTdeh83Hal1HQGf3PsQvAe+l/HLxtj/LOMj8P
fBD0M3PoIX0WXjn6D3qj/xzLWR/WswX4CU3vxA/rNQzyceJ5f8LFnzjKhV2SF/mOIr3pP870E8AP
e+RHHlwTduanq322RY955n5S348h5hpJUN/HHXSDGg5iTYAvafUhMH2pjg2v88UnXPRcf5nmGgVC
tOBbOf2H9aRunMAz0pnrNGDmY+7d/WeS9HiY6T/Ik+sg5mLa7dZV6Bd7rFesV63X0DO41GD0dEhX
wql/YdZZ4516UA981AMfBb7eA09U2AM/BHovuzwLvKws0jjznQR+Vrev6XfTwFG+xgSY+ki9Xpvt
dy1oQ6Z303npeYi0HvoVAd5Lz1k1rtXxMnpKmPlt0A2QhrCJ68DxVMan9b1M64C1mM5GNJ25H8H6
xCCCpNfPTjhobP3M0Js46sJXO+4DSD+EYCNEC17x1MMwnpHOXC06T3MfHVDrXzM6Q9afepgGHeuQ
pfuORT5u+92A8lAmTj3dY/1Ee2AKyxu26wzovPyn06yHzt+pLwG9psNHbv/JS39JN5f+RjT/Eofd
I8rYYYPvQ56UsZmuGn2U+v85vvul7OYYaNg2brqQQ462rjdRtit/3jI9h3Ve+vF5oxb484ZoZDVk
j2DuIzAehKc0Lqjv58vbTUfeEwh8hzFa8Kqnfg3pMpj2H9N5Z/VGrbO59Yt5sfxOvXnL+kuM+act
HtHlBbMu9WYWdJO6cE49CKF+Yx54yXsO/KAHvgV80hrvHueM/XTaBaec6nHSwnkfx30QYQKDwCjC
uWQ/odOOunhk2lMeRDh/6zG/YYQQQrTgNc92qtblyrSLzjPbTmr9ZMohnzQeDoKOdczS5e8XkK4F
ZWXdzUU29OPYpq+62xN0ox7tMMV6eLUzeJv2cbZ/EHjjdznxHcA3ePBJA+/ld5FU9m/EznaOgN7Z
v4dx7zXejQJvGtpZjingveYhE8B7zUNmgfeahwRt73lICPj5zkNoxyKgb5nDX+RaDy9Tfq5LzqV5
0/pZyzq0CwLpTug4rGPiRspQPoQgAu9DOp6L7/nwsyj/LPgHEaIFJz313MYz0pmrXpcnq7/5+7PU
3zjoWPYsnVoPdo93w6BhvbN0+fu35DeuZZOle9xiuU7rSppxdhp0lF2WTq0TG3tn6JiM68Pm4v2H
en5kcGac5TOv/hWB/L36VwvwXv0rDrxX/yJ/r/5FvBl/nX4o8RFHvSEGMYOONXRZVscCWLuyEZw6
MOZ4bvDDSNcC/Oly8EAgPqjjWQd9XOOc/NzwoIuGfKcQ4gjRghc89SuNZ6Qz1wmdZ7b91Lr5jKO+
aTycBB3rm6VT6+du/RpBmVhvZzvTnyr2Z3EEZX8G7QmdT864CPyQB74FeDOe/Z/wp0bB39hX499P
AEc5L9fFz/Gn1mb9e3s9xgQPupCuh9NvC4DWdumT4Wvss9OOky/nVbyYLHC5isOIJ106x+d88d+p
K9OaplTjx3T6uI4HdXwuXTvXM67rt4DHNOJoQbFc95511I/zyVk8I525bJ1nVp+89a4edKxvlk7t
P7j1jtlRHubi/YdWmf9tsdH/K400/lgYewNe9mUE6b3sCz9y4GVfJkFv9NHpd00DH5+jfY09ctKX
gv9c9sjoY878F/RGf5z4KPABna8T3wf8fNdZxrWOhEPz8+lPg74DtCccOjeyUaWtRjyBZ9Pz5GV0
rEGnN/csyxBCPUK0QO3/uPWrBc9IZ660zjOrN95+fAfyYvmdelMs3sqb/+E0dADT6bkvvd42rgs9
BZ42qC8JnABUJO0dJ+Ne9q4UeC97Fwbey97VAx/1aOco8PVe7Q982AM/BLyxN059OcFFAw/6ceC9
/MHTwHv5gzPAe/mDAcjca13aBv6j+IMsopc/2Ac+Da51miHiEDD9lJext/WkW5u142OgYXo3nfEv
nXaZdMa/Hlmc31/GgGtBCOhngzq2XbTTyMzoenoT+gwC74d1bJ59nLjDwYP7HGHsdHRYW7HPVG3R
Lst1TtAYvXT6PVPAe+nlZCXaSRfGSd8CvLF7Tn2aBZ8WD30Kgj7igQ8BH/LAR4APeuBHgDfrcs58
48BPecwXBoE362huP2LUg34M9Ga9102fnmM+0qfxRj4s9igW/3iWB8u6mWtIdzXuqzrbl3urvDd7
q/2lBSnxmty2LujnuxgB+QUY57sGZcAXYtuzlMZLvx9hO+CQAw474GoJK4VnlnivAe9HoGmIB7wS
UYuGIQkR96CHRosRD3wx8KMqrRzFue3K9yZIP6Px0tbSJ0RexAcI575DUUCRlXrgSW974OH5izDw
SlYov4tGyapQRCVevT/S4YDjDrjPg38JvhCS9sAX4dXwYV2XVeIpwSOq7rosxjsgE5qmEAaE25ym
Tafyysl3XgrFtKM8Mw541gGLgmy7cxvV6EDQAZc6YNsBhxxw2AHzyKnhE3HA9RJ26IzWtwYHTYsD
jjpgbu0annEH3OeA0w540AEPufOVutoot5jdci7Fh+lGSS91OCS3mgmvAn5K40kj1+mAJ1xKH1HT
hB1wg4Zt0PCIAmkuBzxE2KWrq/Hm0pimIXzaAc864FIqNfiQptoBc6vIzZNvo/VpmlV4CW/Ig6YM
L+qNappiHCE77UHDtDOe+KgIcj8mr98VizDxeI9pjYiKag1X4uN0LQou4Hm3QQfNkINmNENTKKYA
Gz2fduWl3u0qFjz8aXRj1gHL/UqtYwGOWxoOOmAeP8nYPQccknC+roYdNNVuGq0DEeIBs92jHjRs
u7QHfiXs8IhOS5s85kFTijYdVzSB1Xi58nQ+TaAEHzucVTRFbNMgq5HbRkV8Hy1EfH8giDfNBI+O
umggphLBpXfQlGCEEFFNz3fW5BGbXJ5BHKgTfZqmCjZtRMMbgZ/UMPnIY5yutIWYfwZpmGHnyb9U
wfi6WKGg66Ps/EoRIuzSN/KURzxdeL43Vw96oz8NrrTm3UC6W0YHog64wwHzOKc7X8q5T5cTI6sY
ypQ5IIZ1mVl+unn5aQ+IEx74ZWivcc1nFT6kwuNQ7ncDg6CZUTRFpfwgJZYm2Y4leEOyVMMc68MK
lvKsVrCUZ0TBGDdLBY+CusvGurQAb+QWddEouUGLJV6Ng3EH3OfBk+9OpnW+QXy0c9iDphz6zKOj
tBs8cDeZKXNATOm0lCeXYr1szoymX4U+Mqvrvlf8pSilvw35kGeHgiGTIsFjncRD80Q6gy8UQ4Rz
dSkQ5hitaVbhVdkJD5qtoJl20HBpz82H+UaIh55fjk/D1CsY9hB9CrCReZ8rrbF1aYlXMh/M54/3
W4sFj2uq8WiLGNbwRhyMG9NwKcrP5R3SEJbffNQwl3vy9QH2HHhTtmoXjfGLIhKvylbvgBsccIsD
jjrgDgccd8B9DjjtgHk0M/M+r0eZ6WsNE6/H+lEPGlgTMaFpOI5PedBUim+JWQ883/kN4YigW1Zb
8W5vxAO/Bh/76SMe5SHM44WEywCPabgEPKc80pbiI7WzmoZwMGja7ia8WJqFI4Tz/IrdQh57lO17
k+ibg2bIwWfUg8bGhzMnHDTTHjQFnNhyzRN5hcUjgscUlY7dJHjskTBpeDTR6FIL8Y4y1/bvxMHs
WhF14U1bD7rxmueQB34jPibM44v5trdK8Jgiy1OF3Pi6XL7cqkSQExHYopV8eUnBsp/WE3bJmeNp
g6YnHM3QF4m4xtOH73PwGfTgUwQdGNI0hEcy9AHBI4w5+cKJoEs6pmnY98c1XAYZ8iiikfOMK21t
fzneGF8jeMzQxVO+Iy+PGWp/qXQFaVS/th0wj1YafFjCufNBtnW1B57jJo8lUrYsM48uEuaYFVWw
lHOHZ9oDgkcU3fKnzeeRRvLhOHgin0a+gz+uaKRfNJlPI/2iaUWDblYqZjU97TMYu/MtWi0iAgOv
Ln+hCCtYvo9fnU8v7XO9pmF9G9w0uk15TJH6SflEAZt25BHF3P6ivm8Ql3jVRvIIom6vtAM/6IB5
dNEtQ7YXjzy68Tb60UimXqWCxxxZNvoPkw54yiPtUYz7Mx74Iuhex4X5eQWRVx/xsh2rxKCGafdG
8ukxLj8uxhRN0Rr4QtOaXrYXFyFc4zjlySOELD/H4lIFB4AVdgZfKKrz04I+ICJuvG6vep2WvkQD
4Izv5KI343iHxGvfyQH3OeC0RxnuFA/6eBzTa4ye9sDTV5nRdQziI+vBi5Rs16CP2AqWuhoBnPGT
iXfKzQFzrcfni4JiEqtqfSC8Y+cpTIbVC/MfBLgk1YnvnsYk3vECvfggwCP+bcCfhNX6gPMTfA4v
Ke8X4J4vCBzEBwPN8zCe34fP5JG/14v3HwS4zCXE8cAQI3x8/YMCHihmzMPHjHmgmDEPHNtYmyoX
eywfXmC5Cb8VYrfFdJbPFrWWz3dS+MRJ8Q++JyU30m+x+MQW11rrxV7LwufsG6wWO5fHTjzfaV2C
lJzWPI1fW9Rbp8DrdgyFTL8cr19aYjveAfDmqJ7fgecVSMmPQeMQgeRD3rutBZIr+dZaFWK7dT+e
p/GcdJS/Kf+PdLmfh5by3zIEW+y3UC9rCV5ppBTrY63tsYTd3ZOyD/X0drdvtykVlqxFlvNCTZFU
j+2ebjvRc68ta7Ilpyb1MlWjVRtptvwy506UZxpyPl4yhF8ld95T/vxINGOuLDLm6uEMcsVrLaIc
0j2I9qiCnJmO8j4FDcEna2PlqK8FuUTxxBYdsi3WiXZrnTgEOOpom/Xwb31oobshowOSz0tI0WWF
RLe8U3K6E+1si61Id8DKtvOdLl5sZ9XKue16p1UTiaOk6qL8Z7X8ydfIvyaiXsV4CfN/Z/6ncEy4
pqvLTvHTEvahzu7OZEesvQqlPAq+KdRTXT26/Y+XNEgU5cUXThhzJZTx1Dnl1wX5PAg53K/leb+W
02dR9nUijaeU0p06QxuQs16m/CwH5upoz9lMObjCyfyHdMyVSt6P65irkrznaqNs9wLd7jrmyiDx
DTrmqh7v0zoe1vGojvmih6yvjmd1HMSKGPF8CYQxXwRhzBdEzq1XheKWWKqjp91o1rCUzSnhF7c2
7oFlUPr2Falv1MOAwAeMY22pWCbF0xm58XmhqGlL9bZ2mZTPaH6FoinVmupNGvyIxgfErclYotuh
21/L4QePoDWZvLcnkcnv6znPl4p9vam2niMxm/0jmSnx85r/CnFDrDuWwGeW2+2bmvbttWt72u/L
UFGP2N62UPV22rUK8bDWFz7/iqNvKbtXIQYcz5/2eD7oeP41j+d47V/2b/L/usfzLzqeq/rklm9I
P78Wae9ED/2WtA2PZOTziMTb+PMHWZsdQl9W/cpG/WzUwRakexzhiwhPgs9JqxA25qkM3VPADwH/
guT/WAb/mOav2liVzcl/UFsNWzymoQXgezyT/rjF8YG2raGmqUm1g9ILJ6+T2LE5Kd5H+AnaLYIP
8OZTrsP+SEj8seTcjn4a13x319y4Z26+Z8Hzh/PmS/vG/k87EKQdCMg1bNnPgoW6/+k4omOuQrMf
xnXMlxF4P6JjvrjA+0nEys6fr38tE/tbE534vHjMbr4vHsv0p3HdFoxzx+cK8ZKUyymMgA+UHdWp
k2XbH6gQ38WTlwoqxPcQWxgJOO7b4irweJrjJPzBJ1Eu73G/QnxH6993PEYH1ucl8OGzV6TevJhp
9xdd/gH13+1XVIixDP33rAsx5tsoLf2ISoy+LDnbg/I0+DI5KjO/V2HX6UX4BSqpqQ1dQEAEXb2x
su0c5ceQ84tI8RpKyBRZ7g76FORctr2M9KQ13H2YjaqSmP6kyveQxr6EUlKuLOcU7DBj7mIYOZvx
UZXfEg89pLjxvgJ9mf7DCdAfD/BVTz0uYDVfji865gsIlLNzdDb+0SmkqYDs6R9Fdf4cD0jPmv4W
etNPZMvI/URZrtdl+09YX5J26XV4d6Tcjj5iSWo/OL6RaZc3rCdARz1hnvT4yvGU/iS9Sra9whOT
xbNMrA/LxfGMchnX8RhiyoflMPIx9aA8OB4eD8g9CimHUXjN0g9ArFrH6c9Oav3MeqkV4k20+Jta
/xYgzZuQHb1fWqHu3q6uBdAaWiVll4ipgHaY8pbqcgYR22LS+q+SbqHAHw/oirV240NC5MO7m+Uf
UGK63DxSid6YypeS3QzJ0Z5lS2GJu1vdOL+4DyObKRlzIBdTLtq7tJ/lXSAOtXYlY2gjlphy7ZD4
haKzOxU7HEvI8vl5p0va1dN9mGVeIJIdPfh2F0vN0qiewN8qtN1b8GhNbsynULT39MICMeVC0R5r
6zzS2iU5FkKGRw7GEornoa6eVsXTyOFcfJtQXvb5k+hX5eIvpQZRLqq/ZbHsJaxfN/s/YupFGvHx
wJScQ0m/Bwsvtmj3mfZmvgfkvOPAAVOfNzN6/CZsnclTUTpz86Ln/ICzmAPd2aeUmy1Oz4H/K098
BWrH8o/76FfKPSSlzyg/68FDLdJvxO4GY+78SP9Ox9ypkeOLjnk4Ro4vOubhmRlIInde8basOfFh
lDg73/hr3V/enmO+QbyXH8d5yNvwop3zkLf1HDF/HsK5Bucci8Qf4A8PPZj5OD7Lw5kIPQCTWtHm
YvjsHditKkGvn+08JcurfAj6+v8FFP9NWrZ3QWGeK18g9/l7mecLxd59zXbjrXtBn0szndGTd1Gy
9zB3tMU0fivE3+n8/062be1WW/y1LBs+bwgvyxbvKvw2N/49hb/KjZ9W+Ktz8cY/or7tdPFS9Vf5
7HI9YzrWWaZz5aXSqXLscj1jnzaykGld5WHtid/lwr8Bm8idAVqTndu277oacjLyeEv8d8ijGoGY
GTnv+qVVG/kH2cvz8ZxhvYHezAORcv6H+HmsIhfg3yKEZQgrEEoRLkVYhwodD8i9QdVvsNv1hvhH
i7soMyjRKnFbZ3d9KhWv0nFj7DO9mHdWbavaXC4+lLPtX2XaWfnFnC//CmVcKf/kA+tVJOp6YEi7
U1fQ9TqFla3WeLyrs6011dnT/Ym7k1gdkCmEn6zoU5v5Mn1rA5vxxfjcHDdVOssPcy3tNj8UQr88
d75NP50lKvDDXAjafe54GRzblBftPnfOiF8nCv0V8OlVe/M+gHt8XAcXx1/utj0HOXEHjuUg7Sn8
ubPGXZ+8dVdT813Un107d4BXI9Iu9tP2qtTK7vLA/vOQi2kXbqIdVx+nle3AHULaI36shfGMjplO
2jPEbKdx4NUcg+O3qnV2nYdlfkH6MeTH+nFHkvRqPDF+TBHFJ/2HZX6fzxZF/sdwz/gJxE9q+dhi
mb8cOOOnVIhCihN0hb5ceQf8xJWLYrRnhWPeFPTn0gX97A+VcsUq6Gf5uPNqi6BMr/zKHdI7bcAq
UQXwpp34go2FkrK8WX8wmKkvd3/ZPiOSn5IDx4sh3B8P8FMfev6g5cldTMqVO5iMuXtJ+XLHmv3g
ArG/tqkt0RlPVTXGDmMmj5FWaf9KP/V9pT8kxrSczP0aWVrOE9RzW99T9mVIUwJplkGaqp70iU9h
r/FAcuN2hAMP3H4geaDpjo3Xa7db40Nloeu3335n2YEDdzx44EDVho1lD15x/YH2TcAeqEK84foH
6eM8KB2aB+mGbaDkViFvG7+qJLzjPdtppX+9KEcpnpbjncFcDkyVOvMm9Vv5qKXArhch/FaI1Ui7
2k8ftly8jVZUtaKHZrAb0foGWyDnIXxiSQrqAVfNy8Wv4MeRm9LPGel/rEY7FvvZT6dwfwqWC3b1
rt37bt27E7TmZQ7p/+I5++MYYrb3qKRfIHY1Nu5r1LTS/wX+eIB7z6p9aR/lfEDHp6X9Vb4Oa85y
b5K/G+Xv9fL3TvlbLn9D8neD/L1d/t4hfx+Qvw/J3wexP1UMlaoQlVL+9KvK/EZvbWBVP6v06GeV
kHKLXSGqIOcyhCr0FZaP0PKMb1kGvu5+IOc/GAO4wj6iY3z69ssvzvWnvBp68LdljrR2c+ldNPZ2
Y9mnbKlYvHTxvjgMtOi2oe+w1gWdKV99A/7kVqtPHIVBT6bsJlG/a1fzXc27mtLNTXZNEog/7uo+
7HvGN5QUjyw/sFPUNNXX7qtp3OlbskV/Arco0W7dv6S6flfNzsiuxrsa993mu3JPWY/VvnXp4iXr
d98oGmE/d9Y01ywo2rBtxYb9NY031qRr9+y6q27fnt/Z5PvCRXX1NY3Nd+0trbllV/+KJUl8RVeY
z+gG8B3dxXfu7u1u8z08otao60RPV++R7lDtfZH9rV32vUnfevwVz8Q9WDONpVKVtn9Jx8Ml1o8K
T2y4rL/Ed//SxTb+xGtn4IjdFlzUm0jEru1O+bf5LvA3PXox/irZZ3rT78fafy9QtzYZb4rFum8M
fj/44wtg5/fUtSZjoeahROeR0A+v2LDh5c8v923c3ZMQG3fYm9N2c4+9LRw+tvyKmsK21zcuLq5L
XxG6VySr6mJdXcmQb+FnK+22DVVHW8OYYW+4ofTl4hsP2faeWHfosZINtrjODtvNHbHuih+ufnnx
v716h32j/3I7WnZ7mX35+KaItf2Ldr+1al16cdkdduQBf/+FL4td3e32jYfO/vZ1NQXi8h9f84WG
faHvH6p7bIF4s+1Hn9jV15myL9j15UcLf7Cm5tK9sb5UoZ1eWrdwVyJR1ShaO5Mx++jBfdGDd2O5
MlIwvmWTsLdtvrLSrrQX3/XNQNm7tVz2n5Lr/t+3nysva3zl8pFAXWLPofTa9wsfSx4rxAiM1f6y
0M8njj6/cbBq6WLfRT+8eKS7oPeg/QW7Cb8/HR/CCvY7gyH7pppL/tx31xuHnrvI5+9qjfze1LI/
GDl0cyyx7lMnD/d0re96zu//fqAplno5sUPcUNCZXJfEJPGeqk/tTiVDE3+4s7itfNGO3faiq6du
q7TLaorKfuvdgm+s2CGKdt4+fnNjz77WPZErxj9XvmhDVbQsNN7XdWt8Q9W4/29vLKmJ14y29n++
qqnta4mp5K3x9nFf97N/FMUfo0vG1i/o706ctW94/Fk73h5su/RnG8oPv3Pv//QtL+tOiCPHVr12
5KlOe+vfv3zlSImdeHRxzfafRX1F7x/8ZNT/nL9sZey17951yUgp//TK5ZfckjxcO9TTZ5eN19op
8fvLxV4s1498s72qzKq0m27sPtSTOPKJg+WX3Txe31/U8NhNTbFfdo1EY+01F629OfJHkb+dTvV0
/+2uptq3/ate3vu5Ezcfs75af/i9dN2xwo6iHQ/bnTvqWru6Yl/f5H92+/jWpo5WLGxZoRsKN1Q1
98T3zMYOpWqe//zuC7buDf901dme4kf9r969f9m7lU9c8NVqe21/v++IXML+5NVP/Li4N9HV4zti
N38yyIXe9wHGivrixZf8GaDWtmtTveOp5xN/U/LJQG+yFOvOP/3Cn/nseGtULjFfv3hqTc+4Wk7+
xZri+307ipJy0bqur6c99ovA39R9u2rzwkSq9d7JdHzkiVd8f+Tvse5Z56+Vs/97053/evHo9bfs
XrFpwYjv1jNNKz7ri/7+0eU/8dU8svwF3/eaiovOLHm2sH7FZ/0NLSt+/BPfvq4Vy6wbYg3fGHru
w/bYmhVWXXDflV+97bef333lz2cWN684c5td9pzvg9u+vfRYtOR3LbsMK/It1t/vvvBfAf6Bb9ZK
rTxjfTV+U1npmDVw0Z8B+Zzvn8asW79cesayl7wKgtcuvv6sNfmnT10ya/1q1fV+u2xHgWU/HB70
P7D6q7jrX2mpNfFfrPn5L/xfWvNqzQ77gWeaDl95yfu9h/3+zXXg47/0X1qP1ILNL6x/d8k/Lnh0
5ScW9XSd7S292V+486uFTxT+ANBXdl3/u/5+/6cu/qH/WPBTjTvs8drezq72xhV/unzl6O5Ez5Gh
vicfTTR/+nPXPvSVS95aYB/tW7D+vvv27Fg0dSbWNvUO/fJvFCwMNa2utMcvrvxPt/9O0PduSeXu
/ZV28vKmhQOLmm8PB3u6j8YS46mdWGxIDr33g6bK5p4/WXVZcyiwqNI+e1+JWPAXy7+7rO2aG5at
WPjB0vfTLZFnWn+0/rINf/xI/I1j1plDC+wnx617XjpZ2PTm0m8WTpVxsjj48vvWTJvYUHUjnPpE
Z2lPomr0e+LUDrtR3FAb2nxNdaU9seXKayIifFUk0B/7cNcrf/gEVmzOFLxZ8PBPw0V2Gb3jP1+Y
DPn9V317cduLojepeIlEJ3jV9XTBQONvTDbeUCtCW7Ztq7Q3XyOuqbS3hK/asLSABlqat8AhGJbV
MCdyrJIDk0jFbFr/lMDEwqYwRbuesQiqNWUqGnvuxWiwv1VwNOCgJfhnLDkehDgeVCZ8F+7pCXQf
3uALNOFPXxZ0H/YndnZiQPAvuHRhsVoHHu+Kcdtk+ZYFa1swp1m1Vnyp1wruv88X6k79brjftj79
hZvuXpZMHxWHvrDkUGciuSxlVaT39d9Cla/Gilf/cd/XHu5ATc2u0A/7g/0PPnptW0/XgrWF111w
LHhsefqCwjU77PpYkHut6bcxikVC9yb/F3tfAh7Fca1b3T09Gi2IEUhGCAytYRuBEJIQIIMxWjG7
hAQ2dhSjkTQayRZaZkYIvDFiMcR2bJAJ+H3+LosdP2NnQ861cXKJLWGWxBsoyU0cbvKM4uU6iRNL
Mc/Cvgl6/6nqnumRBGbJe1++993Wp5maqlOnTp2qOnXqVPUpyDk11tG/9exjU+KSrKmbx7bGKmMi
S5J8gYfyA/MCozGnBGhS8aZqy0cFaFLBnNIxTlHsmxfj3s77HI7WEyN8Dlh4HVsiWpcuna9BUrkL
l0QtZM9UdoxduLiklK1ay6fdPCamXW3VyQZp3EO507WMLSOlNSO3jT720HxtaSrm8ECBmMMDacrb
U3OHY9rCrJXjXEXz4OIRKSxFE5PXg+4crxR1gqmR0go27daLbuumXMtJu6TkeB6OV+du2jbi+LjF
8atvbI05Eam0z1/Afr3DfXz2NxM7hwVGr2j4m1975K7hY3JVF5uYfYz1TNYcqRPnnFC6mbolYnv6
s8qjEQ4HZkOLtiTHU+irdDVedC9NzVGYpTrfbetmf7EteHoZW/xh5GbtZOTTUqrWNy1l96gV0f7A
Q8nRzQm5N/0lpTKhMfmtyAdyLKcdb9/VQ/PAubtSxrmnuuXWP0g0R2qV7nn902Nvjd3d3rHKZVV+
NOnIwcj+7//oe3e8/tLPptY88+yPTiiP/LC15N9q/e51PudtypTWZxxPpPzwHnfuz5rdr71prUyd
WPNCi6Oho2X+Oy3R3x3T4XDY5tJcvOCjs69Hu9OfkTaPClhLG1+pq/UvLDo2rOPwvO2HcqX/Ueiq
rNGSrdr5xfW/Z74tts55k5KqI9+s2NOqprDIuY7yno+eV7rVA2NX59FW/B+dkyJSciIydj6hqRXJ
X3xocabLyVOz9i/IsL4W98UjlowULbJVvonNGPkrqf8P03KlL9jvU7tjE6afjmgd3Z+svDf1le1N
hb6c5qk7a7InvjZs6sGkbmnC/Wudh7f7+6QZnyV13/dloFLys8Rl+Reg9DxVGWfdNqLblxm/9uc3
lnqrbwwkuTy+1rHp6VqRV3vd8lfmqO/BOsTx2xtfjg1EKPL4LbYXE/ZP+/CGfLnUXdfcsV3rjdKy
Hm9SAHmx4rOHi6c8zRIzMuTWqCb5buczkf6eex29UV3DHJMznmbPR9/tSpM2ui/63o5pVTPYkmFP
q59F/W/p88K7Su5Sqnetujn6zeHJyoknYmsDP4DldlfsATYMQz39UBM31Vr9ZZafRZa4G+tchyrd
zt49qZKS9utbMvpUmGbTD3RLVW6Nm2UPwiK7dV3FzAgyxhZZjzbUO1LLbyyyFlm1GXGtG6Pya7zO
mVmdd//nSimrZeeYydpu6bMEacR7M5MXuG5xPWmpnO3vqtlzuK8hqzxr/ExNmrBLGWnrYh9+EVt2
cERZmTS8N/4T6b7j0R0ry3o3fcAsi4/K6w9UkBZ1UK5/z3J6xNg4uV8aNW59xQXl/agFc+yNSTvL
pPgXFgfG11YyrkVBa5HYz3W1OtFb5VwV1VTQ4lt1AXr8i3GTuz5TfyHlKtDtEk5LWp3L5x9T/u9d
0rL0abcrmJ4D2FQucdV7vG9v2cx+LTVitzh/Z2V+xfJHlGpXbd0PJK11lw2aHhYEp6WDxyR/g9/1
uKK9KH3a0VkYeNMakEo3R0XE5EdlQz1r/G5l3JmOmtrJvs7h2uJ70m6z+5ysVF8cXChdVfpBvE+R
RzYor38kuZUCsSr4W0nBASU/wh31WJfXdiC5PzfmxITnoj+Vu8aek58ZJimvfaaeSEr+sUP1//j8
+LTA2ExfR19zPev9V1tKWqs2ycnVuW5pyramM7efsNys+boKcj9Zldv71l9d2ogN5VNUZXhTXEry
FFtbstP/MgNn6MREGhNKt2mlwWgVAg28gdWJzfYUS6oW/W1iWYmloSVyR8owrIZgRGKcSw2sud6P
+Sm3kRlWLFafJiaRAFYVNAPJjmR5yeJqp0+uKsGdnR43xHZxbmJpqYMtl5aCwXl1Ny2leVBdWq9x
Jt+0fEWRxi2pqtPfsMzvemjF9nunadvvnGbVdsxVd7b40NKWmrTNLU5HXpbDkOyBvS0RKVuejUmY
ZUuwZrfFzz6yI956a1v8nEVR46YVLNa+ln93XEJh/NSuqIJvTHW8J9e1xa/Kaov/WZajcF1JglS2
oAjrO6+3gXk1uo8E5wE0knZvxOXXMKymimgVwKBs81VXYO0KrLpS0rDPnYZWL21o9gYq3QUuv0vb
Im+dO39pMxkwA8nS8wkl7mqv21ejbIk56by1AeufdDYsiubt0ma5ovUpPl37LW5t22+xvEQRzX43
6TMlO2BgZP1Obe37NkazdJ22zs3ooAXN1AGakOs9qW8qTuir6xWvEpXXYKna2BfBz0Kc0lISYjrv
xKGHmfZz4/K0LdvyoX+2xiyrwXZNcnzKlNgTdQU9mMprzo1NViZkTVKUaZPlmxe56qvq3LHeSZZz
sysj5mv5Xqtbcyo2pwNGz2GuKWkfsConWTwjDm5WUlpj04oa3fWRW1NStQOTU7WFrmysGZyW8pjj
EYvcjJ+LcuRj2awbPJ3YBNUOup57bsbddlg6HSdGGmu75q0pKVoHzY9F3s2y6uzeItMqrzK283nH
yO+ciy1PqmLDarUR2jYbFJNJEamtizpVafHiB6sfjT1085qYnGKbOsFe6PtofE1cfdVWRbN6hvn2
+Dvf2XzzfDkjrdQhj69SH1uT81zFmAMVjQ31Np971Ud+mnH6Tq7NHf2z+XPfj3xhGKupGN/aFOPg
psvyD1o7d8Nwibm0ewtLK3D7yOi1oPXs6w0frpNi6xs7bl7vDniFOsnez895pPKGBqgw53a4b6ma
0BDjWzxm+9ff8jU05/CzJSu0Ot+p7ZbEqvZsnzPliDwy7u/tM2tzpd5ZU6Sn5rVI799zavL87VJO
NBbwv1F/KA0bn+0dmz0xjdSHNucdw1JSDo9L1+6b3j8+YXzC4WU7lTvin7G9Pm6J4wcO6UtnxYhz
Sum5R/1f2laXz019bsHCAxW1jfKytjy2s8pZoaRo5Xc8uDpS2h4hyefkkthjsvNvtSkP7XbZquzt
0nuzJdmiaS9tKdzg97oWVlZP6ihrdju/iE29qFVgknxCKWUTrTvlA0n7i6zPySPlc1Fb2WQtd8mm
UzG75Gn5Lyn7lV22jpmOBY6sETEtllmWLkVq0Wr7J0VvGrdp3PpxZ2JmqjNykh/oPz95Ys5+db86
fuLpW4a3rB19v9v5zPriuHxrx5Ipmamvar+QFs04JT0+N6fe29YYMal+5cjueetc/sqah9wlOd+D
YPb7yzec8qi+ESsVr6tl84QfSQ2/kGbMmz4v4NYO26b7btuaV3psJRkm7SuUzsyMKK87DZesJ1Vo
2+c/31QuLfbUdzR43bBQxLWv8W07IC3PubG2rrZ+l5QN1e+c//XIhGHzNauDsZV3kO7krrVHuDf8
YW5KxpSRssNOlsd0BoMkbJPCIhnl2Ox0fEtyOhxOO1kjmQRrJBPmSMf9GjdINkplacmNjOyR5bb7
q1vbHxDWyIX39p+L8kTuZttS0+eck9p/nTT3Bwfyvrw3LqF57PzDCWkfPFW0VJlgz0z49zHztSrF
GUhPSYOOv/x17fEzLMHWLbctqPY7C8cnYlXQtbonzqHldvcUldR6at6Tukf+6fZd341YXlvlVDNT
K77ROcuaok3XMrUVGUp/Bhuev2nZ012q5Rex/7Z1hqI4lIOvCaPigpzRp6GPbsrdrz6hFpZcgO2w
WO2OPv1g5JEdzt9tme//4y3j59vnVxafaJvXwY+R1E881XtTU9FGrbLm9FzF0tPnzfG6NjodZa13
pq3OZdNy2dRctiCXfedwzl0OiP6JUo5ztZSTIuV8TdK+Lmn3qZL2gGS/34GZxcZtTqS1skY3GtG9
QcuZj9N9G/xq5kI7VkK1bPKyPJp0qpyVuCkq4Eshe9LqSG2k8rDUVlnD1uSX+n/vVaY6a1Owd5so
KVJ0NVdxKt15TlsCZrYadrRsDQY1i7TctII9prFaIYwXBmi1hMWStA/iWTdgD/4y3/+deUmoSyfg
/m9ujjTeMb8SFMsA9Kt5AtJc/sxrK1+isyDYaeN0XAmKHADdQhnwrMWB8DyWi+9imOiLcIS8kOUH
Dyp8NbYkGGlp74jeKeUvuV8BAfWA+d0IAViFzXrvNdTbyEKv7Jvf778SVCsBVEVE46njLtzHyPvl
YqDZJ5P/k/PSAcnC9vGYsyp54CBIme2ykteNcey0dET+GHCvKpTnj6qFe+mgNI9cA8g27gGiTiEf
M2PkNpl8OlxrOe9eRTmN11HOT6+inJrrKKfjKsopv45yXrlEOVOD7fMkb+vXeBuOUunugiPyuWAu
j1wFyL08pkkhryrUknSaJNSu5Hvjatr1Ces4JvrMUYV6lUeeDXz75cPAs1cmzuyVqR1EqXt5H9sr
9yBmlBqHF6yOyORNxCOP45RRDffymFEqvZpxRD7DUwXdFL4U3YHrpruc00198wjv5bac/TL3VGMZ
CUr28DBieHoiPimde6zR0ymcz+9f+w+pXT4AiPBRRMJhF48/Jb8F8eaRqcSLKrXDeDa4JTZcd40e
BP4/q+/g/1P8n2R/hauORLT10C1/teU9xiWHxp6wJus9wFzfkzK9zu/B0iO8po7/yzU9hZr2/z+t
qVFHkqf36RKYwngxXOoAD/wI8xNbFKk/xcFQaBansSJx368SI4mtcH+vEvdCLPE7HSR+gwOu2+WQ
NIYk7mNRopd59QcnH40yOI50/kn9WOb+lWTeQ2W2R58rz+vb3B6Z+ueTXBq8wsenB0eNGetTRyGW
ynpN2YnPEepw1OxF3o9f5TEG3Iwh4Z5EHgNOzCN2YK1TqLdRn6f0OmUR/7VfphL28dKM+Ypizqo0
3qhaNF8xhMV8JeQhQVDP4zAIe2QqoU9Nx6Gzw8Hf6BX97bqsIXnZgxwkpz4GRBuvTY9K5yPOsO0q
+T87Br545GrEfKESrQRN9BL8tdJ7LsiLcHoLOCWUeoTzhSQL1aZflywUXsqvcbmglnNOnJAPsKMK
xZ+S/yenVMhPwrFXJr6MUvHSF+Tnu5wDY1HCXh5eJvXwLrKbnZIP8JzlSGvjPAu1yrXW8mUu8UV7
C/o8cjzw7+Kcj5ZOyfQygUemA+Z7eev0wSpYDN1DtJSQFitMqZlI9fCZRcwgo9TpGE0ehXwyUUoH
vvvUmUEcf1KtXHOhdErzyKtM2LJM2IgzBzi9o1R6HWcgzhBdg3GaKZw1AOfQFFI9RC0HYzNzY7aJ
G3T+odFq9FrBG/F5Q7DvPst7azy6bjhX55jw0MgW3D7O9QMPl1Lrg7IqEa8bkBygWXagpLoFceIh
bRgPSroAzkn81hhZl0skHxRdFlHvk7jPRBpPg/IjjqQRZBn/pDaSuOcwhY9gEIX8Mr+HSg4eHjE0
W0NPICqPczmxl/fc43xcemQxkoiaI8ohfNJIIvlgjCQKL5VoPIdG0lKMJIo/haNXxEU7PvvUm4Cf
6unhEnG9bEj5MqSKEdQDiQK/kSqVM5ctthIOUZrQb/bLVE6/JRvHJPfw8HarTbNZcjgcjdRSK8kB
ykeQom5Pce1sP9eZDqrpiLfhmDnRdbepZGo6UfK8sJL38Hze1R3Eejw54mvQty78H9wg0jfpYDhA
KJ44IwLfobACiWI8MvdO9i38bOLHs/R8IQCWq6+GKCWOhID+tCeEwq3B6Upm3fy4D6UhrM9K9EvT
68LxBLNaWK+pq66kV1f1p9dEfC8NFP1JNsMEY3EFnN6xqdy4YF7EjzSAZJYx2oTfYBKiWk04O03h
9mBeZqoXY8n81VWBy2UKHzSF40gc6I/LFO4yhTtNNGTcGIJvMoUPmsJdpnCrZuKJw1QvE82u8aH4
XaZw05hQfLcpHGfCmSG8AQBQZq5g35BZZ1DVAZ8nmmg2hXvNbRRsF8a6TDDSpFDeXFO41RQ+aArH
TQ7Bd5nie03h7ikhmDhSe/RHIlVBf7onhMKdpvCuaaZ4U7jbHG/iYVxqCH6lKdxqCvea4NtN8d2m
cNx0Ex9M4VZTuN0U7jaF49JCeTNM4SZTuN0U7jWF49JNvMoKhZNnhMLtGSa+mcIZIUWV7TKFe03h
XJPJqMkUPmgKt5vKygjKE5mtnBUqt9cE3xmUAzJrzzbxf66pLjeHwq0mPN2mcBwtbvVnpSm8yxTu
No9lU1m5pnCTKdw5P4RTCk2XbOUCU110yxrFZJhhTOEmU3iXKdxuCneZwr2mcJyprAxTeFeuiT80
uepP3CJT2CSHk03wK03hJnN8YSjvLlO40xTuNYWTF5p4bgq3msLtpnC3KRx3ayhvrincZAqvNNHf
vthUryWhcLcp3mWK/1ifp7KCq6hxyGReRYXrJ2QlFWuOIwppvKSfPMm1BLIxiPB+OcBjJnCrA4XL
LcIOQtoCQVOcRybRLdYs+7nO0KOex0y5XW1EajXXGPYB4guVyjFWL6R1GLae/XINfoWvtijmrEpl
UN1ptUVl0WqrSbEhtF+mtZHQpkR4v/wujxHUUpioJd36VYWw/VGNCtJMGCg/QXnkodckdk75Gyhd
2I361MlB/bqN17MCWn+49o9D66CPKL3e2hKF41iRRFRCv8KnE7+MWv0UqQO196+qpdDXaboW66+9
3IJUgNXXEl5TUpJEvV406ZtNQX1TaPrUbs/yln0U8B75JD57OZW9EtHch7MlwrZmrJb61Kmcb8Sp
PnUaD5OuLda9xEOyAhE0eavtU4mHHmUnwoLvBkd7JeoN6GeyJXDBQtPwMb3fUOuel3bgiPmf1Jiw
NRdhJqyEzeAPXOMBntrI6AViFdJEqix/ElmjRONoDX4NXIWEpoegwqjbRainSawHn4puHTnDY6b/
jjR18eAtCv0ZvPIgWiXusRerEB6meimcWpl7cof7F8TI3Cu6sS4xxJ9dx1ur12KnSVc1yjRgBtbJ
pBIH6ftAz1+jfxurjhJA7Odr1Ysq9cII2CLEqH9PsbE7LZFwYEIvbe7jI+S89FMs057k8Kvh5uUY
71mfQHpQP+rAv1gV7+er8h41Ejl7JcIbwhoNrDHAmg0pYGAtxguSe3geGuPE81cVwniUr7U88m5Q
KdY+TYoZm7EDsYjvQFCOs3zVRNWW2G2WYaCQqNrHP3/D7S114OhtllikkDzbxz9/o66htgVv9kgv
8hXfJ1jtG3UiuD/i9yF8b8A/wQqrlrDajFdVshzI1BvGA47GEVFgttnZ2CFGvKepbmB7acFWCnZZ
3u8Ma11oSifc4iHaFL5SVri3a4VTJ3NP7TL3cy2zar0gshXSE97m6cBwUc0OtnkxQgPbnOKMNif4
8DanVIo17HQJ6BkX8FvsEhj2PBqdwo5vwN0wJNzOQXBkHxyMb8cguMQh4QKD4EYPCUctGk5f0pBw
1AvC4cYMCUejPBxu7JBw5YPgbhwSjnqbga8GrThwX4X4e637ZT1B3F+9L0ftcxqS3Invy1sgOgBR
av0JhxzKDkHpTdZ6WegYCzBjpfEZ6zbUTlhAnuLj9YysAXKGPJTFpYun7eE2GDE/NSnE0RNyyOri
5FhJooqSrg4rSSLqfcKGM9BaI0Z5InALi8rTukXlSz7O7Pg0W3po9JmtYuf4jur9iB0oC4gH4hlo
FaORJmQSNyfrUIYVn3gaHk9yEnMLPknkDzXPEXet3Me9sJMZ0mMRz/mCbsqBh0X+CL3qhEoUnwPE
XllD6LhuBxb657uIv379k5ZOe/h8Q9oh9W9DcxX7ZTSyqXWohuGaazVyCi3DsEYvYF+zFHMMBD2Q
NqFbUgkXVRq3+eyEnM/n5mn8Dp46bkejdMJxSiaLkWHHy7+EHe8Mep3QxY7L1GYemZb6e3m4Ty3A
DPCGXAE8RkwhYk7hvm3C/CtAiv70VZQuDFJqaLO3Ao+oq6DdwJMIGmg1wPDdbxE1pvBF9WNoJYs4
HtJepvFdHCMXaS+Uawe+jVwUvqg24nMxz7WG58rBJ/FJ6GHEJ8JOsKfkFbxWdtSqT12o8+ukTDZX
Yy4SaUuCvKRxdVL+nO8frDDxbSnnG7kDMfi2bFDM8kExKzhvP+fl0Qgw21eFTOhTi0BbhrEmsNzM
d45p3A9FRTEvoQDUGVSsHBRTMiimlFNxP6/TtVIh5I2f18HBewn1yn7LfL6Wo7DoB2KPQKxZiqRs
xPeoAl60v4CncJFErW5IOfNqinIRBKWb9yBDEszJOiSSC6QJD5Rgg7QV9HXip8RvfxH7jdB2eQyN
SWNnMrT4DZr5wnRlogoNyT+JMmPXkXoaNB7+SX1RZjfoatSTQV13BO9JRKmw7QvtTWiUhoX/Ui1z
C/rHAn0N1Kf3j49BhUfeaWqNM8HWMPj9iikGgOh7FCF63UX1AMILIOUIqkjagU+XZQM+cRqAh0nG
iXnNXOafoGlTCkGSRkq5CZ7ye+RL1SAHNZg6ZA0ulyP3qnPkDZlD8J44Fr5HpHFLAHHGznbwK0Vo
DhnYmyYTAH90Q0iwNxFGsT671PxGnJK4Riz6GHoMYmR+a4zM7tSLOjOkdkwcvajuxKdYER1CaKB2
THGGdkzw4doxpYp2KeHjNYQxwNdYV46xB5hojRXC6ATGPrU1qC9SiqFvUx1f1ddNAi6k95rhDpvg
XgC+L1Q7xXwq1moEaayu0hU630Uxv+GWgvWA3iPV8b0t2rmmEuuUHHzSKo/qGX4yjGLOqucwNqkh
Zfa4lWwHm7FC22U9jNAW06orVEuaURqRk/IRPEF6ZLJ3iFVXD1ZdjG3B3HmSz8rUnPfBukEWni94
y1E60Uc1DdFnrpk4uUYxZ3nNBH1Jwd144kgIA1kerqyGSeDDBkAfkIn6A/pK8wKwnZc05QHmVij1
GNdxzC1L2oxRZ4/sBTlPchivzuuLKnGZzvQIPot19ntK+Jqa6nNeOiQ9iFYTvCecd8JSKrPvS7ux
wn0Xv83lHlUI4tIl0tma6yuRpNTVlLj1ukv86SVKpNmwjXOgjcva4/hFsMdx9qSDc4FmojYeDu81
vdIa3h+EXNgGCg0bJ7WpGCH/wdf/xvgTmIjbl8P00BVjIi5eDtP2K8ZENb4cph1XgInM4eYeSf0u
nEff+Ifx6OGrokf018H0PPIP4/SjV0WP6M2D6fnmP6y9HrsCTGSfeoffx+qRH0JYSI/wFnt8AB43
n0m8XD4YMkDY9UTL7zTBH5I26RKnSyWJvYsdU/4XxkMbJP0TOMlHssiQcJTukWl8mG2FYmQNxkmc
6+IzCOE8o5I9UIyxowpJWGF7e0dGEGGS4iHbm43tkGiOvyLtAjlJSit8havo1t+hz5oQTQrXJRRG
0lzY4XBuheugVD+Z0Zwl6+76cKurrpd65DyUs5ef/xh81mQvP8N0nMt98xp7Lz81d1w/tyra8rd8
5qMtVLI29Km36edj+y2rcB5zD5/5/lV6hcu3o3xfyCMTNHZYQP3LCq3wPtZDn0o1XLrt49Bn1Nvh
kMUjk/51ke/brEFLi1XyHbzNSYKcUe/EnCJwDjwFJtbAoVNhk8nOw+lN5xSQVBQUkFSj0AW+Bqc5
26Dha5wGsfYRuzrnAGs+aeblGClNWK0Jj5G7DP3upPwwhxA4SBcNpYZWoTQniPguvqv1dWgkskJ0
GlwbCHEX7gA17EOkX4TsQ1SGWGcKDKK1+9S1oJBmb2OvYxHC+2T6PC+tYavRHyiVznJRD/wJ16de
5ecXB56qymYdvC9VoaSB+nJo63Wg/Yi4LM55KmEnP6kdDD26WBqkb6MMokvid3oq/PQV9i10Zbw8
2KNpd/JSPdoYneQe3ThJRb3uWk5S7UA54as34taVnaXKGHCWimi4+rNUZM0RZ6marOugFRH/jN/1
8nZuB0zlNSXLo9A0RPoZ2QnYS9kzKU3YM41dWcPyKHo+tZ+554v9TMF3SjPOvuMKNZzOpNFp7ISS
VmPO+QYfEWS7FOfo+/A6PfU7of2IkZITTBXjmPbwNODpUyv4fl4iL5FKMXLS2pQgKMWwGtwQHOnP
cr6I04RiRUilDTwJJ1JINl3qPOHH3HV74xA9P2Q/GLrnE1bjnDOVbT7/jLcp8UZPCrDyY4hskA0D
8dQ+uDqM56Q6DmUz3a2Pi0+ualz8kveWkBT3XEKKn0Ophjyi1vuU76CStKSUX6g1+knNgfG1evxf
BsDfjXiaFfvUTH3lPkqt5GdgN/B29fBZn1Kp9Wt469/Dewr1N9qvJrhG/FOacYKS0oTVNkRZnT5H
hOYdQfPPubRdBzldZxU1OyUfDnJjaOh6QNOcRuPO4Aa1/QW1GDE0e1zQzxlQ+QR1lpci2rbNSi3Y
ABxmC/JhxFG8R6ZdcOPsA/XuI9z+TPMkwfRbqoLz6kU1gIhMbgmlvOG5iBbKZeO53DwXhcWOm8jV
A+1gDj+PKuzMVALBnML6k6Qc7QWErNnUpc7gfDtB9aj0y7DVPqmXRVT0W6qZhpmfwkPXMDQTCtxi
pJ2UySVsuFW2McwqG5Kxdt5nmoKpxFnz7oZTov5/dbNTD+g1j0hFf1+B2s/Yfxe7p3yQmuY96oMS
XzVLjPqhOCEM93IID7XDTi5L6Qmd/BVSx2wdvB8S0rAV00FDsUOOl64RTsQNReKE7IKw3apSaEc7
UWZR2G4TxRKFPsRSwWIPimLLEetHLImMUOwixDbr7xCEYtMRux6xZNEIxSYitgWx1KKhWIbYDYgl
KRiKJVv/RsRSLUOxZEedjFiyuIZiX0HsvYiNCos9gNj7EAtPiCZYsr7ORizpoyEMjYi9H7HEOyNW
8HC/TJbafouXn0rqQLBa3sPjXtZ3kgbPAZm8BagnvCR9KIlZ9wx+iblXxNHMuwiY6aoOj5zKcxDM
y4qCUUh5P1KpZIm/j0gj+6hCvz0yzdJD7QuQzZfgSPaRXdl42+wC4gybLIWhqXE8Yq6jMp/llPyc
z7JAbhobdtjNqd8VI3ag5kY0i0ffcAtaOokKQ0Ojmsim3Tx9wkFGogKuAvmKg/hq3g0kF4v8AXPh
tU9/ZLif5E8sk11MliMMIIn16oOETEakYuCJZayUxevxmwxQmbXq5A4CNUAU1hs8W0NeH3Vs0p3A
ZsBYWFyQHzicL2A2mcmLkIPAKjtoOgkcBhQRrIOVxenIiTAyG+LZRAZRXhd5I4uwBoEjWJfpPLGZ
xBBCG4uzi9yEkFqbEIYxx6hNJGuNC4HSxM7LtH6TRfAHl1HyZ0SU6fz6JaodIiCaHTQdSL1EtWNY
6JR3qNrhNSrTaTdIvPJv7+qcAcDG74HxOsNNZ/wHv3lsBy4bTmX1BnHy3lVuY39D/x34EFAIEu5S
g/1YBGzoqO1BTBLkSzZrNXVjQJUz9vdLYBYZs1mcqTcj81fmyGHmNw0wGsstlywjdwDFovbmtwRQ
+BXWfiAuusLVXHtc+8LM5+UxAi+DOSQTYJMfQCW964sryMJO2WPGuQy2ThOG7kHY6C2/nLDz+5ip
wLVLtXlcsA0FKqImO+ydDkiOr2ynDWFvYEB8lCdessQMU4krTeHQiUKJHTTFd5nCuq9rdKcQtYKD
dlyQVhz27gfm18twMSMo2iXmMoWppxptYj6hjnn5CrHtGhJbDjOfZYeEuso2IS0k/KkQPzeFYqXy
FwMS976gsPQXlsg7rbihTF4UyTqejrcuUiBXoTXeLEexsbfl5RZ7Gw6Q61FVnpojR3VJ6xmbL98Q
IXul+CibdUSkPOL1B7LrnLi8Xl6CPDcwSbrJ56/CRRK3SPJouLIFsga8EYqr4+DcRmWT7mLK1LJb
78Pdm+mZ6Vkz06fPY7g/NJ9FybHSiPR0ljX7gQmZaekTWPqE/Lll8PmDaxpa4Opkoy8HrtxmZpYp
7kyW5q+rmFC0rJBpuc3+BvhnyYbXDBwakLYUFlXvqK6t3FpYxKoD7BAurdtcuC8QtzsgSa2FmQUL
swvSs1j+9Fl5C3OnZ7D0jLzpeQWFs54PFObmBvKzAmrmjkIp8KTH61qnLaxldW5fWX7DunWWhnqL
dXltJTwz+Rqq4Wq3xsXggLasaCEcV+UXZswuW15aVJ5WsGzZhO32Lcu1QMbsNDgSqmjthDPa2gqv
y7vxEVwtIQ3f/Eu4fN+s2N5oHL350KqaWh93d0heVJ23ssTAyFWB79QyG2I77vDCR/gLFR1SQ4eU
9E505tnop8fSNaCL8iKY7s7JlmBnFksNHDV5jthZjIU8R3mF/yRvURJiWL7XjTujuNso/4WHqcPC
UZTug1F45J0wxw6lArEArYJT+Fr42nvoNeBFXH3V2x/Yg3cltE2jMuBGCs6L4LsIHlG6Wuwsmulu
ichvyauanakWugbL9w5wSJbacjsQUCb4FGr22lm8hXv4qRQ+i5AFPqWa3Q/MByJLg/BZSq4cP4Nz
AxnZzt9lZxEW7n/H7VtnpR/kgQcuT2vP2NFv4IGHI1iwxY57+uBaR3jWcad/YYef6uUolFzruM9d
JPzcFaHw6fLzBqoheSDL+g6Qwu+juxI+r+54m9CQgxvy+OP2pT+G+jByZ+MfBl6rFvL44rvl20Rc
Zc1bZUQCeRqj9rsnEb8COBoBV9FrC9zVLhR+9zFi0Nq8tfBTuZbcKK1bRa2iR3AvxVO+R0XqMXBj
VzbPlGUZfMau7DZl4aTc9BtTFrjX+XMj6IHaKdVhjIuZdRzfYNBwe0R/vwPHevv7J0E8WvTUFHin
t7BpSLWy6UiN0ONF3lik4m4owA8xS4dF0a2bJGJJeR/4BGecYMJQMXakoiCwjR4bcHHCxR41RZFw
oVfCv5oYM2VUYbzhF6zA4KIZe6ORLl5YxWrgzQXjDUKKBlwFvhvwHQ9hksFm42LLLMxzc9gspKRz
Nl46D5i2TK8tzUUXIOD7+6PgqqaOuSAFMzjSTI50JrsJt6LPQmwWYFIga0NQ/f0jtwg0iYRm3gA0
mUzKyNtw54XAE+Nwz5WaeizZvnuZI882LU8azTq/Gcirc1Vl3Mqk3Q/nBe51Be7PYEnbrXsyH7UW
Jc0rThq5pTRpaubSpMxFSfTJUiqTslcvTYrrmF0M0ab7HWfkZhB+x29NYlJkMbtdemkdc6XjqppV
gbZcpYTd3Zyj1Oco3pwJbcMz8789vCj19a8zttzFUscnH1ub77LjaiSwXO9Bi6gWBCBaSJJC+muO
zq6B3/pU/qDxjUmHnuAcRt2Onjj9m76G6qzhMOogmFB+BUOBP5GdQd0Tmq5peWgPloVOKsI2Q+Vo
8emZbLRap6eGX8lJosyANfqg1wA11jrGXWb8IjQj0Vg1GYlmTGb/S+kmDlxpENIijBNXku8+AEXp
ywlz+RnXVr5E9SPmUqtdCYqvAUhcAELrdrPXpxa27ippiL+G+lejDH1UXqPXqRCR4D+95EBr+Suu
fy1gR+pd6avE8n+n///NAV1GRNP3A/jTuwWZBINhPnPwHobICMwMwcfgzSlc/rQrtNGIC2RDMIaB
RlYfZPG2CJssG2US0NzL4v0p4dUREMYOE17duhYr60YYWZZpEFyKxHKgaqWdPuMJ2dmYJuJiZRjN
4m1B+ojOMsxUBk/oexhf6yAQTb+iDEkfTUWL+Z5WJ/ALG4QjyWSkGDWnOD90hGbAmlvgbuQ0ftMc
shHpvmAMzQfVgKhDnDus5FqU50ecB//eMAwixVwuXLAD2hOG1QfNhXQWf1hd04JlkISlpVczv347
vOwqlFkJvQfTOijTZ75oylEP+HXIYaZJ1IFocKE0u2B8dBa+bweOeqh/fvw14qKW8N8lwNLE+UXL
wDSoO2lQgvQZMppsdiXw3LcSu+6FMGqugj+/hfDqtxhnROg0/vzL9rK3qJelC1LoM8fUNRwiOlZW
aqj3RuiclIzJvAi0ukG3/js6EfC4mERvD2qV6aBmI4cyqNUAA4+Q+KvjF6ETJ2p5nhm4+s3HQzqs
RLD0lKIUXOESxoVFvKcR/71G+dLYMPh8pLjRckQhlSH6jrkOb7leYjiS0T8SkfRuH/0wHlLsYd6F
Y4MyXEt3H3RW+svEfxa4uhA1EzHpYaF8PTaUakDRdxZU1gfYBHynAdMEqK4TwK+5KKEYtDagX3pB
5ToMyYWgl8aSD2nE0XX4I76GpyznHKScxLdq1FIDr+DgWK95me7f0fiew/uN+FUIXTwNvWMZ/iYA
jnT3Zs4jXL6Kbx/HTeXR2NDQn4ZqM1+QXxRIGoJb2eipM6+LWxmgkziVfkXcGoonsGDgDxYMcLSM
c7oBvCqDx8tQPK1RytCya8DvQvAkjYcLUeZQGAvZBo6tDpyhnGmgTgO+CvRhkgjUEstQYgVvUS/G
gPk5yobqVTOB43p6FYw5nEchPgkpQrILl/KgdqWggySIG71pJnpgGf9F6dTXKL8foQrUuYjLDmr1
ZsRR/wuNU3NNnMEWz0RfWogVXgHv5/moySz0I5JD08EhWgPmIZSH9EKkiHrmIlWMiyw9PpOPD6Ik
+4pb/HrGRxEoXAhJSW0uesBycKnINDK+qj9dS/uTZKE/CTPGCEyl8TAU3IB/WqYm4Xss/oXh4b/X
77y3XWL9TraFmUOYAaaZzACZWBlkrxZ9Np0rc9wM4IDIJbHp510bC3IubGmSaYaYrefTCqGfNQT6
VLiU+OrcaLrxVOxl1u81oOdaNWwyBvwzaefE23/21QIpeP9sNP59RdZzjl9/a/Ern8zunvDljk6i
j9a0n2z886st2z9ftnPOou0FqW2fG/Gd71o/mXpoelH7/uJtLW9+cKsRP+1Y5YatTXPynl4/Zcov
P5/4YxoyhIfSSS8L8CGwX3oKV1h25NGPzcyyNRFWVuiqsK4w7LcZv7mT7pjCDZXuugBs1UoMti/6
fk9mW+wiZMzufFAPzsyM2KQHZ2dtQFCxLHdV/vBFMqYjy+zvTdBDc76PUBShwS4IbYK8CqO21Sa2
N54op3BRNTYb3AnNsNHGkDnfsAbf/z4M0CCeW4ML6dYf2OETYbhVY7iBrmOyEczsRDAuBtf+Gbf+
CeNbaSTZrI1bO3BpRwGM08P1CP3OjpKCbf8FCi10fZ+4+ecvH8Gcb6Frg8S1fXSXRt9P7GyEHidu
EhIXCa1ACVEW/QI/3CRRWkAGfv2iPn5P3x0lZOlu8e1Zi9pwSze/qK8T5vQoS8jElX4P2for74St
f5ilku7pq/cLNK2fEc1ecU2tu0pE1n+Xaiay++jevkWwwauMm8Pn/oEb5mEqz/mUIvntIEnfpObi
Vyd9sBTNxWijA7sZMM3T/XiJ51EuwyV5/I48vgkS9S+0b7Ge3+DyAyruu06+JwjW4qo7fvOIr512
E/itIg0tLVupBt5a1CDCglufyNwW2EVmduM2Kr/v6AUiDfdV+Q6u5qThfh9/+nqKpLtPDvwnUYbk
b5wmC35u6E7dZ6fyLZZKL6qKW2Jwzy7uv4gG8hBRRBGyJmN7YLgeyyFxm4evpqLB5a2ygB1WRjfY
NWzYi2ahCosb6/i9vS/Wh3KikrjKB93VXUV4XwM9UZaKZj9ur6N6rTpLmPhVdd5/wcYL7sLmV9X9
181EN11Vh10H4rt8G+3diFvp9rXSMME1H1+8R42ntyjd4nXvTNoFcW9o5AVuOkR5XJX+ZledBAba
LM3/h72ri43rKsL7V7O1u86miRRFldB6ldLE9Ya9u/5PF8feH9utndh13QbFiZLGa5rGiS2vQy3s
NHUKAoGglFZCUR9QUVsJJAoWCgJFSInUByqhKkEKgT4g8UBfENAIpCIeEN/MOXPv3GuvlSjiAZFr
r/ae3ztnzs89882cHXYGcro6hZEZh/aGfdRNf/wBjTWl8aldfI7ItJ7q4Cjk2E/5QcdfgCMU1jh9
a5YG49ypb0AndD910ij7y9v6Dj0RwcmF2WeepzK4Fz9eQ7+yefuZosVLNmgcy9w3ZrNPWho/gtqJ
y49ZMn/xVxq/qPCgp5uqNW6DIo0iB6sAyTGvp4kPH/0lGdoRMy7jzHgXH2fo2OHP0RS0Wib2rGKy
/PqVZOiRGNx3wi2cqxKTBj9FvjyYq8bRyo+vgoMh4+ZtbmEfVHOYH3OzcwtwTABmaqg4/V0dQ8jy
EpRWCTcPK80uODqK4eep79DSRz6gX8FoSMTIWdoi+1h4dra6hP7kdksEF/ktumNLbLDqZqPY779L
nfk46jFKt6OY1vdzmJ988iJNaHbNdQzdHuPbm1+iAXKqSm5mzlb/MUPNm4BLn8Vv7qHJuXgKkzMc
qz1PcxQaUPiMOlGdaaZJB7XZE1jcYzzht+0ydyMzTz1Ms9k6Gnofa0gDAjShtrxP3WyGGjldqr7x
G7OY4PbDB4hS41kI8ztZpSC5YCrSdK/gJdEUM36YOCJ1gxpK3phMhosZLObeosHu76Dl/DvW+E+h
s1DF8MzLO1g1C1WiceizWIsMcV9STB/1E+ljyZ2P8ebz2nGepBwe/FmS34iELUAzLogQ4+mUMIGt
Ick8Z7FJtPhGmN6jdBGuUIM8TBI5oVG0eVxCXo2VFUIWEmyORk6HIvH4rRZbGl8a9mqDqKihHHLJ
MAERcgHi4zzDPQR8fAECyhLCAl+IWmMQAtscRDYFP4W32wcNo9QZFt+rEHBIpHfBq/CDNs8oGjjL
8AjBMbTvtfZdTW8hBwmIrQxQ0PcUYBB4S7Gi5BHE9MHk7ov8bALz0oH8uxGzG3l6Ueoop07BOzE8
quBvL8zTWxG3AlGwD+Fp+CU2eSnNhPcgZQW0GahwRcF/KwyvzYLyPS5HBJSSnbxwyuKizD8SzUmg
9fggvJpgeG+Uu/wEIBTq4M0Q1ZOAzeC+0b1eqgubxTdFZk+hnuu9Xj1XNqlHD69HN6VuiUC9x7xa
k5YZFLPLRDfjDQBQLx6JC2pJQ+IAZF+DHx4Emki/NyFDlYZoGeIRaYye9I30Vl+ozxc66gvt8oV2
+0J7fKEjvtCyL7SyadM/j6Zv3e81fUw13bb0fEpYwJB3XJpII+gAA3BVeSC7V6HcZV4kbA+4sU8y
tGK7rzkau4yJ3tAQ19pOO9HOB1aD1W0eidIBNN5MB6TAYmK/XicGABv6wx2BcGcg3OULlyDL6vIE
w+hwCZKuP90fLgEw9Kd7YRocVH8v+NQlxrnnc5bN0dgbxJeIjy8yPUXPW7S45eIGKzFh7wuMt8lK
LPaYBhmeQDdIjhJPYVeTEJbpT4vojF25n5NFztWD37Aa6CuG5Ip0SZK7nhZeg+oRnqyn4WGxXUbn
20bqng2+ApQW3LY/2lMy9cmTJb4SoMjGN7y61Rs49+7ujgOrKH4BHzL3+TI+X8GHzPnvXf8fHHjt
idZL5rfaDc5K9ihyJ7ogijPvZEJTzOXarQBhuW+DP+wQJ6/YCfuSLbM/8G1XGNcyx64q68K2mhft
OummS/m7/bZL0ovBes4Yet0VMm3p95Yfskn3rnGPPaE/9t04tPOTTz+++knvzalrZ4bCrF0jLZtc
nt1MLPTy148r+6NYyHHPLnhoHd3hV+TVM6JgSL1RSk9Zu4O811Xe8Nc0Nfg1drYGMHTL866pdktc
InRVre4t3o8+hcbV/aq6X1P3t9R9i7wS8ZRxdb+q7tfU/S2dX1H7qo9y+uVSj+xYqEVtNFVz+FZb
bgXTvMMvwRToddUv0QdT7fi2X64IBKTeXB42aIeJnI8JaTRQapETOAwBbhSZs+M5IuN1HSwoL9q/
xfwvYLthiwxYwmA7DdDS2jGLGTPcsFdOUoBNmcWSmQyZ0Q76dO118FXuIktmoU84juSkqbzRs833
RjNph2g4HHd5BSRX9dbYOsZLU7ZtvGdoIh6LWFSTWQtf2eZa10CvTQaUnVqYm1ucMvcw0C4fKpZH
9pYPlYN8Z8BYWramKJY5gnWAH+mdSgmF3lSrwnzdlmUDnWSnWoQ1LChlDNzFvt2Ytxt0OUcW7kKp
HRSwzyS7eiF13PutZHdJEd57x2HgXEHl+3ZdUkeu/TPzr8MPJX/5p9APQ6U/AGHkKyK/Fn+H44kM
48UuXg8maYrpFmnKqkTjmcJ1nBplCtZUWr+aqT9Z1xTbeLbySbtWLTQxaHdN9jJk00NWBtOy2W/a
hVRPiUbiswEWxgArEPxgrBLc3b07uzVKLuP4dxuPY0h25lLgsywWW7wkDdlKsqwyKGkRYlkfdtiC
dfDGYP0BeE6SZfkmoI5Rt+CDPQQvWGMAnha67DIkBvUK6rRtiTxiKb9N0FMq3m6LbXR8QGgTRitD
fzvjEtthfE0ndvyX9NyHG/dc9M3mj997+PVjxR/seKb2wL8vXZb8O+vk31l0tl94tKf/7fnlD776
53cHJf+DG0tL/NszRJFojWxjLbVLAWoDYpl3stJ9LUVv1hEICXfrZwMPYywjgiLFa8hQ4olRk4y8
LQBBIOMez4CM0sZ4QtVgoEFmNh4MKWLrPIuzhFhReTl7GY2+ThJ1fE2yoSaXRXVY+vtAvO1QHOPW
IuVVJV7KAhJgl84iD/3flJ+PqR2R7shRsHuRrQLdDmmi9EnAMiM++b/NtX0jROIgg4sn2F6mClhX
wKSau0oSKwcZcaYOJXyZEIUUPO+S1ckB3A2g9DTsdCxnw7JyeTCrTFLZbso2YzO4Mo61c03ZruLY
mnulzJ01APWhYWMY6xP426zqEQLcxPIOVV1RVVuUml8O2mAvxbg0GUnNMHRDAHwvQrZpbJCbACpu
hQ42QISpt+BuYZo5dA1bU0PCymn+BJlGpkGEj3tmocQ0MY7crFUNaNU4WVbKpTZjdRhGDw824ohq
FD2arEIJttZvWMcdHa3IodlU8zEoxfqGFNuvvYCUzcj/EcgPt3jU71fk275qjqBjpbejP595763v
DT+UvLyEXUv3Y5dzpmxztOEiVhqoL/wKDE3/lK/bCNy3Y7aJnvQ07xPI+vJZtjJLuRahNV++ZatK
oK6UvDXE9ULpIMOC5qCJ0curXwXR6+5Xkpx70dqfUrm0j+/nILmbixXi9iFymD2kT5PY5gq3QmQg
IXGyqzLqXYmV3S8UvVKxvFSVyleSZCqLijcYL8peiZdFwVPyBqnXpwuDaVpZHCROpVmifOd4rove
CYzLusNLzXhvyN3GnZbs6h4gC4VYyFPnukxMzo2hsbpOvLsNC607onWjzMOlQnq5UipXKk7PQKY/
W85n2jt78pnuUlc501UccLp78vlc2SmeSycaS3Mnzp6GPUdBt/qznxnK2kvl4AbXT8r5kkbnps/O
VgvrGJBoJBuFQtoTrEHEUHV2HvtFGr9EfBZRT1cXalCs4nAuDoSehHY6nyuk8z35XC4HypChODpY
SJe6S/1Fp9JV7NAfatbYQCHd1dnV3llxqFQln5U/pA4WC2mn3enMlgacfEd7Od9JnwEHSYnGw0Mw
1UlhH4rzsdWFFBlhHEk0uhxxCsv57nyu1NmezRQrPdmM4xQrme5yez6TzfZniz2Ok8129J/bhxfg
Ps1H1MzGJvPQa6NCze9Cti3l/hcTjczpjSJzwch1/C3ku9tS9HG6nVxbqqejoy2VUFDMXY+uexX8
tznwHwAAAP//AwBQSwMEFAAGAAgAAAAhAJSy2dgDAQAA9QIAACMAAAB4bC93b3Jrc2hlZXRzL19y
ZWxzL3NoZWV0MS54bWwucmVsc6ySwU7DMAyG70i8Q+Q7cTsQQmjpLhPSbgjGA4TUbSPaJErMYG9P
JDa0Tp249Bb7T/7/s5Xl6nvoxY5ist4pKGUBgpzxtXWtgrft080DiMTa1br3jhTsKcGqur5avlCv
OT9KnQ1JZBeXFHTM4RExmY4GnaQP5LLS+DhozmVsMWjzoVvCRVHcYzz1gGrkKTa1gripb0Fs9yEn
/+/tm8YaWnvzOZDjiQjcDf066q88XHbVsSVWICXWv710opcy3wWcZlrMyXQInwY6iKXMa7xEU85J
E6J1TPGVmPOW0ojqTMOzupTv1l2CvJsT0nDsn6MPI7pjM+Hx9Lc1HH3W6gcAAP//AwBQSwMEFAAG
AAgAAAAhANS1WSoLAQAAJQUAACMAAAB4bC93b3Jrc2hlZXRzL19yZWxzL3NoZWV0Mi54bWwucmVs
c8SUzWoDIRSF94W+w3D30ckkmZQSJ5sQyK6U9AFE7/zQGRW1afP2FZqmHZjaTcDd9R4854MjbrYf
Q5+d0LpOKwZzkkOGSmjZqYbBy3E/e4DMea4k77VCBmd0sK3u7zbP2HMfLrm2My4LLsoxaL03j5Q6
0eLAHdEGVVBqbQfuw9E21HDxyhukRZ6X1P72gGrkmR0kA3uQIf94NiH5f29d153AnRZvAyo/EUGF
t/2T1SZ4ctugZ0DIdemu05oEaqDTQIsEQEUMaJ0AqIwBFbcEOg39zvL38B5HpcmvnaM/ekHC/Fdr
81syXcKngS5itLIyQWWrWGWrBEDLGNAyAdDiG4iOPrfqEwAA//8DAFBLAwQUAAYACAAAACEADkT0
37wAAAAlAQAAIwAAAHhsL2RyYXdpbmdzL19yZWxzL2RyYXdpbmcxLnhtbC5yZWxzhI/NCsIwEITv
gu8Q9m7SehCRpr2I0KvUB1jS7Q+2SchGsW9voBcFwdOwO+w3O0X1mifxpMCjsxpymYEga1w72l7D
rbnsjiA4om1xcpY0LMRQldtNcaUJYzriYfQsEsWyhiFGf1KKzUAzsnSebHI6F2aMaQy98mju2JPa
Z9lBhU8GlF9MUbcaQt3mIJrFp+T/bNd1o6GzM4+ZbPwRocyAISYghp6iBinXDa+Sy/QsqLJQX+XK
NwAAAP//AwBQSwMEFAAGAAgAAAAhAOKvn3S5AQAANBUAACcAAAB4bC9wcmludGVyU2V0dGluZ3Mv
cHJpbnRlclNldHRpbmdzMS5iaW7slM9K41AUxr82jlY3KghuXIhLabGlcXSpNFEriSlJKt0Wm4GA
k5Q0RVRciHufwIfxEXwA1y5UXIsb/W6sKFKGDsxm4Nxw7vlzv54kP9JjI8QBEsTo0X4hxSIazENE
WZyyqioGtjBs5ca08Vu05rRSDuq6n4oLHfpptPJ5+lZe424hYLeUezK0y98VcwO58nma8q9c23VP
/9rJqO81l3CDolacnbx8fvjTXX5khytZr3/wiNLiPyTw8V2N8ug3FHm2v6u0M7jGKcpYh85/SRkV
7psowcRPVFkr0Qys8SpRU2XdZFRmrjOv0NeYVbGaZWfs6JqeYVloRmES9FTUaHeDxAtPAlim75su
nCQMorSdhnGEhuP67mbdhxv04sN+VmPodFVUQS0+jBM77gTv0fC3K84C+7phfzC4muouLVD6SNNo
LzmnoN8d2RdPEzvz16vn6v2twRkKnz2VVuXLA6/yDdq+ymdADjHnTR+/OQvUhGly7qip0ECbUQ9H
PE/Qofi70uFZNKK2xh7H6LK/x1+o+6mJlrImSwgIASEgBISAEBACQkAICAEhIASEgBAQAqMQeAMA
AP//AwBQSwMEFAAGAAgAAAAhAEfwmAadAAAAsAAAABoAAAB4bC9jdHJsUHJvcHMvY3RybFByb3Ax
LnhtbAyOMQ7CMAwAdyT+EHmnKUwU0VYCiZmhPCCkDg0kcRUb1P6erDfc3blfYlA/zOwptbCvalCY
LI0+vVp4DLfdERSLSaMJlLCFFRn6brs5O8rxSkkyhXtWxZK4hUlkPmnNdsJouIreZmJyUlmKmpzz
FjXPGc3IE6LEoA913ehGR+MTKHq+0cqwziV0+YpQYYHsZ8BFyhzo7g8AAP//AwBQSwMEFAAGAAgA
AAAhAEfwmAadAAAAsAAAABoAAAB4bC9jdHJsUHJvcHMvY3RybFByb3AyLnhtbAyOMQ7CMAwAdyT+
EHmnKUwU0VYCiZmhPCCkDg0kcRUb1P6erDfc3blfYlA/zOwptbCvalCYLI0+vVp4DLfdERSLSaMJ
lLCFFRn6brs5O8rxSkkyhXtWxZK4hUlkPmnNdsJouIreZmJyUlmKmpzzFjXPGc3IE6LEoA913ehG
R+MTKHq+0cqwziV0+YpQYYHsZ8BFyhzo7g8AAP//AwBQSwMEFAAGAAgAAAAhAEfwmAadAAAAsAAA
ABoAAAB4bC9jdHJsUHJvcHMvY3RybFByb3AzLnhtbAyOMQ7CMAwAdyT+EHmnKUwU0VYCiZmhPCCk
Dg0kcRUb1P6erDfc3blfYlA/zOwptbCvalCYLI0+vVp4DLfdERSLSaMJlLCFFRn6brs5O8rxSkky
hXtWxZK4hUlkPmnNdsJouIreZmJyUlmKmpzzFjXPGc3IE6LEoA913ehGR+MTKHq+0cqwziV0+YpQ
YYHsZ8BFyhzo7g8AAP//AwBQSwMEFAAGAAgAAAAhAEfwmAadAAAAsAAAABoAAAB4bC9jdHJsUHJv
cHMvY3RybFByb3A0LnhtbAyOMQ7CMAwAdyT+EHmnKUwU0VYCiZmhPCCkDg0kcRUb1P6erDfc3blf
YlA/zOwptbCvalCYLI0+vVp4DLfdERSLSaMJlLCFFRn6brs5O8rxSkkyhXtWxZK4hUlkPmnNdsJo
uIreZmJyUlmKmpzzFjXPGc3IE6LEoA913ehGR+MTKHq+0cqwziV0+YpQYYHsZ8BFyhzo7g8AAP//
AwBQSwMEFAAGAAgAAAAhAEfwmAadAAAAsAAAABoAAAB4bC9jdHJsUHJvcHMvY3RybFByb3A1Lnht
bAyOMQ7CMAwAdyT+EHmnKUwU0VYCiZmhPCCkDg0kcRUb1P6erDfc3blfYlA/zOwptbCvalCYLI0+
vVp4DLfdERSLSaMJlLCFFRn6brs5O8rxSkkyhXtWxZK4hUlkPmnNdsJouIreZmJyUlmKmpzzFjXP
Gc3IE6LEoA913ehGR+MTKHq+0cqwziV0+YpQYYHsZ8BFyhzo7g8AAP//AwBQSwMEFAAGAAgAAAAh
AEfwmAadAAAAsAAAABoAAAB4bC9jdHJsUHJvcHMvY3RybFByb3A2LnhtbAyOMQ7CMAwAdyT+EHmn
KUwU0VYCiZmhPCCkDg0kcRUb1P6erDfc3blfYlA/zOwptbCvalCYLI0+vVp4DLfdERSLSaMJlLCF
FRn6brs5O8rxSkkyhXtWxZK4hUlkPmnNdsJouIreZmJyUlmKmpzzFjXPGc3IE6LEoA913ehGR+MT
KHq+0cqwziV0+YpQYYHsZ8BFyhzo7g8AAP//AwBQSwMEFAAGAAgAAAAhAEfwmAadAAAAsAAAABoA
AAB4bC9jdHJsUHJvcHMvY3RybFByb3A3LnhtbAyOMQ7CMAwAdyT+EHmnKUwU0VYCiZmhPCCkDg0k
cRUb1P6erDfc3blfYlA/zOwptbCvalCYLI0+vVp4DLfdERSLSaMJlLCFFRn6brs5O8rxSkkyhXtW
xZK4hUlkPmnNdsJouIreZmJyUlmKmpzzFjXPGc3IE6LEoA913ehGR+MTKHq+0cqwziV0+YpQYYHs
Z8BFyhzo7g8AAP//AwBQSwMEFAAGAAgAAAAhAK0x1F2kAAAA2gAAABUAAAB4bC9wZXJzb25zL3Bl
cnNvbi54bWxkzb0OwjAMBOAdiXeovJO0DKiq+rMxMcIDRKnbRGrsKrZQeXuKGLue7r5rhy0txRuz
RKYOKlNCgeR5jDR38HreLzUUoo5GtzBhBx8UGPrzqV33DdMjihY7QdJBUF0ba8UHTE5Mij6z8KTG
c7I8TdGjlTWjGyUgalrstaxqq+EX4bi3EpIK/L1mO4i8Iu1fE+fkVAzn+eCVN5tcJLD9FwAA//8D
AFBLAwQUAAYACAAAACEA4h0v0EsBAABtAgAAEQAIAWRvY1Byb3BzL2NvcmUueG1sIKIEASigAAEA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAhJJfS8MwFMXfBb9DyYtPXdrOVQ1tByrDBweCE8W3
kNxtweYPSbTbtzdtt1qZ4GNyzv3lnEuK+U7W0RdYJ7QqUTpJUASKaS7UpkQvq0V8jSLnqeK01gpK
tAeH5tX5WcEMYdrCk9UGrBfgokBSjjBToq33hmDs2BYkdZPgUEFcayupD0e7wYayD7oBnCVJjiV4
yqmnuAXGZiCiA5KzAWk+bd0BOMNQgwTlHU4nKf7xerDS/TnQKSOnFH5vQqdD3DGbs14c3DsnBmPT
NJNm2sUI+VP8tnx87qrGQrW7YoCqgjPCLFCvbfUkIDwdXfjoQYMq8Ehq11hT55dh42sB/HZ/4j51
BHZXpX8AeBTCkb7KUXmd3t2vFqjKkiyPk1mcXa2SnMwykubvbYBf823Y/kIeYvxLDNCblng5Jels
RDwCqgKffJDqGwAA//8DAFBLAwQUAAYACAAAACEAobepTpsBAABGAwAAEAAIAWRvY1Byb3BzL2Fw
cC54bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACck0Fv2zAMhe8D9h8M3Ru5
2VAMgaxiSDb0sGEB4nZnRqZjobJkiKyR7NdPjlHHaQ8DdiP5Hp4+U5a6P7Yu6zGSDb4Qt4tcZOhN
qKw/FOKx/H7zRWTE4CtwwWMhTkjiXn/8oLYxdBjZImUpwlMhGuZuJSWZBlugRZJ9UuoQW+DUxoMM
dW0NboJ5adGzXOb5ncQjo6+wuummQDEmrnr+39AqmIGPnspTl4C1KgODK22LOlfy0qivXeesAU5f
r39aEwOFmrNvR4NOybmoEvUOzUu0fBoy5q3aGXC4TgfqGhyhkpeBekAYlrkFG0mrnlc9Gg4xI/sn
rXMpsj0QDpiF6CFa8JxwB9vYnGvXEUf9G+Pz3kGFXslkGIfncu6d1/azXp4Nqbg2DgEjSBKuEUvL
DulXvYXI/yI+M4y8I84GqNkHiNUccYItkZjewZ/3kTDeHLwObQf+lISp+mH9Mz12ZdgA4+uur4dq
10DEKl3PdBfTQD2kNUc3hKwb8AesXj3vheHPeBqfhb69W+Sf8nTps5mSlweg/wIAAP//AwBQSwEC
LQAUAAYACAAAACEA6Mes/xQCAAAHCwAAEwAAAAAAAAAAAAAAAAAAAAAAW0NvbnRlbnRfVHlwZXNd
LnhtbFBLAQItABQABgAIAAAAIQC1VTAj9AAAAEwCAAALAAAAAAAAAAAAAAAAAE0EAABfcmVscy8u
cmVsc1BLAQItABQABgAIAAAAIQBlULX/iQQAAOYKAAAPAAAAAAAAAAAAAAAAAHIHAAB4bC93b3Jr
Ym9vay54bWxQSwECLQAUAAYACAAAACEAc8li2DUBAAA+BAAAGgAAAAAAAAAAAAAAAAAoDAAAeGwv
X3JlbHMvd29ya2Jvb2sueG1sLnJlbHNQSwECLQAUAAYACAAAACEA1WH8EAEEAABoCwAAGAAAAAAA
AAAAAAAAAACdDgAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sUEsBAi0AFAAGAAgAAAAhALkLmJwZ
BwAAYyUAABgAAAAAAAAAAAAAAAAA1BIAAHhsL3dvcmtzaGVldHMvc2hlZXQyLnhtbFBLAQItABQA
BgAIAAAAIQB54NatxQcAABEiAAATAAAAAAAAAAAAAAAAACMaAAB4bC90aGVtZS90aGVtZTEueG1s
UEsBAi0AFAAGAAgAAAAhAGbonXKrAwAAYQ4AAA0AAAAAAAAAAAAAAAAAGSIAAHhsL3N0eWxlcy54
bWxQSwECLQAUAAYACAAAACEAjNzte8EDAAAgDwAAFAAAAAAAAAAAAAAAAADvJQAAeGwvc2hhcmVk
U3RyaW5ncy54bWxQSwECLQAUAAYACAAAACEAr24XWIUEAAA5CwAAGAAAAAAAAAAAAAAAAADiKQAA
eGwvZHJhd2luZ3MvZHJhd2luZzEueG1sUEsBAi0AFAAGAAgAAAAhAJI7gS+lAgAAXAUAABsAAAAA
AAAAAAAAAAAAnS4AAHhsL2RyYXdpbmdzL3ZtbERyYXdpbmcxLnZtbFBLAQItABQABgAIAAAAIQDP
b5agjgQAAFwLAAAUAAAAAAAAAAAAAAAAAHsxAAB4bC9jaGFydHMvY2hhcnQxLnhtbFBLAQItABQA
BgAIAAAAIQDWStfAdQQAAConAAAYAAAAAAAAAAAAAAAAADs2AAB4bC9kcmF3aW5ncy9kcmF3aW5n
Mi54bWxQSwECLQAUAAYACAAAACEA4Tir2SwDAAASFgAAGwAAAAAAAAAAAAAAAADmOgAAeGwvZHJh
d2luZ3Mvdm1sRHJhd2luZzIudm1sUEsBAi0AFAAGAAgAAAAhAPUKwQ/HZgAAAA4BABEAAAAAAAAA
AAAAAAAASz4AAHhsL3ZiYVByb2plY3QuYmluUEsBAi0AFAAGAAgAAAAhAJSy2dgDAQAA9QIAACMA
AAAAAAAAAAAAAAAAQaUAAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQxLnhtbC5yZWxzUEsBAi0A
FAAGAAgAAAAhANS1WSoLAQAAJQUAACMAAAAAAAAAAAAAAAAAhaYAAHhsL3dvcmtzaGVldHMvX3Jl
bHMvc2hlZXQyLnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhAA5E9N+8AAAAJQEAACMAAAAAAAAAAAAA
AAAA0acAAHhsL2RyYXdpbmdzL19yZWxzL2RyYXdpbmcxLnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAh
AOKvn3S5AQAANBUAACcAAAAAAAAAAAAAAAAAzqgAAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVy
U2V0dGluZ3MxLmJpblBLAQItABQABgAIAAAAIQBH8JgGnQAAALAAAAAaAAAAAAAAAAAAAAAAAMyq
AAB4bC9jdHJsUHJvcHMvY3RybFByb3AxLnhtbFBLAQItABQABgAIAAAAIQBH8JgGnQAAALAAAAAa
AAAAAAAAAAAAAAAAAKGrAAB4bC9jdHJsUHJvcHMvY3RybFByb3AyLnhtbFBLAQItABQABgAIAAAA
IQBH8JgGnQAAALAAAAAaAAAAAAAAAAAAAAAAAHasAAB4bC9jdHJsUHJvcHMvY3RybFByb3AzLnht
bFBLAQItABQABgAIAAAAIQBH8JgGnQAAALAAAAAaAAAAAAAAAAAAAAAAAEutAAB4bC9jdHJsUHJv
cHMvY3RybFByb3A0LnhtbFBLAQItABQABgAIAAAAIQBH8JgGnQAAALAAAAAaAAAAAAAAAAAAAAAA
ACCuAAB4bC9jdHJsUHJvcHMvY3RybFByb3A1LnhtbFBLAQItABQABgAIAAAAIQBH8JgGnQAAALAA
AAAaAAAAAAAAAAAAAAAAAPWuAAB4bC9jdHJsUHJvcHMvY3RybFByb3A2LnhtbFBLAQItABQABgAI
AAAAIQBH8JgGnQAAALAAAAAaAAAAAAAAAAAAAAAAAMqvAAB4bC9jdHJsUHJvcHMvY3RybFByb3A3
LnhtbFBLAQItABQABgAIAAAAIQCtMdRdpAAAANoAAAAVAAAAAAAAAAAAAAAAAJ+wAAB4bC9wZXJz
b25zL3BlcnNvbi54bWxQSwECLQAUAAYACAAAACEA4h0v0EsBAABtAgAAEQAAAAAAAAAAAAAAAAB2
sQAAZG9jUHJvcHMvY29yZS54bWxQSwECLQAUAAYACAAAACEAobepTpsBAABGAwAAEAAAAAAAAAAA
AAAAAAD4swAAZG9jUHJvcHMvYXBwLnhtbFBLBQYAAAAAHQAdAOgHAADJtgAAAAA=
"""


def _macro_write_embedded_runner_template(path: Path) -> Path:
    """Write the embedded macro runner template to *path* and return it."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(base64.b64decode(EMBEDDED_RUNNER_TEMPLATE_B64))
    return path


def _xml_text(value: Any) -> str:
    """Escape a value for XML text nodes while preserving valid newlines.

    DMN labels/descriptions can contain characters that are legal in JSON or
    copied text but illegal in XML 1.0. Excel repairs the whole worksheet when
    such a character is written to sheet XML, so strip those characters before
    escaping.
    """
    if value is None:
        return ""
    text = str(value)
    text = "".join(
        ch
        for ch in text
        if ch in "\t\n\r"
        or "\u0020" <= ch <= "\uD7FF"
        or "\uE000" <= ch <= "\uFFFD"
        or "\U00010000" <= ch <= "\U0010FFFF"
    )
    return html.escape(text, quote=False)


def _macro_cell_ref(row: int, col: int) -> str:
    return f"{excel_col_name(col)}{row}"


def _macro_cell_xml(row: int, col: int, value: Any, style: Optional[int] = None) -> str:
    """Return one SpreadsheetML cell using inline strings to avoid shared-string rewrites."""
    if value is None or value == "":
        if style is None:
            return ""
        return f'<c r="{_macro_cell_ref(row, col)}" s="{style}"/>'
    attr_s = f' s="{style}"' if style is not None else ""
    if isinstance(value, bool):
        return f'<c r="{_macro_cell_ref(row, col)}"{attr_s} t="b"><v>{1 if value else 0}</v></c>'
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f'<c r="{_macro_cell_ref(row, col)}"{attr_s}><v>{value}</v></c>'
    return (
        f'<c r="{_macro_cell_ref(row, col)}"{attr_s} t="inlineStr">'
        f'<is><t xml:space="preserve">{_xml_text(value)}</t></is></c>'
    )


def _macro_row_xml(
    row_num: int,
    values: Sequence[Any],
    style_by_col: Optional[Dict[int, int]] = None,
    height: Optional[str] = None,
) -> str:
    style_by_col = style_by_col or {}
    cells = [_macro_cell_xml(row_num, i, value, style_by_col.get(i)) for i, value in enumerate(values, start=1)]
    ht = f' ht="{height}" customHeight="1"' if height else ""
    return (
        f'<row r="{row_num}" spans="1:{len(values)}"{ht} x14ac:dyDescent="0.35">'
        + "".join(cells)
        + "</row>"
    )


def _macro_build_outcome_names(expected: Any) -> str:
    if not expected:
        return ""
    names: List[str] = []
    for part in str(expected).split(","):
        if "=" in part:
            name = part.split("=", 1)[0].strip()
            if name and name not in names:
                names.append(name)
    return ",".join(names)


def _macro_build_variable_types(request_body: JsonObj) -> str:
    variables = (request_body or {}).get("variables", {}) if isinstance(request_body, dict) else {}
    parts: List[str] = []
    for name, spec in variables.items():
        if isinstance(spec, dict) and spec.get("type"):
            parts.append(f"{name}:{spec.get('type')}")
    return ";".join(parts)


def _macro_build_controls_xml(n: int, run_col_zero_based: int) -> str:
    """Build the x14 form-control list for the Tests sheet.

    The returned fragment intentionally does NOT include the outer
    mc:AlternateContent wrapper. Excel expects the worksheet tail to be:

        <drawing .../><legacyDrawing .../>
        <mc:AlternateContent><mc:Choice Requires="x14"><controls>...

    Putting <controls> directly after pageMargins makes Excel repair/replace
    sheet2.xml on some workbooks.
    """
    parts = ["<controls>"]
    for i in range(n):
        shape_id = 2049 + i
        rid = 3 + i
        row0 = 4 + i  # zero-based row for Excel row 5+i
        parts.append(
            f'<mc:AlternateContent>'
            f'<mc:Choice Requires="x14"><control shapeId="{shape_id}" r:id="rId{rid}" name="Button {i + 1}">'
            f'<controlPr defaultSize="0" print="0" autoFill="0" autoPict="0" macro="{RUN_SELECTED_MACRO}">'
            f'<anchor moveWithCells="1" sizeWithCells="1"><from><xdr:col>{run_col_zero_based}</xdr:col>'
            f'<xdr:colOff>0</xdr:colOff><xdr:row>{row0}</xdr:row><xdr:rowOff>63500</xdr:rowOff></from>'
            f'<to><xdr:col>{run_col_zero_based + 1}</xdr:col><xdr:colOff>6350</xdr:colOff>'
            f'<xdr:row>{row0}</xdr:row><xdr:rowOff>368300</xdr:rowOff></to></anchor></controlPr>'
            f'</control></mc:Choice></mc:AlternateContent>'
        )
    parts.append("</controls>")
    return "".join(parts)


def _macro_build_vml(n: int, run_col_zero_based: int) -> str:
    parts = [
        '''<xml xmlns:v="urn:schemas-microsoft-com:vml"
 xmlns:o="urn:schemas-microsoft-com:office:office"
 xmlns:x="urn:schemas-microsoft-com:office:excel">
 <o:shapelayout v:ext="edit">
  <o:idmap v:ext="edit" data="2"/>
 </o:shapelayout><v:shapetype id="_x0000_t201" coordsize="21600,21600" o:spt="201"
  path="m,l,21600r21600,l21600,xe">
  <v:stroke joinstyle="miter"/>
  <v:path shadowok="f" o:extrusionok="f" strokeok="f" fillok="f" o:connecttype="rect"/>
  <o:lock v:ext="edit" shapetype="t"/>
 </v:shapetype>'''
    ]
    for i in range(n):
        shape_id = 2049 + i
        row0 = 4 + i
        margin_top = 69.5 + i * 70
        margin_left = 3060
        parts.append(
            f'''<v:shape id="_x0000_s{shape_id}" type="#_x0000_t201" style='position:absolute;
  margin-left:{margin_left}pt;margin-top:{margin_top:g}pt;width:69.5pt;height:24pt;z-index:{i + 1};
  mso-wrap-style:tight' o:button="t" fillcolor="buttonFace [67]" o:insetmode="auto">
  <v:fill color2="buttonFace [67]" o:detectmouseclick="t"/>
  <o:lock v:ext="edit" rotation="t"/>
  <v:textbox style='mso-direction-alt:auto' o:singleclick="f">
   <div style='text-align:center'><font face="Aptos Narrow" size="220"
   color="#000000">Run</font></div>
  </v:textbox>
  <x:ClientData ObjectType="Button">
   <x:Anchor>
    {run_col_zero_based}, 0, {row0}, 10, {run_col_zero_based + 1}, 1, {row0}, 58</x:Anchor>
   <x:PrintObject>False</x:PrintObject>
   <x:AutoFill>False</x:AutoFill>
   <x:FmlaMacro>{RUN_SELECTED_MACRO}</x:FmlaMacro>
   <x:TextHAlign>Center</x:TextHAlign>
   <x:TextVAlign>Center</x:TextVAlign>
  </x:ClientData>
 </v:shape>'''
        )
    parts.append("</xml>")
    return "".join(parts)


def _macro_build_drawing(n: int, run_col_zero_based: int) -> str:
    parts = [
        '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" xmlns:a14="http://schemas.microsoft.com/office/drawing/2010/main" xmlns:a16="http://schemas.microsoft.com/office/drawing/2014/main">'''
    ]
    for i in range(n):
        shape_id = 2049 + i
        row0 = 4 + i
        cid = f"{{00000000-0008-0000-0100-{i + 1:012X}}}"
        parts.append(
            f'''<mc:AlternateContent><mc:Choice Requires="a14"><xdr:twoCellAnchor><xdr:from><xdr:col>{run_col_zero_based}</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>{row0}</xdr:row><xdr:rowOff>63500</xdr:rowOff></xdr:from><xdr:to><xdr:col>{run_col_zero_based + 1}</xdr:col><xdr:colOff>6350</xdr:colOff><xdr:row>{row0}</xdr:row><xdr:rowOff>368300</xdr:rowOff></xdr:to><xdr:sp macro="" textlink=""><xdr:nvSpPr><xdr:cNvPr id="{shape_id}" name="Button {i + 1}" hidden="1"><a:extLst><a:ext uri="{{63B3BB69-23CF-44E3-9099-C40C66FF867C}}"><a14:compatExt spid="_x0000_s{shape_id}"/></a:ext><a:ext uri="{{FF2B5EF4-FFF2-40B4-BE49-F238E27FC236}}"><a16:creationId id="{cid}"/></a:ext></a:extLst></xdr:cNvPr><xdr:cNvSpPr/></xdr:nvSpPr><xdr:spPr bwMode="auto"><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom><a:noFill/><a:ln w="9525"><a:miter lim="800000"/><a:headEnd/><a:tailEnd/></a:ln></xdr:spPr><xdr:txBody><a:bodyPr vertOverflow="clip" wrap="square" lIns="36576" tIns="36576" rIns="36576" bIns="36576" anchor="ctr" upright="1"/><a:lstStyle/><a:p><a:pPr algn="ctr" rtl="0"><a:defRPr sz="1000"/></a:pPr><a:r><a:rPr lang="en-NL" sz="1100" b="0" i="0" u="none" strike="noStrike" baseline="0"><a:solidFill><a:srgbClr val="000000"/></a:solidFill><a:latin typeface="Aptos Narrow"/></a:rPr><a:t>Run</a:t></a:r></a:p></xdr:txBody></xdr:sp><xdr:clientData fPrintsWithSheet="0"/></xdr:twoCellAnchor></mc:Choice><mc:Fallback/></mc:AlternateContent>'''
        )
    parts.append("</xdr:wsDr>")
    return "".join(parts)


def _macro_build_sheet2_rels(n: int) -> str:
    rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    parts = [
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n<Relationships xmlns="{rel_ns}">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing2.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" Target="../drawings/vmlDrawing2.vml"/>'
    ]
    for i in range(n):
        rid = 3 + i
        cp = 2 + i  # ctrlProp1 belongs to the Dashboard Run All button.
        parts.append(
            f'<Relationship Id="rId{rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/ctrlProp" Target="../ctrlProps/ctrlProp{cp}.xml"/>'
        )
    parts.append("</Relationships>")
    return "".join(parts)


def _macro_update_content_types(xml_bytes: bytes, n: int) -> bytes:
    root = ET.fromstring(xml_bytes)
    ns = "{http://schemas.openxmlformats.org/package/2006/content-types}"
    existing = {el.attrib.get("PartName") for el in root.findall(f"{ns}Override")}
    required = [
        ("/xl/worksheets/sheet1.xml", "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"),
        ("/xl/worksheets/sheet2.xml", "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"),
        ("/xl/drawings/drawing2.xml", "application/vnd.openxmlformats-officedocument.drawing+xml"),
        ("/xl/vbaProject.bin", "application/vnd.ms-office.vbaProject"),
    ]
    for cp in range(1, n + 2):
        required.append((f"/xl/ctrlProps/ctrlProp{cp}.xml", "application/vnd.ms-excel.controlproperties+xml"))
    for part_name, content_type in required:
        if part_name not in existing:
            ET.SubElement(root, f"{ns}Override", {"PartName": part_name, "ContentType": content_type})
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _macro_update_workbook_xml(xml_bytes: bytes, last_col_letter: str, last_row: int) -> bytes:
    txt = xml_bytes.decode("utf-8")
    sheets_xml = '<sheets><sheet name="Dashboard" sheetId="2" r:id="rId1"/><sheet name="Tests" sheetId="1" r:id="rId2"/></sheets>'
    txt = re.sub(r"<sheets>.*?</sheets>", sheets_xml, txt, flags=re.S)
    defined = (
        f'<definedNames><definedName name="_xlnm._FilterDatabase" localSheetId="1" hidden="1">'
        f'Tests!$A$4:${last_col_letter}${last_row}</definedName></definedNames>'
    )
    if re.search(r"<definedNames>.*?</definedNames>", txt, flags=re.S):
        txt = re.sub(r"<definedNames>.*?</definedNames>", defined, txt, flags=re.S)
    elif "<calcPr" in txt:
        txt = txt.replace("<calcPr", defined + "<calcPr", 1)
    else:
        txt = txt.replace("</workbook>", defined + "</workbook>", 1)
    return txt.encode("utf-8")


def _macro_update_app_xml(xml_bytes: bytes) -> bytes:
    """Best-effort update of document properties; stale app.xml is harmless but confusing."""
    try:
        txt = xml_bytes.decode("utf-8")
        txt = re.sub(r"<vt:i4>\d+</vt:i4>", "<vt:i4>2</vt:i4>", txt, count=1)
        txt = re.sub(
            r"<TitlesOfParts>.*?</TitlesOfParts>",
            '<TitlesOfParts><vt:vector size="2" baseType="lpstr"><vt:lpstr>Dashboard</vt:lpstr><vt:lpstr>Tests</vt:lpstr></vt:vector></TitlesOfParts>',
            txt,
            flags=re.S,
        )
        return txt.encode("utf-8")
    except Exception:
        return xml_bytes


def _macro_build_dashboard_xml(test_cases: List[JsonObj], analysis: JsonObj, postman_path: Optional[Path] = None) -> str:
    '''Build a dashboard that keeps the macro status cells intact and restores
    the original per-DMN decision summary table.

    The VBA runner writes totals to B4:B7 and D5:E7, so those cells must stay
    reserved. The generated model metadata and the per-partial-DMN table are
    placed below them to avoid breaking Run All / UpdateDashboard.
    '''
    counts = Counter(tc.get("decisionId", "UNKNOWN") for tc in test_cases)
    decisions = analysis.get("decisions") or {}
    metadata = analysis.get("metadata") or {}
    uncovered = analysis.get("uncoveredConditions", []) or []

    def decision_sort_key(item: Tuple[str, Any]) -> Tuple[int, str]:
        preferred = [
            "BehaalbareHoogteSubsidie",
            "BerekenBasisHoogteSubsidie",
            "BerekenBeschikbaarSubsidiePlafond",
            "SubsidieConstantenThuisbatterij",
            "jaarGebondenBudget",
        ]
        key = item[0]
        try:
            return (preferred.index(key), key)
        except ValueError:
            return (len(preferred), key)

    rows: List[str] = []
    rows.append(_macro_row_xml(1, ["DMN MC/DC Test Generation Summary", None, None, None, None, None, None, None], {1: 2}, height="23.5"))
    rows.append(_macro_row_xml(
        2,
        [
            "Boundary-focused MC/DC cases generated per DMN decision table. Enable macros, then use Run All Tests or row buttons for real POST execution.",
            None, None, None, None, None, None, None,
        ],
        {1: 3},
    ))

    # Keep these cells compatible with the embedded VBA UpdateDashboard macro.
    rows.append(_macro_row_xml(4, ["Total", len(test_cases), None, "RUN ALL TESTS", None, None, "Result", "Count"], {1: 1, 2: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1}))
    rows.append(_macro_row_xml(5, ["Passed", 0, None, None, None, None, "PASS", 0], {1: 1, 2: 1, 7: 1, 8: 1}))
    rows.append(_macro_row_xml(6, ["Failed", 0, None, None, None, None, "FAIL", 0], {1: 1, 2: 1, 7: 1, 8: 1}))
    rows.append(_macro_row_xml(7, ["Not run", len(test_cases), None, "OPEN TESTS SHEET", None, None, "NOT RUN", len(test_cases)], {1: 1, 2: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1}))

    postman_hint = postman_path.name if postman_path else "<generated collection>"
    rows.append(_macro_row_xml(9, ["Metric", "Value", None, "Newman command", f"newman run \"{postman_hint}\"", None, None, None], {1: 1, 2: 1, 4: 1}))
    metric_rows = [
        ("Source DMN", metadata.get("sourceDmn", "")),
        ("Algorithm", metadata.get("algorithm", "")),
        ("Decision count", metadata.get("decisionCount", len(decisions))),
        ("Selected test cases", len(test_cases)),
        ("Uncovered conditions", len(uncovered)),
    ]
    for row_num, (label, value) in enumerate(metric_rows, start=10):
        rows.append(_macro_row_xml(row_num, [label, value, None, None, None, None, None, None], {1: 1}))

    decision_start = 17
    rows.append(_macro_row_xml(decision_start, ["Decision ID", "Decision name", "Hit policy", "Candidate count", "Selected cases", None, None, None], {1: 5, 2: 5, 3: 5, 4: 5, 5: 5}))
    current_row = decision_start + 1
    for decision_id, details in sorted(decisions.items(), key=decision_sort_key):
        rows.append(_macro_row_xml(
            current_row,
            [
                decision_id,
                details.get("decisionName", ""),
                details.get("hitPolicy", ""),
                details.get("candidateCountEvaluated", ""),
                counts.get(decision_id, details.get("selectedTestCaseCount", 0)),
                None, None, None,
            ],
            {},
            height="32",
        ))
        current_row += 1

    if current_row == decision_start + 1:
        rows.append(_macro_row_xml(current_row, ["No decision details available", None, None, None, len(test_cases), None, None, None], {}))
        current_row += 1

    last_row = max(current_row - 1, 17)
    cols_xml = "".join([
        '<col min="1" max="1" width="34" customWidth="1"/>',
        '<col min="2" max="2" width="36" customWidth="1"/>',
        '<col min="3" max="3" width="18" customWidth="1"/>',
        '<col min="4" max="4" width="20" customWidth="1"/>',
        '<col min="5" max="5" width="34" customWidth="1"/>',
        '<col min="6" max="6" width="22" customWidth="1"/>',
        '<col min="7" max="8" width="18" customWidth="1"/>',
    ])
    merge_xml = (
        '<mergeCells count="5">'
        '<mergeCell ref="A1:H1"/>'
        '<mergeCell ref="A2:H2"/>'
        '<mergeCell ref="D4:F5"/>'
        '<mergeCell ref="D7:F8"/>'
        '<mergeCell ref="E9:F9"/>'
        '</mergeCells>'
    )
    auto_filter_xml = f'<autoFilter ref="A{decision_start}:E{last_row}"/>' if last_row >= decision_start else ""
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" mc:Ignorable="x14ac xr xr2 xr3" xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac" xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision" xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2" xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3" xr:uid="{{5D4A7AAF-31BC-43F8-B4A7-32D52E407D0C}}"><sheetPr codeName="Blad1"/><dimension ref="A1:H{last_row}"/><sheetViews><sheetView topLeftCell="A1" workbookViewId="0"><pane ySplit="16" topLeftCell="A17" activePane="bottomLeft" state="frozen"/><selection pane="bottomLeft"/></sheetView></sheetViews><sheetFormatPr defaultRowHeight="14.5" x14ac:dyDescent="0.35"/><cols>{cols_xml}</cols><sheetData>{''.join(rows)}</sheetData>{auto_filter_xml}{merge_xml}<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/><pageSetup orientation="landscape" r:id="rId1"/><drawing r:id="rId2"/><legacyDrawing r:id="rId3"/><mc:AlternateContent xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"><mc:Choice Requires="x14"><controls><mc:AlternateContent xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"><mc:Choice Requires="x14"><control shapeId="1025" r:id="rId4" name="Button 1"><controlPr defaultSize="0" print="0" autoFill="0" autoPict="0" macro="{RUN_ALL_MACRO}"><anchor moveWithCells="1" sizeWithCells="1"><from><xdr:col>3</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>3</xdr:row><xdr:rowOff>0</xdr:rowOff></from><to><xdr:col>6</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>5</xdr:row><xdr:rowOff>0</xdr:rowOff></to></anchor></controlPr></control></mc:Choice></mc:AlternateContent></controls></mc:Choice></mc:AlternateContent></worksheet>'''


def _macro_build_dashboard_vml() -> str:
    return f'''<xml xmlns:v="urn:schemas-microsoft-com:vml"
 xmlns:o="urn:schemas-microsoft-com:office:office"
 xmlns:x="urn:schemas-microsoft-com:office:excel">
 <o:shapelayout v:ext="edit">
  <o:idmap v:ext="edit" data="1"/>
 </o:shapelayout><v:shapetype id="_x0000_t201" coordsize="21600,21600" o:spt="201"
  path="m,l,21600r21600,l21600,xe">
  <v:stroke joinstyle="miter"/>
  <v:path shadowok="f" o:extrusionok="f" strokeok="f" fillok="f" o:connecttype="rect"/>
  <o:lock v:ext="edit" shapetype="t"/>
 </v:shapetype><v:shape id="_x0000_s1025" type="#_x0000_t201" style='position:absolute;
  margin-left:260pt;margin-top:53pt;width:510pt;height:47pt;z-index:1;
  mso-wrap-style:tight' o:button="t" fillcolor="buttonFace [67]" o:insetmode="auto">
  <v:fill color2="buttonFace [67]" o:detectmouseclick="t"/>
  <o:lock v:ext="edit" rotation="t"/>
  <v:textbox style='mso-direction-alt:auto' o:singleclick="f">
   <div style='text-align:center'><font face="Aptos Narrow" size="220"
   color="#000000">Run All Tests</font></div>
  </v:textbox>
  <x:ClientData ObjectType="Button">
   <x:Anchor>3, 0, 3, 0, 6, 0, 5, 0</x:Anchor>
   <x:PrintObject>False</x:PrintObject>
   <x:AutoFill>False</x:AutoFill>
   <x:FmlaMacro>{RUN_ALL_MACRO}</x:FmlaMacro>
   <x:TextHAlign>Center</x:TextHAlign>
   <x:TextVAlign>Center</x:TextVAlign>
  </x:ClientData>
 </v:shape></xml>'''


def _macro_build_dashboard_drawing() -> str:
    shape_id = 1025
    cid = "{00000000-0008-0000-0000-000000000001}"
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" xmlns:a14="http://schemas.microsoft.com/office/drawing/2010/main" xmlns:a16="http://schemas.microsoft.com/office/drawing/2014/main"><mc:AlternateContent><mc:Choice Requires="a14"><xdr:twoCellAnchor><xdr:from><xdr:col>3</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>3</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from><xdr:to><xdr:col>6</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>5</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to><xdr:sp macro="" textlink=""><xdr:nvSpPr><xdr:cNvPr id="{shape_id}" name="Button 1" hidden="1"><a:extLst><a:ext uri="{{63B3BB69-23CF-44E3-9099-C40C66FF867C}}"><a14:compatExt spid="_x0000_s{shape_id}"/></a:ext><a:ext uri="{{FF2B5EF4-FFF2-40B4-BE49-F238E27FC236}}"><a16:creationId id="{cid}"/></a:ext></a:extLst></xdr:cNvPr><xdr:cNvSpPr/></xdr:nvSpPr><xdr:spPr bwMode="auto"><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom><a:noFill/><a:ln w="9525"><a:miter lim="800000"/><a:headEnd/><a:tailEnd/></a:ln></xdr:spPr><xdr:txBody><a:bodyPr vertOverflow="clip" wrap="square" lIns="36576" tIns="36576" rIns="36576" bIns="36576" anchor="ctr" upright="1"/><a:lstStyle/><a:p><a:pPr algn="ctr" rtl="0"><a:defRPr sz="1000"/></a:pPr><a:r><a:rPr lang="en-NL" sz="1100" b="0" i="0" u="none" strike="noStrike" baseline="0"><a:solidFill><a:srgbClr val="000000"/></a:solidFill><a:latin typeface="Aptos Narrow"/></a:rPr><a:t>Run All Tests</a:t></a:r></a:p></xdr:txBody></xdr:sp><xdr:clientData fPrintsWithSheet="0"/></xdr:twoCellAnchor></mc:Choice><mc:Fallback/></mc:AlternateContent></xdr:wsDr>'''


def _macro_build_tests_xml(headers: List[str], test_rows: List[List[Any]]) -> str:
    n_tests = len(test_rows)
    last_col = len(headers)
    last_col_letter = excel_col_name(last_col)
    last_row = 4 + n_tests
    run_col_zero_based = headers.index("Run")

    rows = [
        _macro_row_xml(1, ["Tests"] + [None] * (last_col - 1), {1: 4}, height="21"),
        _macro_row_xml(
            2,
            [
                "Enable macros. Use Run All Tests on the Dashboard or a row Run button; the macro sends POST requests using the Generated JSON Body column.",
            ]
            + [None] * (last_col - 1),
            {1: 3},
        ),
        _macro_row_xml(4, headers, {i: 5 for i in range(1, last_col + 1)}),
    ]

    variable_headers = set(headers[3:headers.index("Expected")])
    style_data: Dict[int, int] = {}
    for i, header in enumerate(headers, start=1):
        if header in variable_headers:
            style_data[i] = 6
        elif header in {"URL", "Expected", "Actual", "Generated JSON Body", "Coverage Reasons"}:
            style_data[i] = 7
    for r_idx, values in enumerate(test_rows, start=5):
        rows.append(_macro_row_xml(r_idx, values, style_data, height="70"))

    col_widths: Dict[int, float] = {
        1: 52,
        2: 10,
        3: 80,
        headers.index("Expected") + 1: 32,
        headers.index("Actual") + 1: 32,
        headers.index("Status") + 1: 12,
        headers.index("Run") + 1: 12,
        headers.index("Username") + 1: 12,
        headers.index("Password") + 1: 22,
        headers.index("Outcome Names") + 1: 28,
        headers.index("Variable Types") + 1: 44,
        headers.index("Generated JSON Body") + 1: 70,
        last_col: 60,
    }
    for col in range(4, headers.index("Expected") + 1):
        col_widths.setdefault(col, 18)
    for col in range(headers.index("Decision ID") + 1, last_col + 1):
        col_widths.setdefault(col, 28)
    cols_xml = "".join(
        f'<col min="{col}" max="{col}" width="{width}" customWidth="1"/>' for col, width in sorted(col_widths.items())
    )
    controls_xml = _macro_build_controls_xml(n_tests, run_col_zero_based)
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" mc:Ignorable="x14ac xr xr2 xr3" xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac" xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision" xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2" xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3" xr:uid="{{01234567-89AB-CDEF-0123-456789ABCDEF}}"><sheetPr codeName="Blad2"/><dimension ref="A1:{last_col_letter}{last_row}"/><sheetViews><sheetView workbookViewId="0"><pane ySplit="4" topLeftCell="A5" activePane="bottomLeft" state="frozen"/><selection pane="bottomLeft"/></sheetView></sheetViews><sheetFormatPr defaultRowHeight="14.5" x14ac:dyDescent="0.35"/><cols>{cols_xml}</cols><sheetData>{''.join(rows)}</sheetData><autoFilter ref="A4:{last_col_letter}{last_row}"/><pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/><drawing r:id="rId1"/><legacyDrawing r:id="rId2"/><mc:AlternateContent><mc:Choice Requires="x14">{controls_xml}</mc:Choice></mc:AlternateContent></worksheet>'''


def generate_excel_workbook(
    test_cases: List[JsonObj],
    analysis: JsonObj,
    output_path: Path,
    *,
    base_url: str = "https://operaton.open-regels.nl",
    tenant_id: str = "46",
    postman_path: Optional[Path] = None,
    runner_template_path: Optional[Path] = None,
    username: str = "demo",
    password: str = "cqa4fpd2jhz*tph5PVC",
) -> None:
    """Create the repaired macro-enabled Excel runner workbook.

    This writes an .xlsm by using the embedded macro runner as the VBA/control
    container, then replacing Dashboard and Tests with the generated MC/DC cases.
    No external Excel template is required. The Tests sheet XML uses Excel-safe control ordering. Pass runner_template_path only when
    you intentionally want to override the embedded runner.

    The macro names are fully qualified so Excel resolves the buttons:
    PostmanTestRunner.RunAllTests and PostmanTestRunner.RunSelectedTest.
    """
    output_path = Path(output_path)
    if output_path.suffix.lower() != ".xlsm":
        output_path = output_path.with_suffix(".xlsm")

    template_tmp: Optional[str] = None
    if runner_template_path is None:
        template_tmp = tempfile.mkdtemp(prefix="mcdc_embedded_runner_")
        runner_template_path = _macro_write_embedded_runner_template(Path(template_tmp) / "embedded_runner_template.xlsm")
    else:
        runner_template_path = Path(runner_template_path)
        if not runner_template_path.exists():
            raise FileNotFoundError(
                f"Macro-enabled runner template override not found: {runner_template_path}. "
                "Omit --excel-runner-template to use the embedded runner."
            )

    with zipfile.ZipFile(runner_template_path) as template_zip:
        names = set(template_zip.namelist())
        if "xl/vbaProject.bin" not in names:
            raise ValueError(f"Runner template {runner_template_path} does not contain xl/vbaProject.bin macros.")
        required_parts = {
            "xl/workbook.xml",
            "[Content_Types].xml",
            "xl/worksheets/sheet1.xml",
            "xl/worksheets/sheet2.xml",
            "xl/worksheets/_rels/sheet2.xml.rels",
        }
        missing = sorted(required_parts - names)
        if missing:
            raise ValueError(f"Runner template {runner_template_path} is missing required parts: {', '.join(missing)}")

    var_names = sorted({
        name
        for tc in test_cases
        for name in (tc.get("requestBody", {}).get("variables", {}) or {}).keys()
    })
    headers = [
        "Name",
        "Method",
        "URL",
        *var_names,
        "Expected",
        "Actual",
        "Status",
        "Run",
        "Username",
        "Password",
        "Outcome Names",
        "Variable Types",
        "Generated JSON Body",
        "Decision ID",
        "Selected Rule ID",
        "Selected Rule Index",
        "Coverage Reasons",
    ]

    def endpoint_for(decision_id: str) -> str:
        return postman_url_for_decision(base_url, decision_id, tenant_id).get("raw", "")

    test_rows: List[List[Any]] = []
    for tc in test_cases:
        variables = tc.get("requestBody", {}).get("variables", {}) or {}
        coverage = tc.get("coverage", {}) or {}
        decision_id = str(tc.get("decisionId", ""))
        request_body = tc.get("requestBody", {}) or {}
        row = [
            tc.get("name", ""),
            "POST",
            endpoint_for(decision_id),
        ]
        for var in var_names:
            spec = variables.get(var)
            row.append(spec.get("value") if isinstance(spec, dict) else None)
        expected = tc.get("expected", "")
        row.extend(
            [
                expected,
                "",
                "NOT RUN",
                "Run",
                username,
                password,
                _macro_build_outcome_names(expected),
                _macro_build_variable_types(request_body),
                json.dumps(request_body, ensure_ascii=False, indent=2),
                decision_id,
                coverage.get("selectedRuleId", ""),
                coverage.get("selectedRuleIndex", ""),
                "\n".join(coverage.get("reasons", []) or []),
            ]
        )
        test_rows.append(row)

    n_tests = len(test_rows)
    last_row = 4 + n_tests
    last_col_letter = excel_col_name(len(headers))
    run_col_zero_based = headers.index("Run")

    tmp = tempfile.mkdtemp(prefix="mcdc_xlsm_")
    try:
        with zipfile.ZipFile(runner_template_path) as zin:
            zin.extractall(tmp)

        Path(tmp, "xl/worksheets").mkdir(parents=True, exist_ok=True)
        Path(tmp, "xl/drawings").mkdir(parents=True, exist_ok=True)
        Path(tmp, "xl/worksheets/_rels").mkdir(parents=True, exist_ok=True)
        Path(tmp, "xl/ctrlProps").mkdir(parents=True, exist_ok=True)

        Path(tmp, "xl/worksheets/sheet1.xml").write_text(_macro_build_dashboard_xml(test_cases, analysis, postman_path), encoding="utf-8")
        Path(tmp, "xl/worksheets/sheet2.xml").write_text(_macro_build_tests_xml(headers, test_rows), encoding="utf-8")
        Path(tmp, "xl/drawings/vmlDrawing1.vml").write_text(_macro_build_dashboard_vml(), encoding="utf-8")
        Path(tmp, "xl/drawings/drawing1.xml").write_text(_macro_build_dashboard_drawing(), encoding="utf-8")
        Path(tmp, "xl/drawings/vmlDrawing2.vml").write_text(_macro_build_vml(n_tests, run_col_zero_based), encoding="utf-8")
        Path(tmp, "xl/drawings/drawing2.xml").write_text(_macro_build_drawing(n_tests, run_col_zero_based), encoding="utf-8")
        Path(tmp, "xl/worksheets/_rels/sheet2.xml.rels").write_text(_macro_build_sheet2_rels(n_tests), encoding="utf-8")

        ctrl_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<formControlPr xmlns="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" objectType="Button" lockText="1"/>'
        )
        for cp in range(2, n_tests + 2):
            Path(tmp, "xl/ctrlProps", f"ctrlProp{cp}.xml").write_text(ctrl_xml, encoding="utf-8")

        content_types_path = Path(tmp, "[Content_Types].xml")
        content_types_path.write_bytes(_macro_update_content_types(content_types_path.read_bytes(), n_tests))
        workbook_path = Path(tmp, "xl/workbook.xml")
        workbook_path.write_bytes(_macro_update_workbook_xml(workbook_path.read_bytes(), last_col_letter, last_row))
        app_path = Path(tmp, "docProps/app.xml")
        if app_path.exists():
            app_path.write_bytes(_macro_update_app_xml(app_path.read_bytes()))

        output_path.parent.mkdir(parents=True, exist_ok=True)
        if output_path.exists():
            output_path.unlink()
        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for root_dir, _dirs, files in os.walk(tmp):
                for filename in files:
                    full = os.path.join(root_dir, filename)
                    arc = os.path.relpath(full, tmp).replace(os.sep, "/")
                    zout.write(full, arc)
    finally:
        shutil.rmtree(tmp)
        if template_tmp is not None:
            shutil.rmtree(template_tmp)



# ---------------------------------------------------------------------------
# V4 Excel dashboard overrides: decision pie chart + dynamic per-decision run counts
# ---------------------------------------------------------------------------


def _macro_formula_cell_xml(
    row: int,
    col: int,
    formula: str,
    style: Optional[int] = None,
    cached_value: Optional[Any] = None,
) -> str:
    """Return a formula cell with an optional cached value for first-open display."""
    attr_s = f' s="{style}"' if style is not None else ""
    formula_text = html.escape(formula.lstrip("="), quote=False)
    value_xml = ""
    if cached_value is not None and cached_value != "":
        if isinstance(cached_value, bool):
            value_xml = f"<v>{1 if cached_value else 0}</v>"
        elif isinstance(cached_value, (int, float)) and not isinstance(cached_value, bool):
            value_xml = f"<v>{cached_value}</v>"
        else:
            value_xml = f"<v>{_xml_text(cached_value)}</v>"
    return f'<c r="{_macro_cell_ref(row, col)}"{attr_s}><f>{formula_text}</f>{value_xml}</c>'


def _macro_row_from_cells_xml(
    row_num: int,
    cells_xml: Sequence[str],
    spans_end: int,
    height: Optional[str] = None,
) -> str:
    ht = f' ht="{height}" customHeight="1"' if height else ""
    return (
        f'<row r="{row_num}" spans="1:{spans_end}"{ht} x14ac:dyDescent="0.35">'
        + "".join(cells_xml)
        + "</row>"
    )


def _macro_update_content_types(xml_bytes: bytes, n: int) -> bytes:
    root = ET.fromstring(xml_bytes)
    ns = "{http://schemas.openxmlformats.org/package/2006/content-types}"
    existing = {el.attrib.get("PartName") for el in root.findall(f"{ns}Override")}
    required = [
        ("/xl/worksheets/sheet1.xml", "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"),
        ("/xl/worksheets/sheet2.xml", "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"),
        ("/xl/drawings/drawing1.xml", "application/vnd.openxmlformats-officedocument.drawing+xml"),
        ("/xl/drawings/drawing2.xml", "application/vnd.openxmlformats-officedocument.drawing+xml"),
        ("/xl/charts/chart1.xml", "application/vnd.openxmlformats-officedocument.drawingml.chart+xml"),
        ("/xl/vbaProject.bin", "application/vnd.ms-office.vbaProject"),
    ]
    for cp in range(1, n + 2):
        required.append((f"/xl/ctrlProps/ctrlProp{cp}.xml", "application/vnd.ms-excel.controlproperties+xml"))
    for part_name, content_type in required:
        if part_name not in existing:
            ET.SubElement(root, f"{ns}Override", {"PartName": part_name, "ContentType": content_type})
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _macro_update_workbook_xml(xml_bytes: bytes, last_col_letter: str, last_row: int) -> bytes:
    txt = xml_bytes.decode("utf-8")
    sheets_xml = '<sheets><sheet name="Dashboard" sheetId="2" r:id="rId1"/><sheet name="Tests" sheetId="1" r:id="rId2"/></sheets>'
    txt = re.sub(r"<sheets>.*?</sheets>", sheets_xml, txt, flags=re.S)
    defined = (
        f'<definedNames><definedName name="_xlnm._FilterDatabase" localSheetId="1" hidden="1">'
        f'Tests!$A$4:${last_col_letter}${last_row}</definedName></definedNames>'
    )
    if re.search(r"<definedNames>.*?</definedNames>", txt, flags=re.S):
        txt = re.sub(r"<definedNames>.*?</definedNames>", defined, txt, flags=re.S)
    elif "<calcPr" in txt:
        txt = txt.replace("<calcPr", defined + "<calcPr", 1)
    else:
        txt = txt.replace("</workbook>", defined + "</workbook>", 1)

    if "<calcPr" in txt:
        def _calc_repl(match: re.Match[str]) -> str:
            tag = match.group(0)
            tag = re.sub(r'\s+(calcMode|fullCalcOnLoad|forceFullCalc)="[^"]*"', '', tag)
            return tag[:-2] + ' calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"/>' if tag.endswith('/>') else tag
        txt = re.sub(r"<calcPr\b[^>]*/>", _calc_repl, txt, count=1)
    return txt.encode("utf-8")


def _macro_decision_sort_key(item: Tuple[str, Any]) -> Tuple[int, str]:
    preferred = [
        "BehaalbareHoogteSubsidie",
        "BerekenBasisHoogteSubsidie",
        "BerekenBeschikbaarSubsidiePlafond",
        "SubsidieConstantenThuisbatterij",
        "jaarGebondenBudget",
    ]
    key = item[0]
    try:
        return (preferred.index(key), key)
    except ValueError:
        return (len(preferred), key)


def _macro_build_dashboard_xml_v4(
    test_cases: List[JsonObj],
    analysis: JsonObj,
    headers: List[str],
    postman_path: Optional[Path] = None,
) -> str:
    counts = Counter(tc.get("decisionId", "UNKNOWN") for tc in test_cases)
    decisions = analysis.get("decisions") or {}
    metadata = analysis.get("metadata") or {}
    uncovered = analysis.get("uncoveredConditions", []) or []
    n_tests = len(test_cases)
    tests_last_row = 4 + n_tests
    status_col = excel_col_name(headers.index("Status") + 1)
    decision_col = excel_col_name(headers.index("Decision ID") + 1)
    status_range = f"Tests!${status_col}$5:${status_col}${tests_last_row}"
    decision_range = f"Tests!${decision_col}$5:${decision_col}${tests_last_row}"
    name_range = f"Tests!$A$5:$A${tests_last_row}"

    rows: List[str] = []
    spans_end = 9
    rows.append(_macro_row_xml(1, ["DMN MC/DC Test Generation Summary"] + [None] * (spans_end - 1), {1: 2}, height="23.5"))
    rows.append(_macro_row_xml(
        2,
        ["Boundary-focused MC/DC cases generated per DMN decision table. Enable macros, then use Run All Tests or row buttons for real POST execution."] + [None] * (spans_end - 1),
        {1: 3},
    ))

    # Preserve the runner's original dashboard addresses: B4:B7 and D5:E7.
    rows.append(_macro_row_from_cells_xml(4, [
        _macro_cell_xml(4, 1, "Total", 1),
        _macro_formula_cell_xml(4, 2, f"COUNTA({name_range})", 1, n_tests),
        _macro_cell_xml(4, 4, "Result", 1),
        _macro_cell_xml(4, 5, "Count", 1),
    ], spans_end))
    rows.append(_macro_row_from_cells_xml(5, [
        _macro_cell_xml(5, 1, "Passed", 1),
        _macro_formula_cell_xml(5, 2, f"COUNTIF({status_range},\"PASS\")", 1, 0),
        _macro_cell_xml(5, 4, "PASS", 1),
        _macro_formula_cell_xml(5, 5, "B5", 1, 0),
    ], spans_end))
    rows.append(_macro_row_from_cells_xml(6, [
        _macro_cell_xml(6, 1, "Failed/Error", 1),
        _macro_formula_cell_xml(6, 2, f"COUNTIFS({status_range},\"<>PASS\",{status_range},\"<>NOT RUN\",{status_range},\"<>\")", 1, 0),
        _macro_cell_xml(6, 4, "FAIL", 1),
        _macro_formula_cell_xml(6, 5, "B6", 1, 0),
    ], spans_end))
    rows.append(_macro_row_from_cells_xml(7, [
        _macro_cell_xml(7, 1, "Not run", 1),
        _macro_formula_cell_xml(7, 2, f"COUNTIF({status_range},\"NOT RUN\")", 1, n_tests),
        _macro_cell_xml(7, 4, "NOT RUN", 1),
        _macro_formula_cell_xml(7, 5, "B7", 1, n_tests),
    ], spans_end))

    rows.append(_macro_row_xml(9, ["RUN ALL TESTS", None, None, "OPEN TESTS SHEET", None, None, None, None, None], {1: 1, 2: 1, 4: 1, 5: 1}, height="28"))
    postman_hint = postman_path.name if postman_path else "<generated collection>"
    rows.append(_macro_row_xml(11, ["Metric", "Value", None, "Newman command", f"newman run \"{postman_hint}\"", None, None, None, None], {1: 1, 2: 1, 4: 1}, height="18"))
    metric_rows = [
        ("Source DMN", metadata.get("sourceDmn", "")),
        ("Algorithm", metadata.get("algorithm", "")),
        ("Decision count", metadata.get("decisionCount", len(decisions))),
        ("Selected test cases", n_tests),
        ("Uncovered conditions", len(uncovered)),
    ]
    for row_num, (label, value) in enumerate(metric_rows, start=12):
        if label == "Selected test cases":
            rows.append(_macro_row_from_cells_xml(row_num, [
                _macro_cell_xml(row_num, 1, label, 1),
                _macro_formula_cell_xml(row_num, 2, "B4", None, n_tests),
            ], spans_end))
        else:
            rows.append(_macro_row_xml(row_num, [label, value, None, None, None, None, None, None, None], {1: 1}))

    decision_start = 20
    decision_headers = ["Decision ID", "Decision name", "Hit policy", "Candidate count", "Total cases", "Run", "Passed", "Failed/Error", "Not run"]
    rows.append(_macro_row_xml(decision_start, decision_headers, {i: 5 for i in range(1, len(decision_headers) + 1)}))
    current_row = decision_start + 1
    for decision_id, details in sorted(decisions.items(), key=_macro_decision_sort_key):
        total_cache = counts.get(decision_id, details.get("selectedTestCaseCount", 0))
        row_ref = current_row
        rows.append(_macro_row_from_cells_xml(row_ref, [
            _macro_cell_xml(row_ref, 1, decision_id),
            _macro_cell_xml(row_ref, 2, details.get("decisionName", "")),
            _macro_cell_xml(row_ref, 3, details.get("hitPolicy", "")),
            _macro_cell_xml(row_ref, 4, details.get("candidateCountEvaluated", "")),
            _macro_formula_cell_xml(row_ref, 5, f"COUNTIF({decision_range},A{row_ref})", None, total_cache),
            _macro_formula_cell_xml(row_ref, 6, f"G{row_ref}+H{row_ref}", None, 0),
            _macro_formula_cell_xml(row_ref, 7, f"COUNTIFS({decision_range},A{row_ref},{status_range},\"PASS\")", None, 0),
            _macro_formula_cell_xml(row_ref, 8, f"COUNTIFS({decision_range},A{row_ref},{status_range},\"<>PASS\",{status_range},\"<>NOT RUN\",{status_range},\"<>\")", None, 0),
            _macro_formula_cell_xml(row_ref, 9, f"COUNTIFS({decision_range},A{row_ref},{status_range},\"NOT RUN\")", None, total_cache),
        ], spans_end, height="32"))
        current_row += 1

    if current_row == decision_start + 1:
        rows.append(_macro_row_xml(current_row, ["No decision details available", None, None, None, n_tests, 0, 0, 0, n_tests], {}))
        current_row += 1

    last_row = max(current_row - 1, decision_start)
    cols_xml = "".join([
        '<col min="1" max="1" width="34" customWidth="1"/>',
        '<col min="2" max="2" width="36" customWidth="1"/>',
        '<col min="3" max="3" width="18" customWidth="1"/>',
        '<col min="4" max="4" width="20" customWidth="1"/>',
        '<col min="5" max="5" width="16" customWidth="1"/>',
        '<col min="6" max="6" width="12" customWidth="1"/>',
        '<col min="7" max="7" width="12" customWidth="1"/>',
        '<col min="8" max="8" width="15" customWidth="1"/>',
        '<col min="9" max="9" width="12" customWidth="1"/>',
        '<col min="10" max="16" width="15" customWidth="1"/>',
    ])
    merge_xml = (
        '<mergeCells count="4">'
        '<mergeCell ref="A1:I1"/>'
        '<mergeCell ref="A2:I2"/>'
        '<mergeCell ref="A9:B10"/>'
        '<mergeCell ref="D9:E10"/>'
        '</mergeCells>'
    )
    auto_filter_xml = f'<autoFilter ref="A{decision_start}:I{last_row}"/>' if last_row >= decision_start else ""
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" mc:Ignorable="x14ac xr xr2 xr3" xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac" xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision" xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2" xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3" xr:uid="{{5D4A7AAF-31BC-43F8-B4A7-32D52E407D0C}}"><sheetPr codeName="Blad1"/><dimension ref="A1:I{last_row}"/><sheetViews><sheetView topLeftCell="A1" workbookViewId="0"><pane ySplit="19" topLeftCell="A20" activePane="bottomLeft" state="frozen"/><selection pane="bottomLeft"/></sheetView></sheetViews><sheetFormatPr defaultRowHeight="14.5" x14ac:dyDescent="0.35"/><cols>{cols_xml}</cols><sheetData>{''.join(rows)}</sheetData>{auto_filter_xml}{merge_xml}<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/><pageSetup orientation="landscape" r:id="rId1"/><drawing r:id="rId2"/><legacyDrawing r:id="rId3"/><mc:AlternateContent xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"><mc:Choice Requires="x14"><controls><mc:AlternateContent xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"><mc:Choice Requires="x14"><control shapeId="1025" r:id="rId4" name="Button 1"><controlPr defaultSize="0" print="0" autoFill="0" autoPict="0" macro="{RUN_ALL_MACRO}"><anchor moveWithCells="1" sizeWithCells="1"><from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>8</xdr:row><xdr:rowOff>0</xdr:rowOff></from><to><xdr:col>1</xdr:col><xdr:colOff>476250</xdr:colOff><xdr:row>10</xdr:row><xdr:rowOff>38100</xdr:rowOff></to></anchor></controlPr></control></mc:Choice></mc:AlternateContent></controls></mc:Choice></mc:AlternateContent></worksheet>'''


def _macro_build_dashboard_vml() -> str:
    return f'''<xml xmlns:v="urn:schemas-microsoft-com:vml"
 xmlns:o="urn:schemas-microsoft-com:office:office"
 xmlns:x="urn:schemas-microsoft-com:office:excel">
 <o:shapelayout v:ext="edit">
  <o:idmap v:ext="edit" data="1"/>
 </o:shapelayout><v:shapetype id="_x0000_t201" coordsize="21600,21600" o:spt="201"
  path="m,l,21600r21600,l21600,xe">
  <v:stroke joinstyle="miter"/>
  <v:path shadowok="f" o:extrusionok="f" strokeok="f" fillok="f" o:connecttype="rect"/>
  <o:lock v:ext="edit" shapetype="t"/>
 </v:shapetype><v:shape id="_x0000_s1025" type="#_x0000_t201" style='position:absolute;
  margin-left:0pt;margin-top:137.25pt;width:107.5pt;height:34.5pt;z-index:1;
  mso-wrap-style:tight' o:button="t" fillcolor="buttonFace [67]" o:insetmode="auto">
  <v:fill color2="buttonFace [67]" o:detectmouseclick="t"/>
  <o:lock v:ext="edit" rotation="t"/>
  <v:textbox style='mso-direction-alt:auto' o:singleclick="f">
   <div style='text-align:center'><font face="Aptos Narrow" size="220"
   color="#000000">Run All Tests</font></div>
  </v:textbox>
  <x:ClientData ObjectType="Button">
   <x:Anchor>0, 0, 8, 0, 1, 75, 10, 6</x:Anchor>
   <x:PrintObject>False</x:PrintObject>
   <x:AutoFill>False</x:AutoFill>
   <x:FmlaMacro>{RUN_ALL_MACRO}</x:FmlaMacro>
   <x:TextHAlign>Center</x:TextHAlign>
   <x:TextVAlign>Center</x:TextVAlign>
  </x:ClientData>
 </v:shape></xml>'''


def _macro_build_dashboard_drawing() -> str:
    shape_id = 1025
    cid = "{00000000-0008-0000-0000-000000000001}"
    chart_cid = "{00000000-0008-0000-0000-000000000002}"
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" xmlns:a14="http://schemas.microsoft.com/office/drawing/2010/main" xmlns:a16="http://schemas.microsoft.com/office/drawing/2014/main"><mc:AlternateContent><mc:Choice Requires="a14"><xdr:twoCellAnchor><xdr:from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>8</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from><xdr:to><xdr:col>1</xdr:col><xdr:colOff>476250</xdr:colOff><xdr:row>10</xdr:row><xdr:rowOff>38100</xdr:rowOff></xdr:to><xdr:sp macro="" textlink=""><xdr:nvSpPr><xdr:cNvPr id="{shape_id}" name="Button 1" hidden="1"><a:extLst><a:ext uri="{{63B3BB69-23CF-44E3-9099-C40C66FF867C}}"><a14:compatExt spid="_x0000_s{shape_id}"/></a:ext><a:ext uri="{{FF2B5EF4-FFF2-40B4-BE49-F238E27FC236}}"><a16:creationId id="{cid}"/></a:ext></a:extLst></xdr:cNvPr><xdr:cNvSpPr/></xdr:nvSpPr><xdr:spPr bwMode="auto"><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom><a:noFill/><a:ln w="9525"><a:miter lim="800000"/><a:headEnd/><a:tailEnd/></a:ln></xdr:spPr><xdr:txBody><a:bodyPr vertOverflow="clip" wrap="square" lIns="36576" tIns="36576" rIns="36576" bIns="36576" anchor="ctr" upright="1"/><a:lstStyle/><a:p><a:pPr algn="ctr" rtl="0"><a:defRPr sz="1000"/></a:pPr><a:r><a:rPr lang="en-NL" sz="1100" b="0" i="0" u="none" strike="noStrike" baseline="0"><a:solidFill><a:srgbClr val="000000"/></a:solidFill><a:latin typeface="Aptos Narrow"/></a:rPr><a:t>Run All Tests</a:t></a:r></a:p></xdr:txBody></xdr:sp><xdr:clientData fPrintsWithSheet="0"/></xdr:twoCellAnchor></mc:Choice><mc:Fallback/></mc:AlternateContent><xdr:twoCellAnchor><xdr:from><xdr:col>6</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>3</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from><xdr:to><xdr:col>15</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>19</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to><xdr:graphicFrame macro=""><xdr:nvGraphicFramePr><xdr:cNvPr id="2" name="DecisionCasesPieChart"><a:extLst><a:ext uri="{{FF2B5EF4-FFF2-40B4-BE49-F238E27FC236}}"><a16:creationId id="{chart_cid}"/></a:ext></a:extLst></xdr:cNvPr><xdr:cNvGraphicFramePr/></xdr:nvGraphicFramePr><xdr:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></xdr:xfrm><a:graphic><a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/chart"><c:chart xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId1"/></a:graphicData></a:graphic></xdr:graphicFrame><xdr:clientData/></xdr:twoCellAnchor></xdr:wsDr>'''


def _macro_build_decision_pie_chart_xml(test_cases: List[JsonObj], analysis: JsonObj, decision_start: int = 20) -> str:
    counts = Counter(tc.get("decisionId", "UNKNOWN") for tc in test_cases)
    decisions = analysis.get("decisions") or {}
    ordered = list(sorted(decisions.items(), key=_macro_decision_sort_key))
    if not ordered:
        ordered = [("All tests", {"selectedTestCaseCount": len(test_cases)})]
    data_start = decision_start + 1
    data_end = data_start + len(ordered) - 1
    labels = [str(decision_id) for decision_id, _details in ordered]
    values = [counts.get(decision_id, details.get("selectedTestCaseCount", 0)) for decision_id, details in ordered]
    pt_count = len(labels)
    str_pts = "".join(f'<c:pt idx="{idx}"><c:v>{_xml_text(label)}</c:v></c:pt>' for idx, label in enumerate(labels))
    num_pts = "".join(f'<c:pt idx="{idx}"><c:v>{value}</c:v></c:pt>' for idx, value in enumerate(values))
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<c:chartSpace xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><c:date1904 val="0"/><c:lang val="nl-NL"/><c:roundedCorners val="0"/><c:style val="10"/><c:chart><c:title><c:tx><c:rich><a:bodyPr/><a:lstStyle/><a:p><a:pPr><a:defRPr/></a:pPr><a:r><a:rPr lang="nl-NL"/><a:t>MC/DC cases by decision</a:t></a:r></a:p></c:rich></c:tx><c:layout/><c:overlay val="0"/></c:title><c:autoTitleDeleted val="0"/><c:plotArea><c:layout/><c:pieChart><c:varyColors val="1"/><c:ser><c:idx val="0"/><c:order val="0"/><c:tx><c:strRef><c:f>Dashboard!$E$20</c:f><c:strCache><c:ptCount val="1"/><c:pt idx="0"><c:v>Total cases</c:v></c:pt></c:strCache></c:strRef></c:tx><c:cat><c:strRef><c:f>Dashboard!$A${data_start}:$A${data_end}</c:f><c:strCache><c:ptCount val="{pt_count}"/>{str_pts}</c:strCache></c:strRef></c:cat><c:val><c:numRef><c:f>Dashboard!$E${data_start}:$E${data_end}</c:f><c:numCache><c:formatCode>General</c:formatCode><c:ptCount val="{pt_count}"/>{num_pts}</c:numCache></c:numRef></c:val><c:dLbls><c:showLegendKey val="0"/><c:showVal val="0"/><c:showCatName val="0"/><c:showSerName val="0"/><c:showPercent val="1"/><c:showLeaderLines val="1"/></c:dLbls></c:ser><c:firstSliceAng val="0"/></c:pieChart></c:plotArea><c:legend><c:legendPos val="r"/><c:layout/><c:overlay val="0"/></c:legend><c:plotVisOnly val="1"/><c:dispBlanksAs val="gap"/><c:showDLblsOverMax val="0"/></c:chart><c:printSettings><c:headerFooter/><c:pageMargins b="0.75" l="0.7" r="0.7" t="0.75" header="0.3" footer="0.3"/><c:pageSetup/></c:printSettings></c:chartSpace>'''


def generate_excel_workbook(
    test_cases: List[JsonObj],
    analysis: JsonObj,
    output_path: Path,
    *,
    base_url: str = "https://operaton.open-regels.nl",
    tenant_id: str = "46",
    postman_path: Optional[Path] = None,
    runner_template_path: Optional[Path] = None,
    username: str = "demo",
    password: str = "cqa4fpd2jhz*tph5PVC",
) -> None:
    """Create the macro-enabled Excel runner with a decision pie chart and dynamic status counts."""
    output_path = Path(output_path)
    if output_path.suffix.lower() != ".xlsm":
        output_path = output_path.with_suffix(".xlsm")

    template_tmp: Optional[str] = None
    if runner_template_path is None:
        template_tmp = tempfile.mkdtemp(prefix="mcdc_embedded_runner_")
        runner_template_path = _macro_write_embedded_runner_template(Path(template_tmp) / "embedded_runner_template.xlsm")
    else:
        runner_template_path = Path(runner_template_path)
        if not runner_template_path.exists():
            raise FileNotFoundError(
                f"Macro-enabled runner template override not found: {runner_template_path}. "
                "Omit --excel-runner-template to use the embedded runner."
            )

    with zipfile.ZipFile(runner_template_path) as template_zip:
        names = set(template_zip.namelist())
        if "xl/vbaProject.bin" not in names:
            raise ValueError(f"Runner template {runner_template_path} does not contain xl/vbaProject.bin macros.")
        required_parts = {
            "xl/workbook.xml",
            "[Content_Types].xml",
            "xl/worksheets/sheet1.xml",
            "xl/worksheets/sheet2.xml",
            "xl/worksheets/_rels/sheet2.xml.rels",
        }
        missing = sorted(required_parts - names)
        if missing:
            raise ValueError(f"Runner template {runner_template_path} is missing required parts: {', '.join(missing)}")

    var_names = sorted({
        name
        for tc in test_cases
        for name in (tc.get("requestBody", {}).get("variables", {}) or {}).keys()
    })
    headers = [
        "Name", "Method", "URL", *var_names, "Expected", "Actual", "Status", "Run", "Username", "Password",
        "Outcome Names", "Variable Types", "Generated JSON Body", "Decision ID", "Selected Rule ID", "Selected Rule Index", "Coverage Reasons",
    ]

    def endpoint_for(decision_id: str) -> str:
        return postman_url_for_decision(base_url, decision_id, tenant_id).get("raw", "")

    test_rows: List[List[Any]] = []
    for tc in test_cases:
        variables = tc.get("requestBody", {}).get("variables", {}) or {}
        coverage = tc.get("coverage", {}) or {}
        decision_id = str(tc.get("decisionId", ""))
        request_body = tc.get("requestBody", {}) or {}
        row = [tc.get("name", ""), "POST", endpoint_for(decision_id)]
        for var in var_names:
            spec = variables.get(var)
            row.append(spec.get("value") if isinstance(spec, dict) else None)
        expected = tc.get("expected", "")
        row.extend([
            expected,
            "",
            "NOT RUN",
            "Run",
            username,
            password,
            _macro_build_outcome_names(expected),
            _macro_build_variable_types(request_body),
            json.dumps(request_body, ensure_ascii=False, indent=2),
            decision_id,
            coverage.get("selectedRuleId", ""),
            coverage.get("selectedRuleIndex", ""),
            "\n".join(coverage.get("reasons", []) or []),
        ])
        test_rows.append(row)

    n_tests = len(test_rows)
    last_row = 4 + n_tests
    last_col_letter = excel_col_name(len(headers))
    run_col_zero_based = headers.index("Run")

    tmp = tempfile.mkdtemp(prefix="mcdc_xlsm_")
    try:
        with zipfile.ZipFile(runner_template_path) as zin:
            zin.extractall(tmp)

        Path(tmp, "xl/worksheets").mkdir(parents=True, exist_ok=True)
        Path(tmp, "xl/drawings").mkdir(parents=True, exist_ok=True)
        Path(tmp, "xl/drawings/_rels").mkdir(parents=True, exist_ok=True)
        Path(tmp, "xl/charts").mkdir(parents=True, exist_ok=True)
        Path(tmp, "xl/worksheets/_rels").mkdir(parents=True, exist_ok=True)
        Path(tmp, "xl/ctrlProps").mkdir(parents=True, exist_ok=True)

        Path(tmp, "xl/worksheets/sheet1.xml").write_text(_macro_build_dashboard_xml_v4(test_cases, analysis, headers, postman_path), encoding="utf-8")
        Path(tmp, "xl/worksheets/sheet2.xml").write_text(_macro_build_tests_xml(headers, test_rows), encoding="utf-8")
        Path(tmp, "xl/drawings/vmlDrawing1.vml").write_text(_macro_build_dashboard_vml(), encoding="utf-8")
        Path(tmp, "xl/drawings/drawing1.xml").write_text(_macro_build_dashboard_drawing(), encoding="utf-8")
        Path(tmp, "xl/drawings/_rels/drawing1.xml.rels").write_text(
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart" Target="../charts/chart1.xml"/>'
            '</Relationships>',
            encoding="utf-8",
        )
        Path(tmp, "xl/charts/chart1.xml").write_text(_macro_build_decision_pie_chart_xml(test_cases, analysis), encoding="utf-8")
        Path(tmp, "xl/drawings/vmlDrawing2.vml").write_text(_macro_build_vml(n_tests, run_col_zero_based), encoding="utf-8")
        Path(tmp, "xl/drawings/drawing2.xml").write_text(_macro_build_drawing(n_tests, run_col_zero_based), encoding="utf-8")
        Path(tmp, "xl/worksheets/_rels/sheet2.xml.rels").write_text(_macro_build_sheet2_rels(n_tests), encoding="utf-8")

        ctrl_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<formControlPr xmlns="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" objectType="Button" lockText="1"/>'
        )
        for cp in range(2, n_tests + 2):
            Path(tmp, "xl/ctrlProps", f"ctrlProp{cp}.xml").write_text(ctrl_xml, encoding="utf-8")

        content_types_path = Path(tmp, "[Content_Types].xml")
        content_types_path.write_bytes(_macro_update_content_types(content_types_path.read_bytes(), n_tests))
        workbook_path = Path(tmp, "xl/workbook.xml")
        workbook_path.write_bytes(_macro_update_workbook_xml(workbook_path.read_bytes(), last_col_letter, last_row))
        app_path = Path(tmp, "docProps/app.xml")
        if app_path.exists():
            app_path.write_bytes(_macro_update_app_xml(app_path.read_bytes()))

        output_path.parent.mkdir(parents=True, exist_ok=True)
        if output_path.exists():
            output_path.unlink()
        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for root_dir, _dirs, files in os.walk(tmp):
                for filename in files:
                    full = os.path.join(root_dir, filename)
                    arc = os.path.relpath(full, tmp).replace(os.sep, "/")
                    zout.write(full, arc)
    finally:
        shutil.rmtree(tmp)
        if template_tmp is not None:
            shutil.rmtree(template_tmp)


# ---------------------------------------------------------------------------
# End-to-end CLI
# ---------------------------------------------------------------------------


def default_prefix_from_dmn(dmn_path: Path) -> str:
    stem = dmn_path.stem
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", stem).strip("_") or "dmn"


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Generate MC/DC + boundary DMN test cases as JSON, an Excel workbook, "
            "and a Postman collection with the generated cases grouped as examples."
        )
    )
    parser.add_argument("dmn", type=Path, help="Path to the DMN XML file")
    parser.add_argument("--postman-template", type=Path, help="Optional base Postman collection with one decision-evaluate request per table")
    parser.add_argument("--out-dir", type=Path, default=Path("."), help="Directory for generated files")
    parser.add_argument("--prefix", help="Output filename prefix. Defaults to the DMN filename stem")
    parser.add_argument("--base-url", default="https://operaton.open-regels.nl", help="Base URL used when no Postman template exists. Accepts either host root or /engine-rest URL.")
    parser.add_argument("--tenant-id", default="46", help="Tenant id used when no Postman template exists")
    parser.add_argument("--excel-runner-template", type=Path, help="Optional macro-enabled Excel runner template override. Omit to use the embedded runner.")
    parser.add_argument("--excel-username", default="demo", help="Username written to the generated Tests sheet")
    parser.add_argument("--excel-password", default="cqa4fpd2jhz*tph5PVC", help="Password written to the generated Tests sheet")
    parser.add_argument("--max-candidates-per-decision", type=int, default=100_000, help="Candidate safety cap per decision table")
    parser.add_argument("--max-cases-per-decision", type=int, help="Optional cap after MC/DC/boundary selection per decision table")
    parser.add_argument("--skip-excel", action="store_true", help="Skip writing the .xlsm workbook")
    parser.add_argument("--skip-postman", action="store_true", help="Skip writing the Postman collection")
    args = parser.parse_args(argv)

    if not args.dmn.exists():
        raise FileNotFoundError(args.dmn)
    if args.postman_template and not args.postman_template.exists():
        raise FileNotFoundError(args.postman_template)

    args.out_dir.mkdir(parents=True, exist_ok=True)
    prefix = args.prefix or default_prefix_from_dmn(args.dmn)
    json_path = args.out_dir / f"{prefix}-mcdc-test-cases.json"
    analysis_path = args.out_dir / f"{prefix}-mcdc-analysis.json"
    excel_path = args.out_dir / f"{prefix}-mcdc-test-cases.xlsm"
    postman_path = args.out_dir / f"{prefix}.postman_collection.mcdc_examples.json"

    cases, analysis = generate(
        args.dmn,
        max_candidates_per_decision=args.max_candidates_per_decision,
        max_cases_per_decision=args.max_cases_per_decision,
    )

    dump_json(cases, json_path)
    dump_json(analysis, analysis_path)

    if not args.skip_excel:
        runner_template = args.excel_runner_template
        if runner_template is not None and not runner_template.exists():
            candidate = args.dmn.parent / runner_template
            if candidate.exists():
                runner_template = candidate
        generate_excel_workbook(
            cases,
            analysis,
            excel_path,
            base_url=args.base_url,
            tenant_id=args.tenant_id,
            postman_path=postman_path if not args.skip_postman else None,
            runner_template_path=runner_template,
            username=args.excel_username,
            password=args.excel_password,
        )

    if not args.skip_postman:
        collection = generate_postman_collection(
            cases,
            postman_template=args.postman_template,
            base_url=args.base_url,
            tenant_id=args.tenant_id,
        )
        dump_json(collection, postman_path)

    counts = Counter(tc.get("decisionId", "UNKNOWN") for tc in cases)
    print(f"Generated MC/DC test cases: {len(cases)}")
    for decision_id, count in sorted(counts.items()):
        print(f"  {decision_id}: {count}")
    print(f"Uncovered conditions: {len(analysis.get('uncoveredConditions', []))}")
    print(f"JSON:     {json_path}")
    print(f"Analysis: {analysis_path}")
    if not args.skip_excel:
        print(f"Excel:    {excel_path}")
    if not args.skip_postman:
        print(f"Postman:  {postman_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
