#!/usr/bin/env python3
"""
dmn_to_pretty_excel.py

Generate a macro-enabled Excel test runner from a DMN decision-table file, using the
working `thuisbatterij-testsPretty.xlsm` workbook as the macro/template base.

The generated workbook keeps the working VBA project from the template, rewrites the
Tests and Dashboard sheets, creates a Run All Tests dashboard button, creates one Run
button per test row, and keeps/refreshes the pie chart area.

Usage:
    python dmn_to_pretty_excel.py input.dmn \
        --template thuisbatterij-testsPretty.xlsm \
        --output generated-tests.xlsm \
        --base-url https://operaton.open-regels.nl/engine-rest \
        --tenant-id 46 \
        --username demo \
        --password 'secret'

Notes:
    - This is a generic DMN decision-table-to-test-runner generator. It creates one
      representative rule-coverage test per DMN rule. True MC/DC for arbitrary FEEL
      expressions often requires semantic/domain-specific case design; use the generated
      workbook as a solid generic baseline and add custom edge cases where needed.
    - Requires openpyxl: pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import sys
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - helpful CLI error
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: pip install openpyxl"
    ) from exc


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class InputColumn:
    name: str
    label: str
    type_ref: str


@dataclass
class OutputColumn:
    name: str
    label: str
    type_ref: str


@dataclass
class DmnRule:
    rule_id: str
    index: int
    input_entries: List[str]
    output_entries: List[str]
    annotation: str = ""


@dataclass
class DmnDecision:
    decision_id: str
    decision_name: str
    hit_policy: str
    inputs: List[InputColumn]
    outputs: List[OutputColumn]
    rules: List[DmnRule]


@dataclass
class TestRow:
    name: str
    decision_id: str
    rule_id: str
    rule_index: int
    input_values: Dict[str, Any]
    input_types: Dict[str, str]
    expected_values: Dict[str, Any]
    output_names: List[str]
    body_json: str
    variable_types: str
    expected_text: str
    coverage_reason: str
    url: str = ""


# ---------------------------------------------------------------------------
# DMN XML parsing helpers
# ---------------------------------------------------------------------------


def local_name(tag: str) -> str:
    if "}" in tag:
        return tag.rsplit("}", 1)[1]
    return tag


def children_by_local(elem: ET.Element, wanted: str) -> List[ET.Element]:
    return [c for c in list(elem) if local_name(c.tag) == wanted]


def first_child(elem: ET.Element, wanted: str) -> Optional[ET.Element]:
    for child in list(elem):
        if local_name(child.tag) == wanted:
            return child
    return None


def first_descendant(elem: ET.Element, wanted: str) -> Optional[ET.Element]:
    for descendant in elem.iter():
        if local_name(descendant.tag) == wanted and descendant is not elem:
            return descendant
    return None


def text_of_first_descendant(elem: ET.Element, wanted: str, default: str = "") -> str:
    child = first_descendant(elem, wanted)
    if child is None or child.text is None:
        return default
    return child.text.strip()


def normalize_var_name(raw: str, fallback: str) -> str:
    """Turn a DMN expression/label into a usable variable name."""
    raw = (raw or "").strip()
    fallback = (fallback or "").strip() or "var"

    # Common DMN inputExpression text is simply the variable name.
    if re.fullmatch(r"[A-Za-z_][\w.:-]*", raw):
        return raw

    # Some modelers use FEEL property paths or variables with spaces in labels.
    if re.fullmatch(r"[A-Za-z_][\w.:-]*(\.[A-Za-z_][\w.:-]*)+", raw):
        return raw

    candidate = re.sub(r"\W+", "_", fallback).strip("_")
    if not candidate:
        candidate = "var"
    if re.match(r"^\d", candidate):
        candidate = f"var_{candidate}"
    return candidate


def parse_dmn(dmn_path: Path) -> List[DmnDecision]:
    tree = ET.parse(dmn_path)
    root = tree.getroot()
    decisions: List[DmnDecision] = []

    for decision_elem in root.iter():
        if local_name(decision_elem.tag) != "decision":
            continue

        decision_table = first_child(decision_elem, "decisionTable")
        if decision_table is None:
            # Some tools wrap or place it a little differently; fall back to descendant.
            decision_table = first_descendant(decision_elem, "decisionTable")
        if decision_table is None:
            continue

        decision_id = decision_elem.attrib.get("id") or decision_elem.attrib.get("name") or "Decision"
        decision_name = decision_elem.attrib.get("name") or decision_id
        hit_policy = decision_table.attrib.get("hitPolicy", "UNIQUE")

        inputs: List[InputColumn] = []
        for i, input_clause in enumerate(children_by_local(decision_table, "input"), start=1):
            label = input_clause.attrib.get("label") or input_clause.attrib.get("id") or f"input{i}"
            input_expr = first_child(input_clause, "inputExpression")
            expr_text = ""
            type_ref = input_clause.attrib.get("typeRef", "")
            if input_expr is not None:
                expr_text = text_of_first_descendant(input_expr, "text", "")
                type_ref = input_expr.attrib.get("typeRef", type_ref)
            name = normalize_var_name(expr_text or label, label)
            inputs.append(InputColumn(name=name, label=label, type_ref=type_ref or "String"))

        outputs: List[OutputColumn] = []
        for i, output_clause in enumerate(children_by_local(decision_table, "output"), start=1):
            label = output_clause.attrib.get("label") or output_clause.attrib.get("name") or output_clause.attrib.get("id") or f"output{i}"
            name = output_clause.attrib.get("name") or normalize_var_name(label, f"output{i}")
            type_ref = output_clause.attrib.get("typeRef", "String")
            outputs.append(OutputColumn(name=name, label=label, type_ref=type_ref or "String"))

        rules: List[DmnRule] = []
        for idx, rule_elem in enumerate(children_by_local(decision_table, "rule"), start=1):
            rule_id = rule_elem.attrib.get("id") or f"R{idx}"
            input_entries = [text_of_first_descendant(e, "text", "-") for e in children_by_local(rule_elem, "inputEntry")]
            output_entries = [text_of_first_descendant(e, "text", "") for e in children_by_local(rule_elem, "outputEntry")]
            annotation_entries = [text_of_first_descendant(e, "text", "") for e in children_by_local(rule_elem, "annotationEntry")]
            annotation = "; ".join(a for a in annotation_entries if a)
            rules.append(DmnRule(rule_id=rule_id, index=idx, input_entries=input_entries, output_entries=output_entries, annotation=annotation))

        decisions.append(
            DmnDecision(
                decision_id=decision_id,
                decision_name=decision_name,
                hit_policy=hit_policy,
                inputs=inputs,
                outputs=outputs,
                rules=rules,
            )
        )

    if not decisions:
        raise ValueError(f"No DMN decision tables found in {dmn_path}")
    return decisions


# ---------------------------------------------------------------------------
# FEEL sampling and JSON/Expected formatting
# ---------------------------------------------------------------------------


def strip_outer_quotes(value: str) -> str:
    value = value.strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
        return value[1:-1]
    return value


def normalize_type_ref(type_ref: str) -> str:
    t = (type_ref or "String").strip().lower()
    t = t.split(":")[-1]
    aliases = {
        "string": "String",
        "str": "String",
        "date": "String",
        "datetime": "String",
        "time": "String",
        "boolean": "Boolean",
        "bool": "Boolean",
        "integer": "Integer",
        "int": "Integer",
        "long": "Long",
        "double": "Double",
        "decimal": "Double",
        "number": "Double",
        "float": "Double",
    }
    return aliases.get(t, "String")


def default_for_type(type_ref: str) -> Any:
    t = normalize_type_ref(type_ref)
    if t in {"Integer", "Long"}:
        return 1
    if t == "Double":
        return 1.0
    if t == "Boolean":
        return False
    return "sample"


def split_top_level_commas(expr: str) -> List[str]:
    parts: List[str] = []
    buf: List[str] = []
    depth = 0
    quote: Optional[str] = None
    for ch in expr:
        if quote:
            buf.append(ch)
            if ch == quote:
                quote = None
            continue
        if ch in {"'", '"'}:
            quote = ch
            buf.append(ch)
            continue
        if ch in "([":
            depth += 1
        elif ch in ")]" and depth > 0:
            depth -= 1
        if ch == "," and depth == 0:
            parts.append("".join(buf).strip())
            buf = []
        else:
            buf.append(ch)
    if buf:
        parts.append("".join(buf).strip())
    return parts


def parse_numeric(value: str) -> Optional[float]:
    value = strip_outer_quotes(value.strip())
    value = value.replace("_", "")
    if re.fullmatch(r"[-+]?\d+(\.\d+)?", value):
        return float(value)
    return None


def cast_number(num: float, type_ref: str) -> Any:
    t = normalize_type_ref(type_ref)
    if t in {"Integer", "Long"}:
        return int(round(num))
    return float(num)


def first_reasonable_choice(expr: str) -> str:
    for part in split_top_level_commas(expr):
        part = part.strip()
        if part and part != "-":
            return part
    return expr.strip()


def sample_value_from_feel(expr: str, type_ref: str) -> Any:
    """Return a representative Python value satisfying a simple FEEL unary test."""
    original = (expr or "").strip()
    if not original or original == "-":
        return default_for_type(type_ref)

    expr = first_reasonable_choice(original)
    lowered = expr.lower().strip()

    # FEEL functions often used for dates.
    m = re.fullmatch(r"date\(\s*['\"]([^'\"]+)['\"]\s*\)", expr, flags=re.I)
    if m:
        return m.group(1)
    m = re.fullmatch(r"date and time\(\s*['\"]([^'\"]+)['\"]\s*\)", expr, flags=re.I)
    if m:
        return m.group(1)

    if lowered in {"true", "false"}:
        return lowered == "true"

    if (expr.startswith("'") and expr.endswith("'")) or (expr.startswith('"') and expr.endswith('"')):
        return strip_outer_quotes(expr)

    # FEEL intervals: [1..10], (1..10], [date("2026-01-01")..date("2026-12-31")]
    m = re.fullmatch(r"[\[\(]\s*(.*?)\s*\.\.\s*(.*?)\s*[\]\)]", expr)
    if m:
        low_raw, high_raw = m.group(1), m.group(2)
        low_num, high_num = parse_numeric(low_raw), parse_numeric(high_raw)
        if low_num is not None and high_num is not None:
            return cast_number((low_num + high_num) / 2, type_ref)
        # For date/string intervals, pick the low endpoint if recognizable.
        low_sample = sample_value_from_feel(low_raw, "String")
        return low_sample

    # Comparisons.
    m = re.fullmatch(r"(<=|>=|<|>)\s*(.+)", expr)
    if m:
        op, rhs = m.group(1), m.group(2).strip()
        rhs_num = parse_numeric(rhs)
        if rhs_num is not None:
            if op == "<":
                return cast_number(rhs_num - 1, type_ref)
            if op == "<=":
                return cast_number(rhs_num, type_ref)
            if op == ">":
                return cast_number(rhs_num + 1, type_ref)
            if op == ">=":
                return cast_number(rhs_num, type_ref)
        return sample_value_from_feel(rhs, type_ref)

    # not(...) cannot be sampled reliably without domain knowledge. Use the default.
    if lowered.startswith("not(") or lowered.startswith("not ("):
        return default_for_type(type_ref)

    # Exact numeric literal.
    n = parse_numeric(expr)
    if n is not None:
        return cast_number(n, type_ref)

    # Anything else is treated as a string literal, with FEEL quotes removed when present.
    return strip_outer_quotes(expr)


def expected_value_from_output(expr: str, type_ref: str) -> Any:
    expr = (expr or "").strip()
    if expr == "":
        return ""
    return sample_value_from_feel(expr, type_ref)


def value_for_json(value: Any, variable_type: str) -> Any:
    t = normalize_type_ref(variable_type)
    if t == "Boolean":
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() == "true"
    if t in {"Integer", "Long"}:
        try:
            return int(value)
        except Exception:
            return value
    if t == "Double":
        try:
            return float(value)
        except Exception:
            return value
    return "" if value is None else str(value)


def format_expected_value(value: Any, type_ref: str) -> str:
    t = normalize_type_ref(type_ref)
    if value is None:
        return ""
    if t == "Boolean" or isinstance(value, bool):
        return "true" if bool(value) else "false"
    if t in {"Integer", "Long", "Double"} and isinstance(value, (int, float)):
        if isinstance(value, float) and math.isfinite(value) and value.is_integer():
            return str(int(value))
        return str(value)
    # Use single quotes to match the existing workbook style for text outputs.
    text = str(value).replace("'", "''")
    return f"'{text}'"


def excel_safe_literal(value: Any) -> Any:
    """Prevent Excel interpreting user/DMN literals beginning with '=' as formulas."""
    if isinstance(value, str) and value.startswith("="):
        return "'" + value
    return value


def slugify(text: str, max_len: int = 60) -> str:
    text = re.sub(r"\s+", "_", text.strip())
    text = re.sub(r"[^A-Za-z0-9_\-.]+", "", text)
    return (text or "rule")[:max_len]


def build_test_rows(
    decisions: Sequence[DmnDecision],
    base_url: str,
    tenant_id: str,
) -> Tuple[List[TestRow], List[str]]:
    all_inputs: List[str] = []
    rows: List[TestRow] = []

    for decision in decisions:
        for inp in decision.inputs:
            if inp.name not in all_inputs:
                all_inputs.append(inp.name)

    for d_idx, decision in enumerate(decisions, start=1):
        for rule in decision.rules:
            input_values: Dict[str, Any] = {}
            input_types: Dict[str, str] = {}
            for i, inp in enumerate(decision.inputs):
                entry = rule.input_entries[i] if i < len(rule.input_entries) else "-"
                typ = normalize_type_ref(inp.type_ref)
                input_values[inp.name] = sample_value_from_feel(entry, typ)
                input_types[inp.name] = typ

            expected_values: Dict[str, Any] = {}
            expected_parts: List[str] = []
            output_names: List[str] = []
            for j, out in enumerate(decision.outputs):
                entry = rule.output_entries[j] if j < len(rule.output_entries) else ""
                typ = normalize_type_ref(out.type_ref)
                value = expected_value_from_output(entry, typ)
                expected_values[out.name] = value
                output_names.append(out.name)
                expected_parts.append(f"{out.name}={format_expected_value(value, typ)}")

            variables = {
                name: {"value": value_for_json(input_values[name], input_types[name]), "type": input_types[name]}
                for name in input_values
            }

            if tenant_id:
                url = f"{base_url.rstrip('/')}/decision-definition/key/{decision.decision_id}/tenant-id/{tenant_id}/evaluate"
            else:
                url = f"{base_url.rstrip('/')}/decision-definition/key/{decision.decision_id}/evaluate"

            rule_label = slugify(rule.rule_id or f"R{rule.index}")
            decision_label = slugify(decision.decision_id)
            name = f"TC-{d_idx}.{rule.index}_{decision_label}_{rule_label}"
            if rule.annotation:
                reason = f"Rule coverage: generated from DMN rule {rule.rule_id}. Annotation: {rule.annotation}"
            else:
                reason = f"Rule coverage: generated from DMN rule {rule.rule_id} / index {rule.index}. Inputs use representative values from inputEntry expressions."

            rows.append(
                TestRow(
                    name=name,
                    decision_id=decision.decision_id,
                    rule_id=rule.rule_id,
                    rule_index=rule.index,
                    input_values=input_values,
                    input_types=input_types,
                    expected_values=expected_values,
                    output_names=output_names,
                    body_json=json.dumps({"variables": variables}, ensure_ascii=False, separators=(",", ":")),
                    variable_types=";".join(f"{n}:{input_types[n]}" for n in input_values),
                    expected_text=", ".join(expected_parts),
                    coverage_reason=reason,
                    url=url,
                )
            )

    return rows, all_inputs


# ---------------------------------------------------------------------------
# Workbook cell generation via openpyxl
# ---------------------------------------------------------------------------


def clear_sheet(ws) -> None:
    if ws.max_row > 1:
        ws.delete_rows(1, ws.max_row)
    elif ws.max_row == 1:
        for cell in ws[1]:
            cell.value = None


def write_workbook_cells(
    template_path: Path,
    output_path: Path,
    test_rows: Sequence[TestRow],
    all_inputs: Sequence[str],
    username: str,
    password: str,
    title: str,
) -> int:
    wb = load_workbook(template_path, keep_vba=True)
    if "Tests" not in wb.sheetnames:
        raise ValueError("Template must contain a sheet named 'Tests'.")
    if "Dashboard" not in wb.sheetnames:
        raise ValueError("Template must contain a sheet named 'Dashboard'.")

    ws = wb["Tests"]
    dash = wb["Dashboard"]
    clear_sheet(ws)
    clear_sheet(dash)

    base_headers = ["Name", "Method", "URL"]
    tail_headers = [
        "Expected", "Actual", "Status", "Run", "Username", "Password",
        "Outcome Names", "Variable Types", "Generated JSON Body", "Decision ID",
        "Selected Rule ID", "Selected Rule Index", "Coverage Reasons",
    ]
    headers = base_headers + list(all_inputs) + tail_headers
    header_row = 4
    first_data_row = 5
    run_col = headers.index("Run") + 1
    status_col = headers.index("Status") + 1
    last_row = first_data_row + len(test_rows) - 1

    # Styles
    title_fill = PatternFill("solid", fgColor="1F4E79")
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    meta_fill = PatternFill("solid", fgColor="E2F0D9")
    status_fill = PatternFill("solid", fgColor="D9EAD3")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Tests"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A2"] = "Generated from DMN. Enable macros, then use Run All Tests on the Dashboard or a row Run button."
    ws["A2"].alignment = Alignment(wrap_text=True)

    for c, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, c, header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for r_idx, row in enumerate(test_rows, start=first_data_row):
        values: Dict[str, Any] = {
            "Name": row.name,
            "Method": "POST",
            "URL": row.url,
            "Expected": row.expected_text,
            "Actual": None,
            "Status": "NOT RUN",
            "Run": "Run",
            "Username": username,
            "Password": password,
            "Outcome Names": ",".join(row.output_names),
            "Variable Types": row.variable_types,
            "Generated JSON Body": row.body_json,
            "Decision ID": row.decision_id,
            "Selected Rule ID": row.rule_id,
            "Selected Rule Index": row.rule_index,
            "Coverage Reasons": row.coverage_reason,
        }
        for inp in all_inputs:
            values[inp] = row.input_values.get(inp)

        for c, header in enumerate(headers, start=1):
            cell = ws.cell(r_idx, c, excel_safe_literal(values.get(header)))
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(header in {"Expected", "Generated JSON Body", "Coverage Reasons"}))
            if header in all_inputs:
                cell.fill = input_fill
            elif header in {"Username", "Password", "Outcome Names", "Variable Types", "Generated JSON Body", "Decision ID", "Selected Rule ID", "Selected Rule Index", "Coverage Reasons"}:
                cell.fill = meta_fill
            elif header == "Status":
                cell.fill = status_fill

    # Fix URL assignment; dataclass is frozen? no. But build_test_rows currently does not save url. Derive from JSON helper below is impossible.
    # This block is superseded by setting row.url dynamically in main.

    # Layout
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(last_row, header_row)}"
    for idx, header in enumerate(headers, start=1):
        if header in {"Name", "URL", "Expected", "Generated JSON Body", "Coverage Reasons"}:
            width = 34 if header != "URL" else 54
        elif header in all_inputs:
            width = 18
        elif header in {"Username", "Password"}:
            width = 16
        else:
            width = 14
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 30
    for r in range(first_data_row, max(first_data_row, last_row) + 1):
        ws.row_dimensions[r].height = 24

    # Dashboard
    status_col_letter = get_column_letter(status_col)
    safe_last_row = max(last_row, first_data_row)
    dash["A1"] = title
    dash["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    dash["A1"].fill = title_fill
    dash["A2"] = "Use the button below to run all tests. You can edit inputs on the Tests sheet."
    dash["A4"] = "Total"
    dash["B4"] = f"=COUNTA(Tests!$A${first_data_row}:$A${safe_last_row})"
    dash["A5"] = "Passed"
    dash["B5"] = f'=COUNTIF(Tests!${status_col_letter}${first_data_row}:${status_col_letter}${safe_last_row},"PASS")'
    dash["A6"] = "Failed"
    dash["B6"] = f'=COUNTIF(Tests!${status_col_letter}${first_data_row}:${status_col_letter}${safe_last_row},"FAIL")'
    dash["A7"] = "Not run"
    dash["B7"] = f'=COUNTIF(Tests!${status_col_letter}${first_data_row}:${status_col_letter}${safe_last_row},"NOT RUN")'
    dash["D4"] = "Result"
    dash["E4"] = "Count"
    dash["D5"] = "PASS"
    dash["E5"] = "=B5"
    dash["D6"] = "FAIL"
    dash["E6"] = "=B6"
    dash["D7"] = "NOT RUN"
    dash["E7"] = "=B7"
    for cell in dash[4]:
        cell.font = Font(bold=True)
    for row in dash.iter_rows(min_row=4, max_row=7, min_col=1, max_col=5):
        for cell in row:
            cell.border = border
    for col in range(1, 6):
        dash.column_dimensions[get_column_letter(col)].width = 18

    # Mark workbook for recalc in Excel.
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    wb.save(output_path)
    return run_col



# ---------------------------------------------------------------------------
# OOXML control/chart patching
# ---------------------------------------------------------------------------


REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def read_zip_entries(path: Path) -> Dict[str, bytes]:
    entries: Dict[str, bytes] = {}
    with zipfile.ZipFile(path, "r") as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            entries[info.filename] = zf.read(info.filename)
    return entries


def write_zip_entries(path: Path, entries: Dict[str, bytes]) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name, data in entries.items():
            zf.writestr(name, data)
    tmp.replace(path)


def get_sheet_path(entries: Dict[str, bytes], sheet_name: str) -> str:
    workbook_xml = ET.fromstring(entries["xl/workbook.xml"])
    rels_xml = ET.fromstring(entries["xl/_rels/workbook.xml.rels"])
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main", "r": R_NS, "rel": REL_NS}
    rid = None
    for sheet in workbook_xml.findall(".//main:sheet", ns):
        if sheet.attrib.get("name") == sheet_name:
            rid = sheet.attrib.get(f"{{{R_NS}}}id")
            break
    if not rid:
        raise ValueError(f"Could not find sheet {sheet_name!r} in workbook.xml")
    target = None
    for rel in rels_xml.findall("rel:Relationship", ns):
        if rel.attrib.get("Id") == rid:
            target = rel.attrib.get("Target")
            break
    if not target:
        raise ValueError(f"Could not resolve relationship for sheet {sheet_name!r}")
    if target.startswith("/"):
        return target.lstrip("/")
    if target.startswith("xl/"):
        return target
    return "xl/" + target


def xml_escape_attr(value: str) -> str:
    return (
        value.replace("&", "&amp;")
        .replace('"', "&quot;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def strip_sheet_drawings_and_controls(sheet_xml: str) -> str:
    sheet_xml = re.sub(r"<drawing\b[^>]*/>", "", sheet_xml)
    sheet_xml = re.sub(r"<legacyDrawing\b[^>]*/>", "", sheet_xml)
    # Controls live inside an mc:AlternateContent wrapper. Remove any wrapper containing <controls>.
    sheet_xml = re.sub(
        r"<mc:AlternateContent\b(?:(?!</mc:AlternateContent>).)*?<controls>.*?</controls>\s*</mc:Choice>\s*</mc:AlternateContent>",
        "",
        sheet_xml,
        flags=re.S,
    )
    return sheet_xml


def dashboard_controls_xml() -> str:
    return (
        '<mc:AlternateContent xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
        'xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
        '<mc:Choice Requires="x14"><controls>'
        '<mc:AlternateContent xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006">'
        '<mc:Choice Requires="x14"><control shapeId="1025" r:id="rId4" name="Button 1">'
        '<controlPr defaultSize="0" print="0" autoFill="0" autoPict="0" macro="RunAllTests" altText="Run All Tests">'
        '<anchor moveWithCells="1" sizeWithCells="1"><from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>8</xdr:row><xdr:rowOff>0</xdr:rowOff></from>'
        '<to><xdr:col>1</xdr:col><xdr:colOff>476250</xdr:colOff><xdr:row>10</xdr:row><xdr:rowOff>38100</xdr:rowOff></to></anchor>'
        '</controlPr></control></mc:Choice></mc:AlternateContent>'
        '</controls></mc:Choice></mc:AlternateContent>'
    )


def tests_controls_xml(num_tests: int, run_col_zero: int, first_row_zero: int = 4) -> str:
    parts = [
        '<mc:AlternateContent xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
        'xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
        '<mc:Choice Requires="x14"><controls>'
    ]
    for i in range(num_tests):
        shape_id = 2049 + i
        row_zero = first_row_zero + i
        rel_id = f"rId{3 + i}"
        parts.append(
            f'<mc:AlternateContent xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006">'
            f'<mc:Choice Requires="x14"><control shapeId="{shape_id}" r:id="{rel_id}" name="Button {i + 1}">'
            f'<controlPr defaultSize="0" print="0" autoFill="0" autoPict="0" macro="RunSelectedTest" altText="Run">'
            f'<anchor moveWithCells="1" sizeWithCells="1">'
            f'<from><xdr:col>{run_col_zero}</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>{row_zero}</xdr:row><xdr:rowOff>63500</xdr:rowOff></from>'
            f'<to><xdr:col>{run_col_zero + 1}</xdr:col><xdr:colOff>6350</xdr:colOff><xdr:row>{row_zero}</xdr:row><xdr:rowOff>368300</xdr:rowOff></to>'
            f'</anchor></controlPr></control></mc:Choice></mc:AlternateContent>'
        )
    parts.append('</controls></mc:Choice></mc:AlternateContent>')
    return "".join(parts)


def make_tests_vml(num_tests: int, run_col_zero: int, first_row_zero: int = 4) -> str:
    left_pt = max(0, run_col_zero) * 95  # approximate only; anchors drive final placement in Excel
    parts = [
        '<xml xmlns:v="urn:schemas-microsoft-com:vml"\n'
        ' xmlns:o="urn:schemas-microsoft-com:office:office"\n'
        ' xmlns:x="urn:schemas-microsoft-com:office:excel">\n'
        ' <o:shapelayout v:ext="edit">\n'
        '  <o:idmap v:ext="edit" data="2"/>\n'
        ' </o:shapelayout><v:shapetype id="_x0000_t201" coordsize="21600,21600" o:spt="201"\n'
        '  path="m,l,21600r21600,l21600,xe">\n'
        '  <v:stroke joinstyle="miter"/>\n'
        '  <v:path shadowok="f" o:extrusionok="f" strokeok="f" fillok="f" o:connecttype="rect"/>\n'
        '  <o:lock v:ext="edit" shapetype="t"/>\n'
        ' </v:shapetype>'
    ]
    for i in range(num_tests):
        shape_id = 2049 + i
        row_zero = first_row_zero + i
        top_pt = 15.5 + row_zero * 18
        parts.append(
            f'<v:shape id="_x0000_s{shape_id}" type="#_x0000_t201" style=\'position:absolute;\n'
            f'  margin-left:{left_pt}pt;margin-top:{top_pt}pt;width:69.5pt;height:18pt;z-index:{i + 1};\n'
            f'  mso-wrap-style:tight\' o:button="t" fillcolor="buttonFace [67]" o:insetmode="auto">\n'
            f'  <v:fill color2="buttonFace [67]" o:detectmouseclick="t"/>\n'
            f'  <o:lock v:ext="edit" rotation="t"/>\n'
            f'  <v:textbox style=\'mso-direction-alt:auto\' o:singleclick="f">\n'
            f'   <div style=\'text-align:center\'><font face="Aptos Narrow" size="220"\n'
            f'   color="#000000">Run</font></div>\n'
            f'  </v:textbox>\n'
            f'  <x:ClientData ObjectType="Button">\n'
            f'   <x:Anchor>\n'
            f'    {run_col_zero}, 0, {row_zero}, 10, {run_col_zero + 1}, 1, {row_zero}, 58</x:Anchor>\n'
            f'   <x:PrintObject>False</x:PrintObject>\n'
            f'   <x:AutoFill>False</x:AutoFill>\n'
            f'   <x:FmlaMacro>RunSelectedTest</x:FmlaMacro>\n'
            f'   <x:TextHAlign>Center</x:TextHAlign>\n'
            f'   <x:TextVAlign>Center</x:TextVAlign>\n'
            f'  </x:ClientData>\n'
            f' </v:shape>'
        )
    parts.append('</xml>')
    return "".join(parts)


def make_tests_drawing(num_tests: int, run_col_zero: int, first_row_zero: int = 4) -> str:
    parts = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
    ]
    for i in range(num_tests):
        shape_id = 2049 + i
        row_zero = first_row_zero + i
        creation = f"{{00000000-0008-0000-{i + 1:04d}-000001080000}}"
        parts.append(
            '<mc:AlternateContent xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006">'
            '<mc:Choice xmlns:a14="http://schemas.microsoft.com/office/drawing/2010/main" Requires="a14">'
            '<xdr:twoCellAnchor>'
            f'<xdr:from><xdr:col>{run_col_zero}</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>{row_zero}</xdr:row><xdr:rowOff>63500</xdr:rowOff></xdr:from>'
            f'<xdr:to><xdr:col>{run_col_zero + 1}</xdr:col><xdr:colOff>6350</xdr:colOff><xdr:row>{row_zero}</xdr:row><xdr:rowOff>368300</xdr:rowOff></xdr:to>'
            '<xdr:sp macro="" textlink="">'
            '<xdr:nvSpPr>'
            f'<xdr:cNvPr id="{shape_id}" name="Button {i + 1}" descr="Run" hidden="1">'
            '<a:extLst>'
            '<a:ext uri="{63B3BB69-23CF-44E3-9099-C40C66FF867C}">'
            f'<a14:compatExt spid="_x0000_s{shape_id}"/>'
            '</a:ext>'
            '<a:ext uri="{FF2B5EF4-FFF2-40B4-BE49-F238E27FC236}">'
            f'<a16:creationId xmlns:a16="http://schemas.microsoft.com/office/drawing/2014/main" id="{creation}"/>'
            '</a:ext>'
            '</a:extLst>'
            '</xdr:cNvPr><xdr:cNvSpPr/></xdr:nvSpPr>'
            '<xdr:spPr bwMode="auto"><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm>'
            '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom><a:noFill/><a:ln w="9525"><a:miter lim="800000"/><a:headEnd/><a:tailEnd/></a:ln></xdr:spPr>'
            '<xdr:txBody><a:bodyPr vertOverflow="clip" wrap="square" lIns="36576" tIns="36576" rIns="36576" bIns="36576" anchor="ctr" upright="1"/><a:lstStyle/>'
            '<a:p><a:pPr algn="ctr" rtl="0"><a:defRPr sz="1000"/></a:pPr>'
            '<a:r><a:rPr lang="en-NL" sz="1100" b="0" i="0" u="none" strike="noStrike" baseline="0"><a:solidFill><a:srgbClr val="000000"/></a:solidFill><a:latin typeface="Aptos Narrow"/></a:rPr><a:t>Run</a:t></a:r>'
            '</a:p></xdr:txBody></xdr:sp><xdr:clientData fPrintsWithSheet="0"/></xdr:twoCellAnchor>'
            '</mc:Choice><mc:Fallback/></mc:AlternateContent>'
        )
    parts.append('</xdr:wsDr>')
    return "".join(parts)


def make_sheet_rels(relationships: Sequence[Tuple[str, str, str]]) -> bytes:
    parts = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n', f'<Relationships xmlns="{REL_NS}">']
    for rid, typ, target in relationships:
        parts.append(f'<Relationship Id="{xml_escape_attr(rid)}" Type="{xml_escape_attr(typ)}" Target="{xml_escape_attr(target)}"/>')
    parts.append('</Relationships>')
    return "".join(parts).encode("utf-8")


def ctrl_prop_xml() -> bytes:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<formControlPr xmlns="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" objectType="Button" lockText="1"/>'
    ).encode("utf-8")


def patch_content_types(entries: Dict[str, bytes], ctrl_prop_count: int) -> None:
    ct_name = "[Content_Types].xml"
    ns = "http://schemas.openxmlformats.org/package/2006/content-types"
    ET.register_namespace("", ns)
    root = ET.fromstring(entries[ct_name])

    def is_override_for(part: str, elem: ET.Element) -> bool:
        return elem.attrib.get("PartName") == part

    # Remove stale ctrlProp overrides and specific drawing/chart overrides we recreate.
    for child in list(root):
        if local_name(child.tag) != "Override":
            continue
        part = child.attrib.get("PartName", "")
        if part.startswith("/xl/ctrlProps/") or part in {
            "/xl/drawings/drawing1.xml",
            "/xl/drawings/drawing2.xml",
            "/xl/charts/chart1.xml",
            "/xl/vbaProject.bin",
        }:
            root.remove(child)

    def add_default(ext: str, content_type: str) -> None:
        for child in root:
            if local_name(child.tag) == "Default" and child.attrib.get("Extension") == ext:
                child.set("ContentType", content_type)
                return
        ET.SubElement(root, f"{{{ns}}}Default", Extension=ext, ContentType=content_type)

    def add_override(part: str, content_type: str) -> None:
        for child in root:
            if local_name(child.tag) == "Override" and is_override_for(part, child):
                child.set("ContentType", content_type)
                return
        ET.SubElement(root, f"{{{ns}}}Override", PartName=part, ContentType=content_type)

    add_default("vml", "application/vnd.openxmlformats-officedocument.vmlDrawing")
    add_override("/xl/drawings/drawing1.xml", "application/vnd.openxmlformats-officedocument.drawing+xml")
    add_override("/xl/drawings/drawing2.xml", "application/vnd.openxmlformats-officedocument.drawing+xml")
    add_override("/xl/charts/chart1.xml", "application/vnd.openxmlformats-officedocument.drawingml.chart+xml")
    add_override("/xl/vbaProject.bin", "application/vnd.ms-office.vbaProject")
    for i in range(1, ctrl_prop_count + 1):
        add_override(f"/xl/ctrlProps/ctrlProp{i}.xml", "application/vnd.ms-excel.controlproperties+xml")

    entries[ct_name] = ET.tostring(root, encoding="utf-8", xml_declaration=True)


def patch_chart_cache(chart_xml: bytes, total_tests: int) -> bytes:
    text = chart_xml.decode("utf-8", errors="ignore")
    # Update cached values to 0,0,total while keeping chart formulas linked to Dashboard D/E.
    text = re.sub(
        r'(<c:numCache>\s*<c:formatCode>General</c:formatCode>\s*<c:ptCount val="3"/>\s*)'
        r'<c:pt idx="0"><c:v>.*?</c:v></c:pt>\s*'
        r'<c:pt idx="1"><c:v>.*?</c:v></c:pt>\s*'
        r'<c:pt idx="2"><c:v>.*?</c:v></c:pt>',
        rf'\1<c:pt idx="0"><c:v>0</c:v></c:pt><c:pt idx="1"><c:v>0</c:v></c:pt><c:pt idx="2"><c:v>{total_tests}</c:v></c:pt>',
        text,
        flags=re.S,
    )
    return text.encode("utf-8")


def patch_ooxml_controls_and_chart(
    output_path: Path,
    template_path: Path,
    num_tests: int,
    run_col_one: int,
) -> None:
    entries = read_zip_entries(output_path)
    template_entries = read_zip_entries(template_path)

    dashboard_sheet = get_sheet_path(entries, "Dashboard")
    tests_sheet = get_sheet_path(entries, "Tests")
    dashboard_rels = str(Path(dashboard_sheet).parent / "_rels" / (Path(dashboard_sheet).name + ".rels")).replace("\\", "/")
    tests_rels = str(Path(tests_sheet).parent / "_rels" / (Path(tests_sheet).name + ".rels")).replace("\\", "/")

    run_col_zero = run_col_one - 1

    # Restore/copy dashboard chart, dashboard button drawings, and chart relationship from the known-good template.
    for name in [
        "xl/drawings/drawing1.xml",
        "xl/drawings/vmlDrawing1.vml",
        "xl/drawings/_rels/drawing1.xml.rels",
        "xl/charts/chart1.xml",
        "xl/ctrlProps/ctrlProp1.xml",
    ]:
        if name in template_entries:
            entries[name] = template_entries[name]
    if "xl/charts/chart1.xml" in entries:
        entries["xl/charts/chart1.xml"] = patch_chart_cache(entries["xl/charts/chart1.xml"], num_tests)

    # Build tests row buttons.
    entries["xl/drawings/vmlDrawing2.vml"] = make_tests_vml(num_tests, run_col_zero).encode("utf-8")
    entries["xl/drawings/drawing2.xml"] = make_tests_drawing(num_tests, run_col_zero).encode("utf-8")

    # Control properties: ctrlProp1 = dashboard; ctrlProp2.. = row buttons.
    for name in list(entries):
        if re.fullmatch(r"xl/ctrlProps/ctrlProp\d+\.xml", name):
            del entries[name]
    entries["xl/ctrlProps/ctrlProp1.xml"] = template_entries.get("xl/ctrlProps/ctrlProp1.xml", ctrl_prop_xml())
    for i in range(2, num_tests + 2):
        entries[f"xl/ctrlProps/ctrlProp{i}.xml"] = ctrl_prop_xml()

    # Patch sheet XML to reference drawings/VML/controls.
    dash_xml = entries[dashboard_sheet].decode("utf-8", errors="ignore")
    dash_xml = strip_sheet_drawings_and_controls(dash_xml)
    dash_insert = (
        '<drawing xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId2"/>'
        '<legacyDrawing xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId3"/>'
        + dashboard_controls_xml()
    )
    dash_xml = dash_xml.replace("</worksheet>", dash_insert + "</worksheet>")
    entries[dashboard_sheet] = dash_xml.encode("utf-8")

    tests_xml = entries[tests_sheet].decode("utf-8", errors="ignore")
    tests_xml = strip_sheet_drawings_and_controls(tests_xml)
    tests_insert = (
        '<drawing xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId1"/>'
        '<legacyDrawing xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId2"/>'
        + tests_controls_xml(num_tests, run_col_zero)
    )
    tests_xml = tests_xml.replace("</worksheet>", tests_insert + "</worksheet>")
    entries[tests_sheet] = tests_xml.encode("utf-8")

    # Patch relationships.
    entries[dashboard_rels] = make_sheet_rels(
        [
            ("rId2", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing", "../drawings/drawing1.xml"),
            ("rId3", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing", "../drawings/vmlDrawing1.vml"),
            ("rId4", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/ctrlProp", "../ctrlProps/ctrlProp1.xml"),
        ]
    )
    test_relationships = [
        ("rId1", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing", "../drawings/drawing2.xml"),
        ("rId2", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing", "../drawings/vmlDrawing2.vml"),
    ]
    for i in range(num_tests):
        test_relationships.append(
            (f"rId{3 + i}", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/ctrlProp", f"../ctrlProps/ctrlProp{i + 2}.xml")
        )
    entries[tests_rels] = make_sheet_rels(test_relationships)

    patch_content_types(entries, ctrl_prop_count=num_tests + 1)
    write_zip_entries(output_path, entries)


# ---------------------------------------------------------------------------
# Main CLI
# ---------------------------------------------------------------------------



def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a Pretty macro-enabled Excel test runner from a DMN file.")
    parser.add_argument("dmn", type=Path, help="Path to the input .dmn XML file")
    parser.add_argument("--template", type=Path, required=True, help="Path to the working Pretty .xlsm template")
    parser.add_argument("--output", type=Path, required=True, help="Output .xlsm file path")
    parser.add_argument("--base-url", default="https://operaton.open-regels.nl/engine-rest", help="Operaton/Camunda REST engine base URL")
    parser.add_argument("--tenant-id", default="", help="Tenant id to include in the evaluate URL. Omit for no tenant segment.")
    parser.add_argument("--username", default="", help="Username to write into the workbook")
    parser.add_argument("--password", default="", help="Password to write into the workbook")
    parser.add_argument("--title", default="DMN Test Runner", help="Dashboard title")
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    if not args.dmn.exists():
        raise FileNotFoundError(args.dmn)
    if not args.template.exists():
        raise FileNotFoundError(args.template)
    if args.output.suffix.lower() != ".xlsm":
        raise ValueError("Output must be an .xlsm file so the VBA macros are preserved.")

    decisions = parse_dmn(args.dmn)
    test_rows, all_inputs = build_test_rows(decisions, args.base_url, args.tenant_id)
    if not test_rows:
        raise ValueError("The DMN has decision tables but no rules to convert into test rows.")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    run_col = write_workbook_cells(
        template_path=args.template,
        output_path=args.output,
        test_rows=test_rows,
        all_inputs=all_inputs,
        username=args.username,
        password=args.password,
        title=args.title,
    )
    patch_ooxml_controls_and_chart(
        output_path=args.output,
        template_path=args.template,
        num_tests=len(test_rows),
        run_col_one=run_col,
    )

    print(f"Generated {args.output} with {len(test_rows)} tests from {len(decisions)} DMN decision table(s).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
